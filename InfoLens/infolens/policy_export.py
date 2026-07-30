"""雪花出库政策核销明细 Excel 导出。"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins


EXPORT_HEADERS = (
    "序号",
    "对象编码",
    "对象名称",
    "产品",
    "数量",
    "单价",
    "核销金额",
    "是否达标",
)


def _excel_safe_text(value: Any) -> str:
    """避免外部文本被 Excel 当作公式执行。"""

    text = str(value or "")
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def build_policy_reimbursement_workbook(
    policy: dict[str, Any],
    rows: list[dict[str, Any]],
    *,
    operator_name: str,
    distributor_name: str,
) -> io.BytesIO:
    """按照业务部核销模板生成可下载的工作簿。"""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.sheet_view.showGridLines = False

    title = (
        f"经销商：{distributor_name}"
        f"                              "
        f"{int(policy['year'])}.{int(policy['month'])}月（{policy['name']}）"
    )
    sheet.merge_cells("A1:H1")
    sheet["A1"] = _excel_safe_text(title)
    sheet["A1"].font = Font(name="宋体", size=12, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 25

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    base_font = Font(name="宋体", size=10)

    for column, header in enumerate(EXPORT_HEADERS, start=1):
        cell = sheet.cell(row=2, column=column, value=header)
        cell.font = Font(name="宋体", size=10, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[2].height = 25

    first_data_row = 3
    for index, item in enumerate(rows, start=1):
        row_number = first_data_row + index - 1
        values = (
            index,
            _excel_safe_text(item["terminal_code"]),
            _excel_safe_text(item["customer_name"]),
            _excel_safe_text(item["product_name"]),
            item["quantity"],
            item["settlement_price"],
            (
                f"=E{row_number}*F{row_number}"
                if item["settlement_price"] is not None
                else None
            ),
            "是",
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.font = base_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if column in {3, 4} else "center",
                vertical="center",
                wrap_text=column in {3, 4},
            )
        sheet.cell(row=row_number, column=2).number_format = "@"
        sheet.cell(row=row_number, column=5).number_format = "0.##"
        sheet.cell(row=row_number, column=6).number_format = "0.00"
        sheet.cell(row=row_number, column=7).number_format = "0.00"
        sheet.row_dimensions[row_number].height = 25

    last_data_row = first_data_row + len(rows) - 1
    total_row = last_data_row + 1
    sheet.cell(row=total_row, column=4, value="合计")
    sheet.cell(
        row=total_row,
        column=5,
        value=f"=SUM(E{first_data_row}:E{last_data_row})",
    )
    sheet.cell(
        row=total_row,
        column=7,
        value=f"=SUM(G{first_data_row}:G{last_data_row})",
    )
    for column in range(1, 9):
        cell = sheet.cell(row=total_row, column=column)
        cell.font = Font(name="宋体", size=10, bold=column in {4, 5, 7})
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=total_row, column=5).number_format = "0.##"
    sheet.cell(row=total_row, column=7).number_format = "0.00"
    sheet.row_dimensions[total_row].height = 25

    signature_row = total_row + 1
    sheet.merge_cells(
        start_row=signature_row,
        start_column=1,
        end_row=signature_row,
        end_column=3,
    )
    sheet.merge_cells(
        start_row=signature_row,
        start_column=4,
        end_row=signature_row,
        end_column=8,
    )
    sheet.cell(
        row=signature_row,
        column=1,
        value=_excel_safe_text(f"制表人：{operator_name}"),
    )
    sheet.cell(row=signature_row, column=4, value="业务部经理签字：")
    for column in range(1, 9):
        cell = sheet.cell(row=signature_row, column=column)
        cell.font = base_font
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[signature_row].height = 36

    widths = (6.375, 14.5, 41, 67.625, 9, 9, 9, 18.5)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + column)].width = width

    sheet.freeze_panes = "A3"
    sheet.print_title_rows = "1:2"
    sheet.print_area = f"A1:H{signature_row}"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.scale = 59
    sheet.sheet_properties.pageSetUpPr.fitToPage = False
    sheet.page_margins = PageMargins(
        left=0.0784722,
        right=0.0388889,
        top=0.1180556,
        bottom=0.0388889,
        header=0.0388889,
        footer=0.2986111,
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.properties.creator = operator_name
    workbook.properties.title = f"{policy['display_name']}核销明细"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
