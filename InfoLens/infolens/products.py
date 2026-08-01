"""产品档案、编码映射和雪花库存文件合并更新。"""

from __future__ import annotations

import json
import math
import re
import sqlite3
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


PRODUCT_STATUSES = {"正常", "待完善"}
AUXILIARY_UNITS = {"瓶", "听", "罐"}
STOCK_REQUIRED_HEADERS = (
    "年月",
    "商品编号",
    "商品名称",
    "单位",
    "入库千升数",
    "可用（箱）",
)
PACK_PATTERN = re.compile(r"(\d+)\s*[*×xX]\s*(\d+)")


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _clean(value: Any, *, label: str, limit: int, required: bool = False) -> str:
    text = _text(value)
    if required and not text:
        raise ValueError(f"{label}不能为空")
    if len(text) > limit:
        raise ValueError(f"{label}不能超过{limit}个字符")
    return text


def _codes(value: Any, *, label: str, required: bool) -> list[str]:
    if isinstance(value, str):
        values = re.split(r"[,，;；\n]+", value)
    elif isinstance(value, (list, tuple)):
        values = value
    else:
        values = []
    result: list[str] = []
    seen: set[str] = set()
    for raw in values:
        code = _clean(raw, label=label, limit=100)
        if not code or code in seen:
            continue
        seen.add(code)
        result.append(code)
    if required and not result:
        raise ValueError(f"{label}不能为空")
    if len(result) > 50:
        raise ValueError(f"{label}最多支持50个")
    return result


def _nonnegative_number(
    value: Any,
    *,
    label: str,
    integer: bool = False,
    required: bool = False,
) -> int | float | None:
    if value is None or _text(value) == "":
        if required:
            raise ValueError(f"{label}不能为空")
        return None
    if isinstance(value, bool):
        raise ValueError(f"{label}格式不正确")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label}必须为数值") from exc
    if not math.isfinite(number) or number < 0:
        raise ValueError(f"{label}必须为非负数")
    if integer:
        if not number.is_integer():
            raise ValueError(f"{label}必须为非负整数")
        return int(number)
    return number


def _number_or_zero(value: Any, *, label: str) -> float:
    if value is None or _text(value) == "":
        return 0.0
    if isinstance(value, bool):
        raise ValueError(f"{label}格式不正确")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label}必须为数值") from exc
    if not math.isfinite(number):
        raise ValueError(f"{label}必须为有限数值")
    return number


def _month(value: Any) -> str:
    digits = re.sub(r"\D", "", _text(value))
    if len(digits) < 6:
        raise ValueError("年月格式不正确")
    month = digits[:6]
    if not 1 <= int(month[4:6]) <= 12:
        raise ValueError("年月格式不正确")
    return month


def parse_packaging(product_name: str) -> tuple[int | None, str]:
    """从商品名称解析每箱规格及瓶/听/罐辅助单位。"""
    match = PACK_PATTERN.search(str(product_name or ""))
    if not match:
        return None, ""
    left = int(match.group(1))
    right = int(match.group(2))
    specification = left * right
    prefix = str(product_name)[: match.start()]
    positions = [(prefix.rfind(unit), unit) for unit in AUXILIARY_UNITS]
    position, auxiliary_unit = max(positions, default=(-1, ""))
    return specification, auxiliary_unit if position >= 0 else ""


def normalize_product(
    payload: dict[str, Any],
    *,
    allow_incomplete: bool = False,
) -> dict[str, Any]:
    product = {
        "product_codes": _codes(
            payload.get("product_codes"),
            label="商品编码",
            required=True,
        ),
        "short_name": _clean(
            payload.get("short_name"),
            label="商品简称",
            limit=100,
            required=not allow_incomplete,
        ),
        "product_name": _clean(
            payload.get("product_name"),
            label="商品名称",
            limit=500,
            required=True,
        ),
        "snow_inventory": _nonnegative_number(
            payload.get("snow_inventory"),
            label="雪花库存",
            required=True,
        ),
        "housekeeper_codes": _codes(
            payload.get("housekeeper_codes"),
            label="管家婆编码",
            required=not allow_incomplete,
        ),
        "specification": _nonnegative_number(
            payload.get("specification"),
            label="规格",
            integer=True,
            required=not allow_incomplete,
        ),
        "auxiliary_unit": _clean(
            payload.get("auxiliary_unit"),
            label="辅助单位",
            limit=10,
            required=not allow_incomplete,
        ),
        "settlement_price": _nonnegative_number(
            payload.get("settlement_price"),
            label="结算价",
        ),
    }
    if product["settlement_price"] is not None:
        product["settlement_price"] = round(product["settlement_price"], 2)
    if (
        product["auxiliary_unit"]
        and product["auxiliary_unit"] not in AUXILIARY_UNITS
    ):
        raise ValueError("辅助单位只能是瓶、听或罐")
    complete = (
        bool(product["short_name"])
        and bool(product["housekeeper_codes"])
        and product["specification"] is not None
        and bool(product["auxiliary_unit"])
    )
    product["status"] = "正常" if complete else "待完善"
    return product


def parse_stock_workbook(stream: BinaryIO) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError, TypeError) as exc:
        raise ValueError("无法读取Excel文件，请确认文件格式正确") from exc
    try:
        worksheet = workbook.active
        # 部分雪花系统导出的工作簿把 dimension 错写成 A1，
        # read_only 模式会因此只读取第一列，需要让 openpyxl 重新扫描实际范围。
        if worksheet.calculate_dimension() in {"A1", "A1:A1"}:
            worksheet.reset_dimensions()
        rows = worksheet.iter_rows(values_only=True)
        header_values = next(rows, None)
        if not header_values:
            raise ValueError("Excel文件没有表头")
        headers = [_text(value) for value in header_values]
        missing = [name for name in STOCK_REQUIRED_HEADERS if name not in headers]
        if missing:
            raise ValueError(f"Excel缺少必要字段：{'、'.join(missing)}")
        indexes = {name: headers.index(name) for name in STOCK_REQUIRED_HEADERS}
        parsed: list[dict[str, Any]] = []
        for row_number, values in enumerate(rows, start=2):
            if not any(value is not None and _text(value) for value in values):
                continue
            raw = {
                header: values[index] if index < len(values) else None
                for header, index in indexes.items()
            }
            product_code = _clean(
                raw["商品编号"],
                label="商品编号",
                limit=100,
                required=True,
            )
            product_name = _clean(
                raw["商品名称"],
                label="商品名称",
                limit=500,
                required=True,
            )
            source_unit = _text(raw["单位"])
            specification, auxiliary_unit = parse_packaging(product_name)
            parsed.append(
                {
                    "row_number": row_number,
                    "source_month": _month(raw["年月"]),
                    "product_code": product_code,
                    "product_name": product_name,
                    "source_unit": source_unit,
                    "inbound_kiloliters": _number_or_zero(
                        raw["入库千升数"],
                        label="入库千升数",
                    ),
                    "snow_inventory": (
                        _nonnegative_number(
                            raw["可用（箱）"],
                            label="可用（箱）",
                            required=True,
                        )
                        if source_unit == "箱"
                        else None
                    ),
                    "specification": specification,
                    "auxiliary_unit": auxiliary_unit,
                }
            )
        return parsed
    finally:
        workbook.close()


class ProductStore:
    def __init__(self, database_path: str | Path):
        self.database_path = Path(database_path)
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self._initialize()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.database_path, timeout=30)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys=ON")
        connection.execute("PRAGMA journal_mode=WAL")
        connection.execute("PRAGMA busy_timeout=30000")
        return connection

    def _initialize(self) -> None:
        with self._connect() as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS products (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    short_name TEXT NOT NULL DEFAULT '',
                    product_name TEXT NOT NULL,
                    snow_inventory REAL NOT NULL DEFAULT 0,
                    specification INTEGER,
                    auxiliary_unit TEXT NOT NULL DEFAULT '',
                    settlement_price REAL,
                    status TEXT NOT NULL DEFAULT '待完善',
                    stock_month TEXT NOT NULL DEFAULT '',
                    version INTEGER NOT NULL DEFAULT 1,
                    created_by TEXT NOT NULL,
                    created_by_name TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    updated_by TEXT NOT NULL,
                    updated_by_name TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    deleted_by TEXT NOT NULL DEFAULT '',
                    deleted_by_name TEXT NOT NULL DEFAULT '',
                    deleted_at TEXT NOT NULL DEFAULT '',
                    CHECK(status IN ('正常', '待完善')),
                    CHECK(snow_inventory >= 0),
                    CHECK(specification IS NULL OR specification >= 0),
                    CHECK(settlement_price IS NULL OR settlement_price >= 0)
                );
                CREATE INDEX IF NOT EXISTS idx_products_filter
                    ON products(deleted_at, status, updated_at DESC, id DESC);
                CREATE INDEX IF NOT EXISTS idx_products_name
                    ON products(product_name, short_name);
                CREATE TABLE IF NOT EXISTS product_codes (
                    product_id INTEGER NOT NULL,
                    code TEXT NOT NULL UNIQUE,
                    PRIMARY KEY(product_id, code),
                    FOREIGN KEY(product_id) REFERENCES products(id)
                );
                CREATE INDEX IF NOT EXISTS idx_product_codes_product
                    ON product_codes(product_id);

                CREATE TABLE IF NOT EXISTS product_housekeeper_codes (
                    product_id INTEGER NOT NULL,
                    code TEXT NOT NULL UNIQUE,
                    PRIMARY KEY(product_id, code),
                    FOREIGN KEY(product_id) REFERENCES products(id)
                );
                CREATE INDEX IF NOT EXISTS idx_product_housekeeper_product
                    ON product_housekeeper_codes(product_id);

                CREATE TABLE IF NOT EXISTS product_audit_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER NOT NULL,
                    operator TEXT NOT NULL,
                    operator_name TEXT NOT NULL,
                    action_type TEXT NOT NULL,
                    action_summary TEXT NOT NULL,
                    changes_json TEXT NOT NULL DEFAULT '{}',
                    operated_at TEXT NOT NULL,
                    FOREIGN KEY(product_id) REFERENCES products(id)
                );

                CREATE TABLE IF NOT EXISTS product_import_previews (
                    id TEXT PRIMARY KEY,
                    operator TEXT NOT NULL,
                    operator_name TEXT NOT NULL,
                    filename TEXT NOT NULL,
                    rows_json TEXT NOT NULL,
                    result_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    committed_at TEXT NOT NULL DEFAULT ''
                );

                CREATE TABLE IF NOT EXISTS product_imports (
                    id TEXT PRIMARY KEY,
                    operator TEXT NOT NULL,
                    operator_name TEXT NOT NULL,
                    filename TEXT NOT NULL,
                    total_count INTEGER NOT NULL,
                    created_count INTEGER NOT NULL,
                    updated_count INTEGER NOT NULL,
                    unchanged_count INTEGER NOT NULL,
                    skipped_count INTEGER NOT NULL,
                    warning_count INTEGER NOT NULL,
                    failed_count INTEGER NOT NULL,
                    monthly_inbound_tons REAL NOT NULL DEFAULT 0,
                    snow_inventory_boxes REAL NOT NULL DEFAULT 0,
                    committed_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS product_import_monthly_summaries (
                    import_id TEXT NOT NULL,
                    month TEXT NOT NULL,
                    inbound_tons REAL NOT NULL DEFAULT 0,
                    inventory_boxes REAL NOT NULL DEFAULT 0,
                    PRIMARY KEY(import_id, month),
                    FOREIGN KEY(import_id) REFERENCES product_imports(id)
                );
                CREATE INDEX IF NOT EXISTS idx_product_import_monthly_summary
                    ON product_import_monthly_summaries(month, import_id);
                """
            )
            import_columns = {
                row["name"]
                for row in connection.execute("PRAGMA table_info(product_imports)")
            }
            if "monthly_inbound_tons" not in import_columns:
                connection.execute(
                    """
                    ALTER TABLE product_imports
                    ADD COLUMN monthly_inbound_tons REAL NOT NULL DEFAULT 0
                    """
                )
            if "snow_inventory_boxes" not in import_columns:
                connection.execute(
                    """
                    ALTER TABLE product_imports
                    ADD COLUMN snow_inventory_boxes REAL NOT NULL DEFAULT 0
                    """
                )
            product_columns = {
                row["name"]
                for row in connection.execute("PRAGMA table_info(products)")
            }
            if "stock_month" not in product_columns:
                connection.execute(
                    """
                    ALTER TABLE products
                    ADD COLUMN stock_month TEXT NOT NULL DEFAULT ''
                    """
                )
            connection.execute(
                """
                UPDATE products
                SET stock_month = ?
                WHERE stock_month = ''
                """,
                (datetime.now().strftime("%Y%m"),),
            )
            connection.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_products_stock_month
                ON products(stock_month, deleted_at, updated_at DESC)
                """
            )
            self._backfill_monthly_summaries(connection)

    def _backfill_monthly_summaries(
        self,
        connection: sqlite3.Connection,
    ) -> None:
        imports = connection.execute(
            """
            SELECT product_imports.id, product_import_previews.rows_json
            FROM product_imports
            JOIN product_import_previews
              ON product_import_previews.committed_at = product_imports.committed_at
             AND product_import_previews.filename = product_imports.filename
             AND product_import_previews.operator = product_imports.operator
            WHERE NOT EXISTS (
                SELECT 1
                FROM product_import_monthly_summaries
                WHERE product_import_monthly_summaries.import_id =
                      product_imports.id
            )
            """
        ).fetchall()
        for imported in imports:
            try:
                rows = json.loads(imported["rows_json"])
            except (TypeError, ValueError):
                continue
            for summary in self._monthly_summaries(rows):
                connection.execute(
                    """
                    INSERT OR IGNORE INTO product_import_monthly_summaries (
                        import_id, month, inbound_tons, inventory_boxes
                    ) VALUES (?, ?, ?, ?)
                    """,
                    (
                        imported["id"],
                        summary["month"],
                        summary["inbound_tons"],
                        summary["inventory_boxes"],
                    ),
                )

    def list_products(
        self,
        *,
        name: str = "",
        product_code: str = "",
        housekeeper_code: str = "",
        status: str = "",
        inventory_sort: str = "",
        summary_month: str = "",
        page: int = 1,
        page_size: int = 20,
    ) -> dict[str, Any]:
        page = max(1, int(page))
        page_size = max(1, min(100, int(page_size)))
        current_month = datetime.now().strftime("%Y%m")
        conditions = [
            "products.deleted_at = ''",
            "products.stock_month = ?",
        ]
        parameters: list[Any] = [current_month]
        if name.strip():
            conditions.append(
                "(products.product_name LIKE ? OR products.short_name LIKE ?)"
            )
            parameters.extend([f"%{name.strip()}%", f"%{name.strip()}%"])
        if product_code.strip():
            conditions.append(
                """
                EXISTS (
                    SELECT 1 FROM product_codes
                    WHERE product_codes.product_id = products.id
                      AND product_codes.code LIKE ?
                )
                """
            )
            parameters.append(f"%{product_code.strip()}%")
        if housekeeper_code.strip():
            conditions.append(
                """
                EXISTS (
                    SELECT 1 FROM product_housekeeper_codes
                    WHERE product_housekeeper_codes.product_id = products.id
                      AND product_housekeeper_codes.code LIKE ?
                )
                """
            )
            parameters.append(f"%{housekeeper_code.strip()}%")
        if status.strip():
            if status not in PRODUCT_STATUSES:
                raise ValueError("档案状态格式不正确")
            conditions.append("products.status = ?")
            parameters.append(status)
        inventory_sort = inventory_sort.strip().lower()
        if inventory_sort not in {"", "asc", "desc"}:
            raise ValueError("库存排序格式不正确")
        summary_month = summary_month.strip()
        if (
            summary_month
            and summary_month != "all"
            and not re.fullmatch(r"\d{6}", summary_month)
        ):
            raise ValueError("汇总月份格式不正确")
        order_by = (
            f"products.snow_inventory {inventory_sort.upper()}, "
            "products.updated_at DESC, products.id DESC"
            if inventory_sort
            else "products.updated_at DESC, products.id DESC"
        )
        where = " AND ".join(conditions)
        with self._connect() as connection:
            total = connection.execute(
                f"SELECT COUNT(*) FROM products WHERE {where}",
                parameters,
            ).fetchone()[0]
            rows = connection.execute(
                f"""
                SELECT products.*
                FROM products
                WHERE {where}
                ORDER BY {order_by}
                LIMIT ? OFFSET ?
                """,
                [*parameters, page_size, (page - 1) * page_size],
            ).fetchall()
            items = [self._public_product(connection, row) for row in rows]
            summary_months = [
                row["month"]
                for row in connection.execute(
                    """
                    SELECT DISTINCT month
                    FROM product_import_monthly_summaries
                    ORDER BY month DESC
                    """
                ).fetchall()
            ]
            selected_month = (
                summary_month
                if summary_month
                else current_month
                if current_month in summary_months
                else summary_months[0]
                if summary_months
                else current_month
            )
            if selected_month == "all":
                inbound_summary = connection.execute(
                    """
                    WITH ranked AS (
                        SELECT product_import_monthly_summaries.month,
                               product_import_monthly_summaries.inbound_tons,
                               product_imports.committed_at,
                               ROW_NUMBER() OVER (
                                   PARTITION BY
                                       product_import_monthly_summaries.month
                                   ORDER BY product_imports.committed_at DESC,
                                            product_imports.rowid DESC
                               ) AS rank_number
                        FROM product_import_monthly_summaries
                        JOIN product_imports
                          ON product_imports.id =
                             product_import_monthly_summaries.import_id
                    )
                    SELECT COALESCE(SUM(inbound_tons), 0) AS inbound_tons,
                           MAX(committed_at) AS committed_at
                    FROM ranked
                    WHERE rank_number = 1
                    """
                ).fetchone()
            else:
                inbound_summary = connection.execute(
                    """
                    SELECT product_imports.committed_at,
                           product_import_monthly_summaries.inbound_tons
                    FROM product_import_monthly_summaries
                    JOIN product_imports
                      ON product_imports.id =
                         product_import_monthly_summaries.import_id
                    WHERE product_import_monthly_summaries.month = ?
                    ORDER BY product_imports.committed_at DESC,
                             product_imports.rowid DESC
                    LIMIT 1
                    """,
                    (selected_month,),
                ).fetchone()
            current_inventory = connection.execute(
                """
                SELECT product_import_monthly_summaries.inventory_boxes
                FROM product_import_monthly_summaries
                JOIN product_imports
                  ON product_imports.id =
                     product_import_monthly_summaries.import_id
                WHERE product_import_monthly_summaries.month = ?
                ORDER BY product_imports.committed_at DESC,
                         product_imports.rowid DESC
                LIMIT 1
                """,
                (current_month,),
            ).fetchone()
            if not summary_months:
                legacy_import = connection.execute(
                    """
                    SELECT committed_at,
                           monthly_inbound_tons AS inbound_tons,
                           snow_inventory_boxes AS inventory_boxes
                    FROM product_imports
                    ORDER BY committed_at DESC, rowid DESC
                    LIMIT 1
                    """
                ).fetchone()
                if legacy_import:
                    summary_months = [current_month]
                    inbound_summary = legacy_import
                    current_inventory = legacy_import
        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "summary_month": selected_month,
            "summary_months": summary_months,
            "latest_upload_at": (
                inbound_summary["committed_at"]
                if inbound_summary and inbound_summary["committed_at"]
                else ""
            ),
            "monthly_inbound_tons": (
                float(inbound_summary["inbound_tons"])
                if inbound_summary
                else 0.0
            ),
            "snow_inventory_boxes": (
                float(current_inventory["inventory_boxes"])
                if current_inventory
                else 0.0
            ),
        }

    def get_product(self, product_id: int) -> dict[str, Any]:
        with self._connect() as connection:
            row = connection.execute(
                "SELECT * FROM products WHERE id = ? AND deleted_at = ''",
                (product_id,),
            ).fetchone()
            if not row:
                raise ValueError("产品档案不存在")
            return self._public_product(connection, row)

    def normal_product_options(self) -> list[dict[str, Any]]:
        current_month = datetime.now().strftime("%Y%m")
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT *
                FROM products
                WHERE deleted_at = '' AND status = '正常'
                  AND stock_month = ?
                ORDER BY short_name, product_name, id
                """,
                (current_month,),
            ).fetchall()
            return [
                {
                    "id": int(row["id"]),
                    "short_name": row["short_name"],
                    "product_name": row["product_name"],
                    "product_codes": [
                        code["code"]
                        for code in connection.execute(
                            """
                            SELECT code FROM product_codes
                            WHERE product_id = ?
                            ORDER BY code
                            """,
                            (row["id"],),
                        ).fetchall()
                    ],
                }
                for row in rows
            ]

    def create_product(
        self,
        payload: dict[str, Any],
        *,
        operator: str,
        operator_name: str,
        source: str = "manual",
        allow_incomplete: bool = False,
    ) -> dict[str, Any]:
        product = normalize_product(payload, allow_incomplete=allow_incomplete)
        now = _now()
        with self._connect() as connection:
            try:
                cursor = connection.execute(
                    """
                    INSERT INTO products (
                        short_name, product_name, snow_inventory, specification,
                        auxiliary_unit, settlement_price, status, stock_month,
                        created_by, created_by_name, created_at,
                        updated_by, updated_by_name, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        product["short_name"],
                        product["product_name"],
                        product["snow_inventory"],
                        product["specification"],
                        product["auxiliary_unit"],
                        product["settlement_price"],
                        product["status"],
                        datetime.now().strftime("%Y%m"),
                        operator,
                        operator_name,
                        now,
                        operator,
                        operator_name,
                        now,
                    ),
                )
                product_id = int(cursor.lastrowid)
                self._replace_codes(
                    connection,
                    product_id,
                    product["product_codes"],
                    product["housekeeper_codes"],
                )
                self._log(
                    connection,
                    product_id,
                    operator,
                    operator_name,
                    "create",
                    "上传创建待完善产品" if source == "upload" else "新建产品档案",
                    product,
                )
            except sqlite3.IntegrityError as exc:
                raise self._integrity_error(exc) from exc
            row = connection.execute(
                "SELECT * FROM products WHERE id = ?", (product_id,)
            ).fetchone()
            return self._public_product(connection, row)

    def update_product(
        self,
        product_id: int,
        payload: dict[str, Any],
        *,
        operator: str,
        operator_name: str,
        allow_incomplete: bool = False,
        source: str = "manual",
    ) -> dict[str, Any]:
        product = normalize_product(payload, allow_incomplete=allow_incomplete)
        with self._connect() as connection:
            current = connection.execute(
                "SELECT * FROM products WHERE id = ? AND deleted_at = ''",
                (product_id,),
            ).fetchone()
            if not current:
                raise ValueError("产品档案不存在")
            expected_version = payload.get("version")
            if expected_version is not None and int(expected_version) != current["version"]:
                raise ValueError("产品档案已被其他人修改，请刷新后重试")
            now = _now()
            try:
                connection.execute(
                    """
                    UPDATE products
                    SET short_name = ?, product_name = ?, snow_inventory = ?,
                        specification = ?, auxiliary_unit = ?,
                        settlement_price = ?, status = ?, version = version + 1,
                        updated_by = ?, updated_by_name = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        product["short_name"],
                        product["product_name"],
                        product["snow_inventory"],
                        product["specification"],
                        product["auxiliary_unit"],
                        product["settlement_price"],
                        product["status"],
                        operator,
                        operator_name,
                        now,
                        product_id,
                    ),
                )
                self._replace_codes(
                    connection,
                    product_id,
                    product["product_codes"],
                    product["housekeeper_codes"],
                )
                self._log(
                    connection,
                    product_id,
                    operator,
                    operator_name,
                    "update",
                    "上传更新产品档案" if source == "upload" else "修改产品档案",
                    product,
                )
            except sqlite3.IntegrityError as exc:
                raise self._integrity_error(exc) from exc
            row = connection.execute(
                "SELECT * FROM products WHERE id = ?", (product_id,)
            ).fetchone()
            return self._public_product(connection, row)

    def delete_product(
        self,
        product_id: int,
        *,
        operator: str,
        operator_name: str,
    ) -> None:
        with self._connect() as connection:
            row = connection.execute(
                "SELECT * FROM products WHERE id = ? AND deleted_at = ''",
                (product_id,),
            ).fetchone()
            if not row:
                raise ValueError("产品档案不存在")
            snapshot = self._public_product(connection, row)
            now = _now()
            connection.execute(
                """
                UPDATE products
                SET deleted_by = ?, deleted_by_name = ?, deleted_at = ?,
                    updated_by = ?, updated_by_name = ?, updated_at = ?,
                    version = version + 1
                WHERE id = ?
                """,
                (
                    operator,
                    operator_name,
                    now,
                    operator,
                    operator_name,
                    now,
                    product_id,
                ),
            )
            connection.execute(
                "DELETE FROM product_codes WHERE product_id = ?", (product_id,)
            )
            connection.execute(
                "DELETE FROM product_housekeeper_codes WHERE product_id = ?",
                (product_id,),
            )
            self._log(
                connection,
                product_id,
                operator,
                operator_name,
                "delete",
                "删除产品档案",
                snapshot,
            )

    def create_import_preview(
        self,
        *,
        filename: str,
        operator: str,
        operator_name: str,
        rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        preview_id = uuid.uuid4().hex
        now = _now()
        with self._connect() as connection:
            connection.execute(
                "DELETE FROM product_import_previews WHERE created_at < ?",
                ((datetime.now() - timedelta(hours=6)).isoformat(timespec="seconds"),),
            )
            result = self._analyze_import(connection, rows)
            connection.execute(
                """
                INSERT INTO product_import_previews (
                    id, operator, operator_name, filename,
                    rows_json, result_json, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    preview_id,
                    operator,
                    operator_name,
                    Path(filename).name,
                    json.dumps(rows, ensure_ascii=False),
                    json.dumps(result, ensure_ascii=False),
                    now,
                ),
            )
        return {"preview_id": preview_id, "filename": Path(filename).name, **result}

    def commit_import_preview(
        self,
        preview_id: str,
        *,
        operator: str,
        operator_name: str,
    ) -> dict[str, Any]:
        with self._connect() as connection:
            preview = connection.execute(
                """
                SELECT * FROM product_import_previews
                WHERE id = ? AND committed_at = ''
                """,
                (preview_id,),
            ).fetchone()
            if not preview:
                raise ValueError("上传预览不存在或已提交，请重新上传")
            if preview["operator"] != operator:
                raise ValueError("只能提交本人创建的上传预览")
            created_at = datetime.fromisoformat(preview["created_at"])
            if datetime.now() - created_at > timedelta(hours=6):
                raise ValueError("上传预览已过期，请重新上传")
            rows = json.loads(preview["rows_json"])
            analysis = self._analyze_import(connection, rows)
            if analysis["failed_count"]:
                raise ValueError("文件中仍有失败记录，请修正后重新上传")

            created_count = 0
            updated_count = 0
            unchanged_count = 0
            for item in analysis["details"]:
                if item["result"] in {"跳过", "失败", "仅保存月度汇总"}:
                    continue
                row = next(
                    source
                    for source in rows
                    if source["row_number"] == item["row_number"]
                )
                existing_id = self._find_product_id_by_code(
                    connection, row["product_code"]
                )
                if existing_id is None:
                    product = {
                        "product_codes": [row["product_code"]],
                        "short_name": "",
                        "product_name": row["product_name"],
                        "snow_inventory": row["snow_inventory"],
                        "housekeeper_codes": [],
                        "specification": row["specification"],
                        "auxiliary_unit": row["auxiliary_unit"],
                        "settlement_price": None,
                        "stock_month": row["source_month"],
                    }
                    self._create_product_in_connection(
                        connection,
                        product,
                        operator=operator,
                        operator_name=operator_name,
                    )
                    created_count += 1
                elif item["result"] == "无变化":
                    unchanged_count += 1
                else:
                    self._update_from_stock_in_connection(
                        connection,
                        existing_id,
                        row,
                        operator=operator,
                        operator_name=operator_name,
                    )
                    updated_count += 1

            committed_at = _now()
            import_id = uuid.uuid4().hex
            connection.execute(
                "UPDATE product_import_previews SET committed_at = ? WHERE id = ?",
                (committed_at, preview_id),
            )
            connection.execute(
                """
                INSERT INTO product_imports (
                    id, operator, operator_name, filename, total_count,
                    created_count, updated_count, unchanged_count,
                    skipped_count, warning_count, failed_count,
                    monthly_inbound_tons, snow_inventory_boxes, committed_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    operator,
                    operator_name,
                    preview["filename"],
                    analysis["total_count"],
                    created_count,
                    updated_count,
                    unchanged_count,
                    analysis["skipped_count"],
                    analysis["warning_count"],
                    0,
                    analysis["monthly_inbound_tons"],
                    analysis["snow_inventory_boxes"],
                    committed_at,
                ),
            )
            for summary in analysis["monthly_summaries"]:
                connection.execute(
                    """
                    INSERT INTO product_import_monthly_summaries (
                        import_id, month, inbound_tons, inventory_boxes
                    ) VALUES (?, ?, ?, ?)
                    """,
                    (
                        import_id,
                        summary["month"],
                        summary["inbound_tons"],
                        summary["inventory_boxes"],
                    ),
                )
        return {
            "id": import_id,
            "total_count": analysis["total_count"],
            "created_count": created_count,
            "updated_count": updated_count,
            "unchanged_count": unchanged_count,
            "skipped_count": analysis["skipped_count"],
            "summary_only_count": analysis["summary_only_count"],
            "warning_count": analysis["warning_count"],
            "failed_count": 0,
            "monthly_inbound_tons": analysis["monthly_inbound_tons"],
            "snow_inventory_boxes": analysis["snow_inventory_boxes"],
            "monthly_summaries": analysis["monthly_summaries"],
            "committed_at": committed_at,
        }

    def _analyze_import(
        self,
        connection: sqlite3.Connection,
        rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        details: list[dict[str, Any]] = []
        counts = {
            "created_count": 0,
            "updated_count": 0,
            "unchanged_count": 0,
            "skipped_count": 0,
            "warning_count": 0,
            "failed_count": 0,
            "summary_only_count": 0,
        }
        seen_codes: set[tuple[str, str]] = set()
        current_month = datetime.now().strftime("%Y%m")
        for row in rows:
            code = row["product_code"]
            row_key = (row["source_month"], code)
            detail = {
                "row_number": row["row_number"],
                "product_code": code,
                "product_name": row["product_name"],
                "result": "",
                "message": "",
            }
            if row["source_unit"] != "箱":
                detail.update(
                    result="跳过",
                    message=f"单位为“{row['source_unit'] or '空'}”，仅导入箱",
                )
                counts["skipped_count"] += 1
            elif row_key in seen_codes:
                detail.update(result="失败", message="文件内商品编号重复")
                counts["failed_count"] += 1
            elif row["source_month"] != current_month:
                seen_codes.add(row_key)
                detail.update(
                    result="仅保存月度汇总",
                    message="非本月商品数据不写入产品明细",
                )
                counts["summary_only_count"] += 1
            else:
                seen_codes.add(row_key)
                existing_id = self._find_product_id_by_code(connection, code)
                warning_parts = []
                if row["specification"] is None:
                    warning_parts.append("未解析到规格")
                if not row["auxiliary_unit"]:
                    warning_parts.append("未解析到辅助单位")
                if warning_parts:
                    counts["warning_count"] += 1
                if existing_id is None:
                    detail["result"] = "新增待完善"
                    counts["created_count"] += 1
                else:
                    current = connection.execute(
                        "SELECT * FROM products WHERE id = ?", (existing_id,)
                    ).fetchone()
                    changes = (
                        current["product_name"] != row["product_name"]
                        or float(current["snow_inventory"]) != float(row["snow_inventory"])
                        or (
                            row["specification"] is not None
                            and current["specification"] != row["specification"]
                        )
                        or (
                            row["auxiliary_unit"]
                            and current["auxiliary_unit"] != row["auxiliary_unit"]
                        )
                    )
                    detail["result"] = "更新" if changes else "无变化"
                    counts[
                        "updated_count" if changes else "unchanged_count"
                    ] += 1
                detail["message"] = "；".join(warning_parts)
            details.append(detail)
        monthly_summaries = self._monthly_summaries(rows)
        selected_summary = next(
            (
                summary
                for summary in monthly_summaries
                if summary["month"] == current_month
            ),
            monthly_summaries[-1]
            if monthly_summaries
            else {"inbound_tons": 0.0, "inventory_boxes": 0.0},
        )
        return {
            "total_count": len(rows),
            "details": details,
            "monthly_inbound_tons": selected_summary["inbound_tons"],
            "snow_inventory_boxes": selected_summary["inventory_boxes"],
            "monthly_summaries": monthly_summaries,
            **counts,
        }

    @staticmethod
    def _monthly_summaries(
        rows: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        months = sorted(
            {
                str(row.get("source_month") or "")
                for row in rows
                if re.fullmatch(r"\d{6}", str(row.get("source_month") or ""))
            }
        )
        return [
            {
                "month": month,
                "inbound_tons": math.fsum(
                    float(row.get("inbound_kiloliters") or 0)
                    for row in rows
                    if row.get("source_month") == month
                ),
                "inventory_boxes": math.fsum(
                    float(row.get("snow_inventory") or 0)
                    for row in rows
                    if row.get("source_month") == month
                    and row.get("source_unit") == "箱"
                ),
            }
            for month in months
        ]

    def _create_product_in_connection(
        self,
        connection: sqlite3.Connection,
        product: dict[str, Any],
        *,
        operator: str,
        operator_name: str,
    ) -> int:
        normalized = normalize_product(product, allow_incomplete=True)
        now = _now()
        cursor = connection.execute(
            """
            INSERT INTO products (
                short_name, product_name, snow_inventory, specification,
                auxiliary_unit, settlement_price, status, stock_month,
                created_by, created_by_name, created_at,
                updated_by, updated_by_name, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                normalized["short_name"],
                normalized["product_name"],
                normalized["snow_inventory"],
                normalized["specification"],
                normalized["auxiliary_unit"],
                normalized["settlement_price"],
                normalized["status"],
                product.get("stock_month") or datetime.now().strftime("%Y%m"),
                operator,
                operator_name,
                now,
                operator,
                operator_name,
                now,
            ),
        )
        product_id = int(cursor.lastrowid)
        self._replace_codes(
            connection,
            product_id,
            normalized["product_codes"],
            normalized["housekeeper_codes"],
        )
        self._log(
            connection,
            product_id,
            operator,
            operator_name,
            "create",
            "上传创建待完善产品",
            normalized,
        )
        return product_id

    def _update_from_stock_in_connection(
        self,
        connection: sqlite3.Connection,
        product_id: int,
        row: dict[str, Any],
        *,
        operator: str,
        operator_name: str,
    ) -> None:
        current = connection.execute(
            "SELECT * FROM products WHERE id = ?", (product_id,)
        ).fetchone()
        specification = (
            row["specification"]
            if row["specification"] is not None
            else current["specification"]
        )
        auxiliary_unit = row["auxiliary_unit"] or current["auxiliary_unit"]
        complete = (
            bool(current["short_name"])
            and bool(
                connection.execute(
                    """
                    SELECT 1 FROM product_housekeeper_codes
                    WHERE product_id = ? LIMIT 1
                    """,
                    (product_id,),
                ).fetchone()
            )
            and specification is not None
            and bool(auxiliary_unit)
        )
        now = _now()
        connection.execute(
            """
            UPDATE products
            SET product_name = ?, snow_inventory = ?, specification = ?,
                auxiliary_unit = ?, status = ?, stock_month = ?,
                version = version + 1,
                updated_by = ?, updated_by_name = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                row["product_name"],
                row["snow_inventory"],
                specification,
                auxiliary_unit,
                "正常" if complete else "待完善",
                row["source_month"],
                operator,
                operator_name,
                now,
                product_id,
            ),
        )
        self._log(
            connection,
            product_id,
            operator,
            operator_name,
            "update",
            "上传更新产品档案",
            row,
        )

    @staticmethod
    def _find_product_id_by_code(
        connection: sqlite3.Connection, code: str
    ) -> int | None:
        row = connection.execute(
            """
            SELECT product_codes.product_id
            FROM product_codes
            JOIN products ON products.id = product_codes.product_id
            WHERE product_codes.code = ? AND products.deleted_at = ''
            """,
            (code,),
        ).fetchone()
        return int(row[0]) if row else None

    @staticmethod
    def _replace_codes(
        connection: sqlite3.Connection,
        product_id: int,
        product_codes: list[str],
        housekeeper_codes: list[str],
    ) -> None:
        connection.execute(
            "DELETE FROM product_codes WHERE product_id = ?", (product_id,)
        )
        connection.execute(
            "DELETE FROM product_housekeeper_codes WHERE product_id = ?",
            (product_id,),
        )
        connection.executemany(
            "INSERT INTO product_codes(product_id, code) VALUES (?, ?)",
            [(product_id, code) for code in product_codes],
        )
        connection.executemany(
            """
            INSERT INTO product_housekeeper_codes(product_id, code)
            VALUES (?, ?)
            """,
            [(product_id, code) for code in housekeeper_codes],
        )

    @staticmethod
    def _log(
        connection: sqlite3.Connection,
        product_id: int,
        operator: str,
        operator_name: str,
        action_type: str,
        action_summary: str,
        changes: dict[str, Any],
    ) -> None:
        connection.execute(
            """
            INSERT INTO product_audit_logs (
                product_id, operator, operator_name, action_type,
                action_summary, changes_json, operated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                product_id,
                operator,
                operator_name,
                action_type,
                action_summary,
                json.dumps(changes, ensure_ascii=False, default=str),
                _now(),
            ),
        )

    @staticmethod
    def _integrity_error(error: sqlite3.IntegrityError) -> ValueError:
        message = str(error)
        if "product_codes.code" in message:
            return ValueError("商品编码已被其他产品使用")
        if "product_housekeeper_codes.code" in message:
            return ValueError("管家婆编码已被其他产品使用")
        return ValueError("产品数据与已有记录冲突")

    @staticmethod
    def _public_product(
        connection: sqlite3.Connection, row: sqlite3.Row
    ) -> dict[str, Any]:
        item = dict(row)
        item["product_codes"] = [
            code_row[0]
            for code_row in connection.execute(
                "SELECT code FROM product_codes WHERE product_id = ? ORDER BY code",
                (row["id"],),
            ).fetchall()
        ]
        item["housekeeper_codes"] = [
            code_row[0]
            for code_row in connection.execute(
                """
                SELECT code FROM product_housekeeper_codes
                WHERE product_id = ? ORDER BY code
                """,
                (row["id"],),
            ).fetchall()
        ]
        for key in ("deleted_by", "deleted_by_name", "deleted_at"):
            item.pop(key, None)
        return item
