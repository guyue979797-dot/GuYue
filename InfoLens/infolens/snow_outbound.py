"""雪花出库文件解析、政策标签规则和按月覆盖存储。"""

from __future__ import annotations

import json
import math
import re
import sqlite3
import uuid
import zipfile
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.utils.exceptions import InvalidFileException

from infolens.customers import (
    CUSTOMER_FIELDS,
    SALESPEOPLE,
    SNOW_SALESPEOPLE,
    CustomerStore,
    normalize_customer,
    normalize_terminal_code,
)


POLICY_TAGS = (
    "主推店",
    "超勇冰冻10+2",
    "超勇冰冻10+1",
    "花车",
    "旺季套餐陈列",
    "纯生3+1",
    "夜市陈列",
)
POLICY_COLORS = (
    "blue",
    "cyan",
    "green",
    "orange",
    "purple",
    "magenta",
    "red",
    "lime",
    "gold",
    "arcoblue",
)
RULE_FIELDS = {
    "outbound_remark": {
        "label": "出库单备注",
        "operators": {"equals", "contains"},
    },
    "sale_type": {
        "label": "售卖类型",
        "operators": {"equals", "contains"},
    },
    "converted_boxes": {
        "label": "折合箱数",
        "operators": {"equals", "greater_than", "less_than"},
    },
}
RULE_OPERATOR_LABELS = {
    "equals": "等于",
    "contains": "包含",
    "greater_than": "大于",
    "less_than": "小于",
}
POLICY_GIFT_TYPES = (
    "试业用酒-协议终端",
    "促销赠酒-临时搭赠",
    "促销赠酒-渠道营销",
    "促销赠酒-置换用酒",
    "陈列赠酒",
)
OUTBOUND_REQUIRED_HEADERS = (
    "票号",
    "开票日期",
    "业务员",
    "对象编码",
    "对象名称",
    "折合箱数",
    "售卖类型",
    "出库单备注",
)
REVERSAL_TICKET_PATTERN = re.compile(r"^CX1-(\d{6})-(.+)$")
ORIGINAL_TICKET_MONTH_PATTERN = re.compile(r"^CK(\d{4})(\d{2})\d{2}")


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _excel_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _invoice_date(value: Any) -> str:
    parsed: date | None = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value)
            parsed = converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            parsed = None
    else:
        raw = _excel_text(value).replace("/", "-")
        match = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", raw)
        if match:
            try:
                parsed = date(
                    int(match.group(1)),
                    int(match.group(2)),
                    int(match.group(3)),
                )
            except ValueError:
                parsed = None
    if not parsed:
        raise ValueError("开票日期格式不正确")
    return parsed.isoformat()


def _converted_boxes(value: Any) -> float:
    raw = _excel_text(value)
    if not raw:
        raise ValueError("折合箱数不能为空")
    try:
        number = float(raw)
    except ValueError as exc:
        raise ValueError("折合箱数必须为数值") from exc
    if not math.isfinite(number):
        raise ValueError("折合箱数必须为有限数值")
    return number


def _parse_reversal_ticket(ticket_no: str) -> tuple[str, str] | None:
    match = REVERSAL_TICKET_PATTERN.fullmatch(str(ticket_no or "").strip())
    if not match:
        return None
    encoded_month = f"{match.group(1)[:4]}-{match.group(1)[4:6]}"
    try:
        datetime.strptime(encoded_month, "%Y-%m")
    except ValueError as exc:
        raise ValueError("冲销票号中的发生月份格式不正确") from exc
    original_ticket_no = match.group(2).strip()
    if not original_ticket_no:
        raise ValueError("冲销票号缺少被冲销的原票号")
    return encoded_month, original_ticket_no


def _month_from_original_ticket(ticket_no: str) -> str:
    match = ORIGINAL_TICKET_MONTH_PATTERN.match(str(ticket_no or "").strip())
    if not match:
        return ""
    month = f"{match.group(1)}-{match.group(2)}"
    try:
        datetime.strptime(month, "%Y-%m")
    except ValueError:
        return ""
    return month


def _effective_preview_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    reversed_tickets = {
        reversal[1]
        for row in rows
        if (reversal := _parse_reversal_ticket(row.get("ticket_no", "")))
    }
    return [
        row
        for row in rows
        if not _parse_reversal_ticket(row.get("ticket_no", ""))
        and row.get("ticket_no") not in reversed_tickets
    ]


def parse_outbound_workbook(file_stream: BinaryIO) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(file_stream, read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise ValueError("无法读取Excel文件，请确认文件为有效的.xlsx格式") from exc
    try:
        worksheet = workbook.active
        # 部分第三方导出的Excel把工作表尺寸错误标记为A1:A1。
        # read_only模式会信任该标记，重置后才能读到实际的全部列。
        worksheet.reset_dimensions()
        first_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )
        headers = [_excel_text(value) for value in first_row]
        while headers and not headers[-1]:
            headers.pop()
        duplicates = sorted({header for header in headers if headers.count(header) > 1})
        if duplicates:
            raise ValueError(f"Excel表头存在重复字段：{'、'.join(duplicates)}")
        missing = [header for header in OUTBOUND_REQUIRED_HEADERS if header not in headers]
        if missing:
            raise ValueError(f"Excel缺少表头：{'、'.join(missing)}")
        header_index = {header: index for index, header in enumerate(headers)}

        rows: list[dict[str, Any]] = []
        errors: list[str] = []
        ticket_identity: dict[str, tuple[str, str]] = {}
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if not any(value is not None and _excel_text(value) for value in values):
                continue

            def value(header: str) -> Any:
                index = header_index.get(header)
                return values[index] if index is not None and index < len(values) else None

            try:
                ticket_no = _excel_text(value("票号"))
                if not ticket_no:
                    raise ValueError("票号不能为空")
                invoice_date = _invoice_date(value("开票日期"))
                reversal = _parse_reversal_ticket(ticket_no)
                if reversal:
                    original_month = _month_from_original_ticket(reversal[1])
                    if original_month and reversal[0] != original_month:
                        raise ValueError(
                            "冲销票号中的发生月份与原票号月份不一致"
                        )
                terminal_code = normalize_terminal_code(value("对象编码"))
                customer_name = _excel_text(value("对象名称"))
                if not customer_name:
                    raise ValueError("对象名称不能为空")
                if len(ticket_no) > 100:
                    raise ValueError("票号不能超过100个字符")
                identity = (invoice_date[:7], terminal_code)
                previous_identity = ticket_identity.get(ticket_no)
                if previous_identity and previous_identity != identity:
                    raise ValueError("同一票号关联了不同月份或终端")
                ticket_identity[ticket_no] = identity
                rows.append(
                    {
                        "row_number": row_number,
                        "ticket_no": ticket_no,
                        "invoice_date": invoice_date,
                        "month": invoice_date[:7],
                        "terminal_code": terminal_code,
                        "customer_name": customer_name[:200],
                        "salesperson_raw": _excel_text(value("业务员"))[:50],
                        "address": _excel_text(value("地址"))[:500],
                        "phone": _excel_text(value("电话号码"))[:50],
                        "converted_boxes": _converted_boxes(value("折合箱数")),
                        "sale_type": _excel_text(value("售卖类型"))[:100],
                        "outbound_remark": _excel_text(value("出库单备注"))[:500],
                        "product_code": _excel_text(
                            value("商品编码")
                            if "商品编码" in header_index
                            else value("商品编号")
                        )[:100],
                        "product_name": _excel_text(
                            value("商品简称")
                            if "商品简称" in header_index
                            else value("商品名称")
                        )[:500],
                    }
                )
            except ValueError as exc:
                errors.append(f"第{row_number}行：{exc}")
                if len(errors) >= 100:
                    break
        if errors:
            suffix = "；错误超过100条，仅展示前100条" if len(errors) >= 100 else ""
            raise ValueError("；".join(errors) + suffix)
        if not rows:
            raise ValueError("Excel中没有可导入的出库数据")
        return rows
    finally:
        workbook.close()


def validate_rules(raw_rules: Any) -> list[dict[str, Any]]:
    if not isinstance(raw_rules, list) or not raw_rules:
        raise ValueError("至少需要配置一个标签规则")
    if len(raw_rules) > len(POLICY_TAGS):
        raise ValueError(f"标签规则最多{len(POLICY_TAGS)}组")
    normalized = []
    used_tags: set[str] = set()
    for group_index, group in enumerate(raw_rules, start=1):
        if not isinstance(group, dict):
            raise ValueError(f"第{group_index}个标签规则格式不正确")
        tag = str(group.get("tag") or "").strip()
        if tag not in POLICY_TAGS:
            raise ValueError(f"第{group_index}个规则的映射标签不正确")
        if tag in used_tags:
            raise ValueError(f"标签“{tag}”只能配置一个条件组")
        used_tags.add(tag)
        conditions = group.get("conditions")
        if not isinstance(conditions, list) or not 1 <= len(conditions) <= 3:
            raise ValueError(f"标签“{tag}”必须配置1至3个条件")
        normalized_conditions = []
        for condition_index, condition in enumerate(conditions, start=1):
            if not isinstance(condition, dict):
                raise ValueError(f"标签“{tag}”第{condition_index}个条件格式不正确")
            field = str(condition.get("field") or "").strip()
            operator = str(condition.get("operator") or "").strip()
            raw_value = str(condition.get("value") or "").strip()
            if field not in RULE_FIELDS:
                raise ValueError(f"标签“{tag}”第{condition_index}个条件字段不正确")
            if operator not in RULE_FIELDS[field]["operators"]:
                raise ValueError(f"标签“{tag}”第{condition_index}个条件操作符不正确")
            if not raw_value:
                raise ValueError(f"标签“{tag}”第{condition_index}个条件值不能为空")
            if field == "converted_boxes":
                try:
                    numeric_value = float(raw_value)
                except ValueError as exc:
                    raise ValueError(
                        f"标签“{tag}”第{condition_index}个折合箱数必须为数值"
                    ) from exc
                if not math.isfinite(numeric_value):
                    raise ValueError(
                        f"标签“{tag}”第{condition_index}个折合箱数必须为有限数值"
                    )
                value: str | float = numeric_value
            else:
                value = raw_value
            normalized_conditions.append(
                {"field": field, "operator": operator, "value": value}
            )
        normalized.append({"tag": tag, "conditions": normalized_conditions})
    return normalized


def _condition_matches(row: dict[str, Any], condition: dict[str, Any]) -> bool:
    field = condition["field"]
    operator = condition["operator"]
    expected = condition["value"]
    actual = row[field]
    if field == "converted_boxes":
        actual_number = float(actual)
        expected_number = float(expected)
        if operator == "equals":
            return math.isclose(actual_number, expected_number, abs_tol=1e-9)
        if operator == "greater_than":
            return actual_number > expected_number
        return actual_number < expected_number
    actual_text = str(actual or "").strip()
    expected_text = str(expected).strip()
    if operator == "equals":
        return actual_text == expected_text
    return expected_text in actual_text


def evaluate_policy_tags(
    rows: list[dict[str, Any]],
    rules: list[dict[str, Any]],
) -> dict[tuple[str, str, str], dict[str, Any]]:
    matches: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        if float(row["converted_boxes"]) < 0:
            continue
        for group in rules:
            if all(_condition_matches(row, condition) for condition in group["conditions"]):
                key = (row["month"], row["terminal_code"], group["tag"])
                matches.setdefault(
                    key,
                    {
                        "month": row["month"],
                        "terminal_code": row["terminal_code"],
                        "tag": group["tag"],
                        "ticket_no": row["ticket_no"],
                        "row_number": row["row_number"],
                    },
                )
    return matches


def _optional_nonnegative_integer(value: Any, label: str) -> int | None:
    if value in ("", None):
        return None
    if isinstance(value, bool):
        raise ValueError(f"{label}必须为非负整数")
    try:
        numeric_value = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label}必须为非负整数") from exc
    if (
        not math.isfinite(numeric_value)
        or numeric_value < 0
        or not numeric_value.is_integer()
    ):
        raise ValueError(f"{label}必须为非负整数")
    return int(numeric_value)


def normalize_policy(payload: dict[str, Any]) -> dict[str, Any]:
    name = str(payload.get("name") or "").strip()
    outbound_code = str(payload.get("outbound_code") or "").strip()
    explanation = str(payload.get("explanation") or "").strip()
    if not name:
        raise ValueError("标签名字不能为空")
    if len(name) > 10:
        raise ValueError("标签名字不能超过10个字")
    if not outbound_code:
        raise ValueError("出库编码不能为空")
    if len(outbound_code) > 100:
        raise ValueError("出库编码不能超过100个字符")
    if not explanation:
        raise ValueError("出库解释不能为空")
    if len(explanation) > 50:
        raise ValueError("出库解释不能超过50个字")
    try:
        year = int(payload.get("year"))
        month = int(payload.get("month"))
    except (TypeError, ValueError) as exc:
        raise ValueError("年份和月份必须为有效数值") from exc
    if year != 2026:
        raise ValueError("政策年份目前仅支持2026年")
    if not 1 <= month <= 12:
        raise ValueError("月份必须在1至12之间")

    set_limit = _optional_nonnegative_integer(
        payload.get("set_limit"),
        "套数限制",
    )
    month_target = _optional_nonnegative_integer(
        payload.get("month_target"),
        "月目标",
    )
    raw_conflict_policy_ids = payload.get("conflict_policy_ids", [])
    if raw_conflict_policy_ids in (None, ""):
        raw_conflict_policy_ids = []
    if not isinstance(raw_conflict_policy_ids, list):
        raise ValueError("冲突政策格式不正确")
    conflict_policy_ids = list(
        dict.fromkeys(
            str(policy_id).strip()
            for policy_id in raw_conflict_policy_ids
            if str(policy_id).strip()
        )
    )
    if len(conflict_policy_ids) > 50:
        raise ValueError("冲突政策最多选择50项")

    product_fields_supplied = any(
        key in payload
        for key in (
            "normal_sale_product_ids",
            "gift_product_ids",
            "gift_product_id",
            "gift_type",
        )
    )
    raw_normal_product_ids = payload.get("normal_sale_product_ids", [])
    if raw_normal_product_ids in (None, ""):
        raw_normal_product_ids = []
    if not isinstance(raw_normal_product_ids, list):
        raise ValueError("正常销售产品格式不正确")
    try:
        normal_sale_product_ids = list(
            dict.fromkeys(int(product_id) for product_id in raw_normal_product_ids)
        )
    except (TypeError, ValueError) as exc:
        raise ValueError("正常销售产品格式不正确") from exc
    raw_gift_product_ids = payload.get("gift_product_ids")
    if raw_gift_product_ids is None:
        legacy_gift_product_id = payload.get("gift_product_id")
        raw_gift_product_ids = (
            [] if legacy_gift_product_id in ("", None) else [legacy_gift_product_id]
        )
    if raw_gift_product_ids in (None, ""):
        raw_gift_product_ids = []
    if not isinstance(raw_gift_product_ids, list):
        raise ValueError("赠送产品格式不正确")
    try:
        gift_product_ids = list(
            dict.fromkeys(int(product_id) for product_id in raw_gift_product_ids)
        )
    except (TypeError, ValueError) as exc:
        raise ValueError("赠送产品格式不正确") from exc
    gift_type = str(payload.get("gift_type") or "").strip()
    if product_fields_supplied:
        if not normal_sale_product_ids:
            raise ValueError("正常销售产品不能为空")
        if not gift_product_ids:
            raise ValueError("赠送产品不能为空")
        if gift_type not in POLICY_GIFT_TYPES:
            raise ValueError("售卖类型不能为空或格式不正确")
        if len(normal_sale_product_ids) > 50:
            raise ValueError("正常销售产品最多选择50项")
        if len(gift_product_ids) > 50:
            raise ValueError("赠送产品最多选择50项")

    conditions = payload.get("conditions")
    if not isinstance(conditions, list) or not 1 <= len(conditions) <= 3:
        raise ValueError("标签必须配置1至3个命中条件")
    normalized_conditions: list[dict[str, Any]] = []
    used_fields: set[str] = set()
    for index, condition in enumerate(conditions, start=1):
        if not isinstance(condition, dict):
            raise ValueError(f"第{index}个条件格式不正确")
        field = str(condition.get("field") or "").strip()
        operator = str(condition.get("operator") or "").strip()
        raw_value = str(condition.get("value") or "").strip()
        if field not in RULE_FIELDS:
            raise ValueError(f"第{index}个条件字段不正确")
        if field in used_fields:
            raise ValueError(f"字段“{RULE_FIELDS[field]['label']}”只能定义一次")
        used_fields.add(field)
        if operator not in RULE_FIELDS[field]["operators"]:
            raise ValueError(f"第{index}个条件操作符不正确")
        if not raw_value:
            raise ValueError(f"第{index}个条件值不能为空")
        if field == "converted_boxes":
            try:
                value: str | float = float(raw_value)
            except ValueError as exc:
                raise ValueError(f"第{index}个折合箱数必须为数值") from exc
            if not math.isfinite(value):
                raise ValueError(f"第{index}个折合箱数必须为有限数值")
        else:
            value = raw_value
        normalized_conditions.append(
            {"field": field, "operator": operator, "value": value}
        )
    return {
        "name": name,
        "display_name": f"{month}月-{name}",
        "outbound_code": outbound_code,
        "explanation": explanation,
        "requires_photo": bool(payload.get("requires_photo")),
        "set_limit": set_limit,
        "month_target": month_target,
        "year": year,
        "month": month,
        "conditions": normalized_conditions,
        "conflict_policy_ids": conflict_policy_ids,
        "normal_sale_product_ids": normal_sale_product_ids,
        "gift_product_ids": gift_product_ids,
        "gift_type": gift_type,
    }


def evaluate_policies(
    rows: list[dict[str, Any]],
    policies: list[dict[str, Any]],
) -> dict[tuple[str, str, str], dict[str, Any]]:
    matches: dict[tuple[str, str, str], dict[str, Any]] = {}
    policies_by_period: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for policy in policies:
        policies_by_period[f"{int(policy['year']):04d}-{int(policy['month']):02d}"].append(
            policy
        )
    rows_by_ticket: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if float(row["converted_boxes"]) >= 0:
            rows_by_ticket[
                (row["month"], row["ticket_no"], row["terminal_code"])
            ].append(row)
    for (period, _ticket_no, terminal_code), ticket_rows in rows_by_ticket.items():
        for policy in policies_by_period.get(period, []):
            matched_row = _policy_ticket_match_row(ticket_rows, policy)
            if matched_row is not None:
                key = (period, terminal_code, policy["id"])
                matches.setdefault(
                    key,
                    {
                        "month": period,
                        "terminal_code": terminal_code,
                        "policy_id": policy["id"],
                        "tag": policy["display_name"],
                        "color": policy["color"],
                        "ticket_no": matched_row["ticket_no"],
                        "row_number": matched_row["row_number"],
                        "rule_snapshot": policy,
                    },
                )
    return matches


def _policy_ticket_match_row(
    ticket_rows: list[dict[str, Any]],
    policy: dict[str, Any],
) -> dict[str, Any] | None:
    eligible_rows = [
        row for row in ticket_rows if float(row.get("converted_boxes") or 0) >= 0
    ]
    if not eligible_rows:
        return None
    condition_row = next(
        (
            row
            for row in eligible_rows
            if all(
                _condition_matches(row, condition)
                for condition in policy.get("conditions", [])
            )
        ),
        None,
    )
    if condition_row is None:
        return None

    return condition_row


class SnowOutboundStore:
    def __init__(self, database_path: str | Path):
        self.database_path = Path(database_path)
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self._initialize()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.database_path, timeout=30)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys=ON")
        connection.execute("PRAGMA journal_mode=WAL")
        connection.execute("PRAGMA busy_timeout=30000")
        return connection

    def _initialize(self) -> None:
        with self._connect() as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS snow_rule_templates (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL UNIQUE,
                    rules_json TEXT NOT NULL,
                    is_default INTEGER NOT NULL DEFAULT 0,
                    created_by TEXT NOT NULL,
                    created_by_name TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    updated_by TEXT NOT NULL,
                    updated_by_name TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS snow_outbound_previews (
                    id TEXT PRIMARY KEY,
                    filename TEXT NOT NULL,
                    operator TEXT NOT NULL,
                    operator_name TEXT NOT NULL,
                    rows_json TEXT NOT NULL,
                    rules_json TEXT NOT NULL,
                    summary_json TEXT NOT NULL,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS snow_outbound_imports (
                    id TEXT PRIMARY KEY,
                    filename TEXT NOT NULL,
                    operator TEXT NOT NULL,
                    operator_name TEXT NOT NULL,
                    rules_json TEXT NOT NULL,
                    months_json TEXT NOT NULL,
                    row_count INTEGER NOT NULL,
                    ticket_count INTEGER NOT NULL,
                    terminal_count INTEGER NOT NULL,
                    tag_count INTEGER NOT NULL,
                    auto_customer_count INTEGER NOT NULL,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS snow_outbound_tickets (
                    month TEXT NOT NULL,
                    ticket_no TEXT NOT NULL,
                    invoice_date TEXT NOT NULL,
                    terminal_code TEXT NOT NULL,
                    customer_name TEXT NOT NULL,
                    salesperson_raw TEXT NOT NULL DEFAULT '',
                    address TEXT NOT NULL DEFAULT '',
                    phone TEXT NOT NULL DEFAULT '',
                    import_id TEXT NOT NULL,
                    PRIMARY KEY(month, ticket_no),
                    FOREIGN KEY(import_id) REFERENCES snow_outbound_imports(id)
                );
                CREATE INDEX IF NOT EXISTS idx_snow_tickets_terminal
                    ON snow_outbound_tickets(month, terminal_code);

                CREATE TABLE IF NOT EXISTS snow_outbound_lines (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    month TEXT NOT NULL,
                    ticket_no TEXT NOT NULL,
                    row_number INTEGER NOT NULL,
                    terminal_code TEXT NOT NULL,
                    converted_boxes REAL NOT NULL,
                    sale_type TEXT NOT NULL DEFAULT '',
                    outbound_remark TEXT NOT NULL DEFAULT '',
                    salesperson_raw TEXT NOT NULL DEFAULT '',
                    raw_json TEXT NOT NULL,
                    import_id TEXT NOT NULL,
                    UNIQUE(month, ticket_no, row_number),
                    FOREIGN KEY(month, ticket_no)
                        REFERENCES snow_outbound_tickets(month, ticket_no)
                        ON DELETE CASCADE,
                    FOREIGN KEY(import_id) REFERENCES snow_outbound_imports(id)
                );

                CREATE TABLE IF NOT EXISTS snow_outbound_reversals (
                    reversal_month TEXT NOT NULL,
                    reversal_ticket_no TEXT NOT NULL,
                    original_ticket_no TEXT NOT NULL,
                    original_month TEXT NOT NULL DEFAULT '',
                    terminal_code TEXT NOT NULL DEFAULT '',
                    reversal_date TEXT NOT NULL DEFAULT '',
                    reason TEXT NOT NULL DEFAULT '',
                    import_id TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    PRIMARY KEY(reversal_month, reversal_ticket_no),
                    FOREIGN KEY(import_id) REFERENCES snow_outbound_imports(id)
                );
                CREATE INDEX IF NOT EXISTS idx_snow_reversals_original
                    ON snow_outbound_reversals(original_ticket_no, original_month);

                CREATE TABLE IF NOT EXISTS customer_policy_tags (
                    month TEXT NOT NULL,
                    terminal_code TEXT NOT NULL,
                    tag TEXT NOT NULL,
                    matched_ticket_no TEXT NOT NULL,
                    matched_row_number INTEGER NOT NULL,
                    import_id TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    PRIMARY KEY(month, terminal_code, tag),
                    FOREIGN KEY(import_id) REFERENCES snow_outbound_imports(id)
                );
                CREATE INDEX IF NOT EXISTS idx_policy_tags_terminal
                    ON customer_policy_tags(terminal_code, month);

                CREATE TABLE IF NOT EXISTS snow_policy_reversal_matches (
                    month TEXT NOT NULL,
                    policy_id TEXT NOT NULL,
                    terminal_code TEXT NOT NULL,
                    original_ticket_no TEXT NOT NULL,
                    reversal_ticket_no TEXT NOT NULL,
                    reversal_date TEXT NOT NULL DEFAULT '',
                    reason TEXT NOT NULL DEFAULT '',
                    import_id TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    PRIMARY KEY (
                        month, policy_id, terminal_code,
                        original_ticket_no, reversal_ticket_no
                    )
                );
                CREATE INDEX IF NOT EXISTS idx_policy_reversal_matches_policy
                    ON snow_policy_reversal_matches(policy_id, month, terminal_code);

                CREATE TABLE IF NOT EXISTS snow_policies (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    display_name TEXT NOT NULL,
                    outbound_code TEXT NOT NULL UNIQUE,
                    explanation TEXT NOT NULL,
                    requires_photo INTEGER NOT NULL DEFAULT 0,
                    set_limit INTEGER,
                    month_target INTEGER,
                    year INTEGER NOT NULL,
                    month INTEGER NOT NULL,
                    color TEXT NOT NULL,
                    enabled INTEGER NOT NULL DEFAULT 1,
                    created_by TEXT NOT NULL,
                    created_by_name TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    updated_by TEXT NOT NULL,
                    updated_by_name TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    deleted_by TEXT NOT NULL DEFAULT '',
                    deleted_by_name TEXT NOT NULL DEFAULT '',
                    deleted_at TEXT NOT NULL DEFAULT '',
                    UNIQUE(year, month, name),
                    CHECK(month BETWEEN 1 AND 12),
                    CHECK(set_limit IS NULL OR set_limit >= 0),
                    CHECK(month_target IS NULL OR month_target >= 0)
                );
                CREATE INDEX IF NOT EXISTS idx_snow_policies_filter
                    ON snow_policies(deleted_at, year, month, enabled, updated_at DESC);

                CREATE TABLE IF NOT EXISTS snow_policy_conditions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    policy_id TEXT NOT NULL,
                    position INTEGER NOT NULL,
                    field TEXT NOT NULL,
                    operator TEXT NOT NULL,
                    value_text TEXT NOT NULL,
                    UNIQUE(policy_id, field),
                    UNIQUE(policy_id, position),
                    FOREIGN KEY(policy_id) REFERENCES snow_policies(id)
                );

                CREATE TABLE IF NOT EXISTS snow_policy_products (
                    policy_id TEXT NOT NULL,
                    product_id INTEGER NOT NULL,
                    role TEXT NOT NULL,
                    position INTEGER NOT NULL,
                    PRIMARY KEY(policy_id, role, product_id),
                    UNIQUE(policy_id, role, position),
                    CHECK(role IN ('normal_sale', 'gift')),
                    FOREIGN KEY(policy_id) REFERENCES snow_policies(id),
                    FOREIGN KEY(product_id) REFERENCES products(id)
                );
                CREATE INDEX IF NOT EXISTS idx_snow_policy_products_product
                    ON snow_policy_products(product_id, role);

                CREATE TABLE IF NOT EXISTS snow_policy_conflicts (
                    policy_id TEXT NOT NULL,
                    conflict_policy_id TEXT NOT NULL,
                    position INTEGER NOT NULL,
                    PRIMARY KEY(policy_id, conflict_policy_id),
                    UNIQUE(policy_id, position),
                    FOREIGN KEY(policy_id) REFERENCES snow_policies(id),
                    FOREIGN KEY(conflict_policy_id) REFERENCES snow_policies(id)
                );
                CREATE INDEX IF NOT EXISTS idx_snow_policy_conflicts_target
                    ON snow_policy_conflicts(conflict_policy_id);

                CREATE TABLE IF NOT EXISTS snow_policy_alert_results (
                    month TEXT NOT NULL,
                    policy_id TEXT NOT NULL,
                    terminal_code TEXT NOT NULL,
                    alert_type TEXT NOT NULL,
                    conflict_policy_ids_json TEXT NOT NULL DEFAULT '[]',
                    ticket_count INTEGER,
                    set_limit INTEGER,
                    details_json TEXT NOT NULL DEFAULT '[]',
                    evaluated_at TEXT NOT NULL,
                    PRIMARY KEY(month, policy_id, terminal_code, alert_type),
                    FOREIGN KEY(policy_id) REFERENCES snow_policies(id)
                );
                CREATE INDEX IF NOT EXISTS idx_snow_policy_alerts_policy
                    ON snow_policy_alert_results(policy_id, month, alert_type);

                CREATE TABLE IF NOT EXISTS snow_policy_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    policy_id TEXT NOT NULL,
                    operator TEXT NOT NULL,
                    operator_name TEXT NOT NULL,
                    action_type TEXT NOT NULL,
                    action_summary TEXT NOT NULL,
                    changes_json TEXT NOT NULL DEFAULT '{}',
                    operated_at TEXT NOT NULL,
                    FOREIGN KEY(policy_id) REFERENCES snow_policies(id)
                );
                CREATE INDEX IF NOT EXISTS idx_snow_policy_logs
                    ON snow_policy_logs(policy_id, operated_at DESC, id DESC);
                """
            )
            self._ensure_column(
                connection,
                "snow_outbound_previews",
                "update_policy",
                "INTEGER NOT NULL DEFAULT 1",
            )
            self._ensure_column(
                connection,
                "snow_outbound_imports",
                "update_policy",
                "INTEGER NOT NULL DEFAULT 1",
            )
            self._ensure_column(
                connection,
                "customer_policy_tags",
                "policy_id",
                "TEXT NOT NULL DEFAULT ''",
            )
            self._ensure_column(
                connection,
                "customer_policy_tags",
                "color",
                "TEXT NOT NULL DEFAULT 'gray'",
            )
            self._ensure_column(
                connection,
                "customer_policy_tags",
                "rule_snapshot_json",
                "TEXT NOT NULL DEFAULT '{}'",
            )
            self._ensure_column(
                connection,
                "snow_policies",
                "month_target",
                "INTEGER",
            )
            self._ensure_column(
                connection,
                "snow_policies",
                "gift_type",
                "TEXT NOT NULL DEFAULT ''",
            )
            self._ensure_column(
                connection,
                "snow_policy_alert_results",
                "details_json",
                "TEXT NOT NULL DEFAULT '[]'",
            )

    @staticmethod
    def _ensure_column(
        connection: sqlite3.Connection,
        table: str,
        column: str,
        definition: str,
    ) -> None:
        existing = {
            row["name"]
            for row in connection.execute(f"PRAGMA table_info({table})").fetchall()
        }
        if column not in existing:
            connection.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")

    @staticmethod
    def _product_requirements_active(connection: sqlite3.Connection) -> bool:
        products_table = connection.execute(
            """
            SELECT 1 FROM sqlite_master
            WHERE type = 'table' AND name = 'products'
            """
        ).fetchone()
        if not products_table:
            return False
        return bool(
            connection.execute(
                """
                SELECT 1 FROM products
                WHERE deleted_at = '' AND status = '正常'
                LIMIT 1
                """
            ).fetchone()
        )

    @classmethod
    def _policy_required_fields_complete(
        cls,
        connection: sqlite3.Connection,
        policy: dict[str, Any],
    ) -> bool:
        if not (
            str(policy.get("name") or "").strip()
            and str(policy.get("outbound_code") or "").strip()
            and str(policy.get("explanation") or "").strip()
            and policy.get("conditions")
        ):
            return False
        if not cls._product_requirements_active(connection):
            return True
        return bool(
            policy.get("normal_sale_product_ids")
            and policy.get("gift_product_ids")
            and str(policy.get("gift_type") or "").strip()
        )

    def _disable_incomplete_policies(
        self,
        connection: sqlite3.Connection,
    ) -> None:
        if not self._product_requirements_active(connection):
            return
        candidates = connection.execute(
            """
            SELECT * FROM snow_policies
            WHERE deleted_at = '' AND enabled = 1
            """
        ).fetchall()
        incomplete: list[str] = []
        for row in candidates:
            policy = self._hydrate_policy(connection, dict(row))
            if not policy["required_fields_complete"]:
                incomplete.append(policy["id"])
        if not incomplete:
            return
        placeholders = ",".join("?" for _ in incomplete)
        connection.execute(
            f"UPDATE snow_policies SET enabled = 0 WHERE id IN ({placeholders})",
            incomplete,
        )
        connection.execute(
            f"DELETE FROM snow_policy_alert_results WHERE policy_id IN ({placeholders})",
            incomplete,
        )

    def list_policies(
        self,
        *,
        year: str = "",
        month: str = "",
        outbound_code: str = "",
        name: str = "",
        enabled: str = "",
        requires_photo: bool | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> dict[str, Any]:
        page = max(1, int(page))
        page_size = max(1, min(100, int(page_size)))
        conditions = ["deleted_at = ''"]
        parameters: list[Any] = []
        if year:
            conditions.append("year = ?")
            parameters.append(int(year))
        if month:
            conditions.append("month = ?")
            parameters.append(int(month))
        if outbound_code.strip():
            conditions.append("outbound_code = ?")
            parameters.append(outbound_code.strip())
        if name.strip():
            conditions.append("(name LIKE ? OR display_name LIKE ?)")
            parameters.extend([f"%{name.strip()}%", f"%{name.strip()}%"])
        if enabled in {"true", "false", "1", "0"}:
            conditions.append("enabled = ?")
            parameters.append(1 if enabled in {"true", "1"} else 0)
        if requires_photo is not None:
            conditions.append("requires_photo = ?")
            parameters.append(int(requires_photo))
        where = " AND ".join(conditions)
        with self._connect() as connection:
            self._disable_incomplete_policies(connection)
            latest_upload_at = connection.execute(
                "SELECT MAX(created_at) FROM snow_outbound_imports"
            ).fetchone()[0]
            total = connection.execute(
                f"SELECT COUNT(*) FROM snow_policies WHERE {where}",
                parameters,
            ).fetchone()[0]
            rows = connection.execute(
                f"""
                SELECT * FROM snow_policies
                WHERE {where}
                ORDER BY year DESC, month DESC, updated_at DESC, rowid ASC
                LIMIT ? OFFSET ?
                """,
                [*parameters, page_size, (page - 1) * page_size],
            ).fetchall()
            items = [self._hydrate_policy(connection, dict(row)) for row in rows]
            if items:
                policy_ids = [item["id"] for item in items]
                placeholders = ",".join("?" for _ in policy_ids)
                shipped_rows = connection.execute(
                    f"""
                    SELECT policy_id, COUNT(DISTINCT terminal_code) AS shipped_count
                    FROM customer_policy_tags
                    WHERE policy_id IN ({placeholders})
                    GROUP BY policy_id
                    """,
                    policy_ids,
                ).fetchall()
                shipped_counts = {
                    row["policy_id"]: int(row["shipped_count"])
                    for row in shipped_rows
                }
                reversed_rows = connection.execute(
                    f"""
                    SELECT policy_id,
                           COUNT(DISTINCT terminal_code) AS reversed_count
                    FROM snow_policy_reversal_matches
                    WHERE policy_id IN ({placeholders})
                    GROUP BY policy_id
                    """,
                    policy_ids,
                ).fetchall()
                reversed_counts = {
                    row["policy_id"]: int(row["reversed_count"])
                    for row in reversed_rows
                }
                alert_rows = connection.execute(
                    f"""
                    SELECT policy_id, COUNT(DISTINCT terminal_code) AS alert_count
                    FROM snow_policy_alert_results
                    WHERE policy_id IN ({placeholders})
                    GROUP BY policy_id
                    """,
                    policy_ids,
                ).fetchall()
                alert_counts = {
                    row["policy_id"]: int(row["alert_count"])
                    for row in alert_rows
                }
                for item in items:
                    item["shipped_count"] = shipped_counts.get(item["id"], 0)
                    item["reversed_count"] = reversed_counts.get(item["id"], 0)
                    item["alert_count"] = alert_counts.get(item["id"], 0)
                self._attach_policy_reimbursement_metrics(connection, items)
        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "latest_upload_at": latest_upload_at or "",
        }

    def get_policy(
        self, policy_id: str, *, include_deleted: bool = False
    ) -> dict[str, Any] | None:
        conditions = ["id = ?"]
        if not include_deleted:
            conditions.append("deleted_at = ''")
        with self._connect() as connection:
            self._disable_incomplete_policies(connection)
            row = connection.execute(
                f"SELECT * FROM snow_policies WHERE {' AND '.join(conditions)}",
                (policy_id.strip(),),
            ).fetchone()
            return (
                self._hydrate_policy(connection, dict(row))
                if row is not None
                else None
            )

    def policy_summaries(
        self,
        policy_ids: list[str],
        *,
        include_deleted: bool = False,
    ) -> dict[str, dict[str, Any]]:
        normalized_ids = list(
            dict.fromkeys(
                policy_id.strip()
                for policy_id in policy_ids
                if policy_id.strip()
            )
        )
        if not normalized_ids:
            return {}
        placeholders = ",".join("?" for _ in normalized_ids)
        conditions = [f"id IN ({placeholders})"]
        if not include_deleted:
            conditions.append("deleted_at = ''")
        with self._connect() as connection:
            rows = connection.execute(
                f"""
                SELECT id, name, display_name, color, enabled, deleted_at
                FROM snow_policies
                WHERE {' AND '.join(conditions)}
                """,
                normalized_ids,
            ).fetchall()
        return {
            row["id"]: {
                "policy_id": row["id"],
                "name": row["name"],
                "tag": row["display_name"],
                "color": row["color"] or "arcoblue",
                "enabled": bool(row["enabled"]),
                "deleted": bool(row["deleted_at"]),
            }
            for row in rows
        }

    def active_policy_options(
        self,
        month: str,
        *,
        requires_photo_only: bool = False,
    ) -> list[dict[str, Any]]:
        match = re.fullmatch(r"(\d{4})-(\d{2})", month.strip())
        if not match:
            raise ValueError("照片月份格式不正确")
        year, month_number = int(match.group(1)), int(match.group(2))
        photo_condition = "AND requires_photo = 1" if requires_photo_only else ""
        with self._connect() as connection:
            rows = connection.execute(
                f"""
                SELECT id, display_name, name, color, year, month
                FROM snow_policies
                WHERE deleted_at = '' AND enabled = 1
                  {photo_condition}
                  AND year = ? AND month = ?
                ORDER BY created_at, rowid
                """,
                (year, month_number),
            ).fetchall()
        return [dict(row) for row in rows]

    def policy_options(
        self,
        *,
        year: int,
        month: int,
        exclude_id: str = "",
    ) -> list[dict[str, Any]]:
        if year != 2026 or not 1 <= month <= 12:
            raise ValueError("政策月份格式不正确")
        conditions = ["deleted_at = ''", "year = ?", "month = ?"]
        parameters: list[Any] = [year, month]
        if exclude_id.strip():
            conditions.append("id != ?")
            parameters.append(exclude_id.strip())
        with self._connect() as connection:
            rows = connection.execute(
                f"""
                SELECT id, display_name, name, color, enabled
                FROM snow_policies
                WHERE {' AND '.join(conditions)}
                ORDER BY created_at, rowid
                """,
                parameters,
            ).fetchall()
        return [
            {
                **dict(row),
                "enabled": bool(row["enabled"]),
            }
            for row in rows
        ]

    def policy_alert_terminals(self, policy_id: str) -> list[dict[str, Any]]:
        with self._connect() as connection:
            policy_row = connection.execute(
                "SELECT * FROM snow_policies WHERE id = ? AND deleted_at = ''",
                (policy_id.strip(),),
            ).fetchone()
            if not policy_row:
                raise ValueError("雪花政策标签不存在")
            rows = connection.execute(
                """
                SELECT alerts.terminal_code, alerts.alert_type,
                       alerts.conflict_policy_ids_json, alerts.ticket_count,
                       alerts.set_limit, alerts.details_json,
                       alerts.evaluated_at,
                       COALESCE(customers.customer_name, '') AS customer_name
                FROM snow_policy_alert_results AS alerts
                LEFT JOIN customers
                  ON customers.terminal_code = alerts.terminal_code
                 AND customers.deleted_at = ''
                WHERE alerts.policy_id = ?
                GROUP BY alerts.terminal_code, alerts.alert_type,
                         alerts.conflict_policy_ids_json, alerts.ticket_count,
                         alerts.set_limit, alerts.details_json,
                         alerts.evaluated_at, customers.customer_name
                ORDER BY alerts.terminal_code, alerts.alert_type
                """,
                (policy_id.strip(),),
            ).fetchall()
            conflict_ids = list(
                dict.fromkeys(
                    conflict_id
                    for row in rows
                    for conflict_id in json.loads(
                        row["conflict_policy_ids_json"] or "[]"
                    )
                )
            )
            conflict_policies: dict[str, dict[str, str]] = {}
            if conflict_ids:
                placeholders = ",".join("?" for _ in conflict_ids)
                conflict_policies = {
                    row["id"]: {
                        "id": row["id"],
                        "name": row["display_name"],
                        "color": row["color"] or "gray",
                    }
                    for row in connection.execute(
                        f"""
                        SELECT id, display_name, color FROM snow_policies
                        WHERE id IN ({placeholders})
                        """,
                        conflict_ids,
                    ).fetchall()
                }

        items_by_terminal: dict[str, dict[str, Any]] = {}
        for row in rows:
            item = items_by_terminal.setdefault(
                row["terminal_code"],
                {
                    "terminal_code": row["terminal_code"],
                    "customer_name": row["customer_name"],
                    "alert_names": [],
                    "conflict_policy_names": [],
                    "conflict_policies": [],
                    "ticket_count": None,
                    "set_limit": None,
                    "details": [],
                    "evaluated_at": row["evaluated_at"],
                },
            )
            if row["alert_type"] == "policy_conflict":
                item["alert_names"].append("雪花政策冲突告警")
                item["conflict_policies"] = [
                    conflict_policies.get(
                        conflict_id,
                        {
                            "id": conflict_id,
                            "name": conflict_id,
                            "color": "gray",
                        },
                    )
                    for conflict_id in json.loads(
                        row["conflict_policy_ids_json"] or "[]"
                    )
                ]
                item["conflict_policy_names"] = [
                    conflict_policy["name"]
                    for conflict_policy in item["conflict_policies"]
                ]
            elif row["alert_type"] == "duplicate_outbound":
                item["alert_names"].append("政策重复出库告警")
                item["ticket_count"] = row["ticket_count"]
                item["set_limit"] = row["set_limit"]
            elif row["alert_type"] == "normal_sale_product_error":
                item["alert_names"].append("正常销售产品错误告警")
            elif row["alert_type"] == "gift_product_error":
                item["alert_names"].append("赠送产品错误告警")
            elif row["alert_type"] == "sale_type_error":
                item["alert_names"].append("售卖类型错误告警")
            if row["alert_type"] in {
                "normal_sale_product_error",
                "gift_product_error",
                "sale_type_error",
            }:
                alert_name = {
                    "normal_sale_product_error": "正常销售产品错误告警",
                    "gift_product_error": "赠送产品错误告警",
                    "sale_type_error": "售卖类型错误告警",
                }[row["alert_type"]]
                for detail in json.loads(row["details_json"] or "[]"):
                    item["details"].append(
                        {
                            **detail,
                            "alert_name": alert_name,
                        }
                    )
        return list(items_by_terminal.values())

    def policy_months(self, *, requires_photo_only: bool = False) -> list[str]:
        photo_condition = "AND requires_photo = 1" if requires_photo_only else ""
        with self._connect() as connection:
            rows = connection.execute(
                f"""
                SELECT DISTINCT year, month
                FROM snow_policies
                WHERE deleted_at = ''
                  {photo_condition}
                ORDER BY year DESC, month DESC
                """
            ).fetchall()
        return [
            f"{int(row['year']):04d}-{int(row['month']):02d}" for row in rows
        ]

    def shipped_terminals(self, policy_id: str) -> list[dict[str, str]]:
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT tags.terminal_code,
                       COALESCE(customers.customer_name, '') AS customer_name,
                       COALESCE(
                           NULLIF(customers.salesperson, ''),
                           MAX(NULLIF(tickets.salesperson_raw, '')),
                           ''
                       ) AS salesperson
                FROM customer_policy_tags AS tags
                LEFT JOIN customers
                  ON customers.terminal_code = tags.terminal_code
                 AND customers.deleted_at = ''
                LEFT JOIN snow_outbound_tickets AS tickets
                  ON tickets.month = tags.month
                 AND tickets.terminal_code = tags.terminal_code
                WHERE tags.policy_id = ?
                GROUP BY tags.terminal_code, customers.customer_name,
                         customers.salesperson
                ORDER BY tags.terminal_code
                """,
                (policy_id.strip(),),
            ).fetchall()
        return [
            {
                "terminal_code": row["terminal_code"],
                "customer_name": row["customer_name"],
                "salesperson": row["salesperson"],
            }
            for row in rows
        ]

    def reversed_terminals(self, policy_id: str) -> list[dict[str, str]]:
        with self._connect() as connection:
            policy = connection.execute(
                "SELECT id FROM snow_policies WHERE id = ? AND deleted_at = ''",
                (policy_id.strip(),),
            ).fetchone()
            if not policy:
                raise ValueError("雪花政策标签不存在")
            rows = connection.execute(
                """
                SELECT matches.terminal_code,
                       COALESCE(customers.customer_name, '') AS customer_name,
                       COALESCE(
                           NULLIF(customers.salesperson, ''),
                           MAX(NULLIF(tickets.salesperson_raw, '')),
                           ''
                       ) AS salesperson,
                       MAX(matches.reversal_date) AS reversal_date
                FROM snow_policy_reversal_matches AS matches
                LEFT JOIN customers
                  ON customers.terminal_code = matches.terminal_code
                 AND customers.deleted_at = ''
                LEFT JOIN snow_outbound_tickets AS tickets
                  ON tickets.month = matches.month
                 AND tickets.ticket_no = matches.original_ticket_no
                WHERE matches.policy_id = ?
                GROUP BY matches.terminal_code, customers.customer_name,
                         customers.salesperson
                ORDER BY reversal_date DESC, matches.terminal_code
                """,
                (policy_id.strip(),),
            ).fetchall()
            items = []
            for row in rows:
                reason_rows = connection.execute(
                    """
                    SELECT DISTINCT reason
                    FROM snow_policy_reversal_matches
                    WHERE policy_id = ? AND terminal_code = ?
                      AND reason != ''
                    ORDER BY reversal_date DESC, reversal_ticket_no
                    """,
                    (policy_id.strip(), row["terminal_code"]),
                ).fetchall()
                items.append(
                    {
                        "terminal_code": row["terminal_code"],
                        "customer_name": row["customer_name"],
                        "salesperson": row["salesperson"],
                        "reversal_date": row["reversal_date"] or "",
                        "reason": "；".join(
                            reason["reason"] for reason in reason_rows
                        ),
                    }
                )
        return items

    def shipped_terminal_codes_by_policy(
        self,
        policy_ids: list[str],
    ) -> dict[str, set[str]]:
        normalized_ids = list(
            dict.fromkeys(
                policy_id.strip()
                for policy_id in policy_ids
                if policy_id.strip()
            )
        )
        if not normalized_ids:
            return {}
        placeholders = ",".join("?" for _ in normalized_ids)
        with self._connect() as connection:
            rows = connection.execute(
                f"""
                SELECT policy_id, terminal_code
                FROM customer_policy_tags
                WHERE policy_id IN ({placeholders})
                GROUP BY policy_id, terminal_code
                """,
                normalized_ids,
            ).fetchall()
        result: dict[str, set[str]] = defaultdict(set)
        for row in rows:
            result[row["policy_id"]].add(row["terminal_code"])
        return dict(result)

    def policy_tags_for_terminals(
        self,
        month: str,
        terminal_codes: list[str],
    ) -> dict[str, list[dict[str, str]]]:
        normalized_codes = list(
            dict.fromkeys(
                str(item).strip()
                for item in terminal_codes
                if str(item).strip()
            )
        )
        if not month.strip() or not normalized_codes:
            return {}
        placeholders = ",".join("?" for _ in normalized_codes)
        with self._connect() as connection:
            rows = connection.execute(
                f"""
                SELECT terminal_code, tag, color, policy_id
                FROM customer_policy_tags
                WHERE month = ? AND terminal_code IN ({placeholders})
                ORDER BY tag, policy_id
                """,
                [month.strip(), *normalized_codes],
            ).fetchall()
        result: dict[str, list[dict[str, str]]] = defaultdict(list)
        seen: set[tuple[str, str]] = set()
        for row in rows:
            key = (row["terminal_code"], row["policy_id"] or row["tag"])
            if key in seen:
                continue
            seen.add(key)
            result[row["terminal_code"]].append(
                {
                    "policy_id": row["policy_id"],
                    "tag": row["tag"],
                    "color": row["color"] or "arcoblue",
                }
            )
        return dict(result)

    def active_policies_for_months(
        self,
        connection: sqlite3.Connection,
        months: list[str],
    ) -> list[dict[str, Any]]:
        if not months:
            return []
        period_pairs = [(int(value[:4]), int(value[5:7])) for value in months]
        period_sql = " OR ".join("(year = ? AND month = ?)" for _ in period_pairs)
        parameters = [item for pair in period_pairs for item in pair]
        rows = connection.execute(
            f"""
            SELECT * FROM snow_policies
            WHERE deleted_at = '' AND enabled = 1 AND ({period_sql})
            ORDER BY year, month, created_at, id
            """,
            parameters,
        ).fetchall()
        return [self._hydrate_policy(connection, dict(row)) for row in rows]

    @staticmethod
    def _latest_policy_snapshot_for_month(
        connection: sqlite3.Connection,
        month: str,
    ) -> tuple[list[dict[str, Any]], bool]:
        rows = connection.execute(
            """
            SELECT rules_json
            FROM snow_outbound_imports
            WHERE update_policy = 1
            ORDER BY created_at DESC, rowid DESC
            """
        ).fetchall()
        for row in rows:
            try:
                rules = json.loads(row["rules_json"] or "[]")
            except (TypeError, ValueError, json.JSONDecodeError):
                continue
            if not isinstance(rules, list):
                continue
            modern = [
                rule
                for rule in rules
                if isinstance(rule, dict)
                and rule.get("id")
                and f"{int(rule.get('year', 0)):04d}-{int(rule.get('month', 0)):02d}"
                == month
            ]
            if modern:
                return modern, False
            if rules and all(
                isinstance(rule, dict) and "tag" in rule and "id" not in rule
                for rule in rules
            ):
                return rules, True
        return [], False

    @staticmethod
    def _effective_outbound_rows(
        connection: sqlite3.Connection,
        month: str,
    ) -> list[dict[str, Any]]:
        reversed_ticket_rows = connection.execute(
            """
            SELECT DISTINCT original_ticket_no
            FROM snow_outbound_reversals
            WHERE original_month = ?
            """,
            (month,),
        ).fetchall()
        reversed_tickets = {
            row["original_ticket_no"] for row in reversed_ticket_rows
        }
        rows = connection.execute(
            """
            SELECT ticket_no, raw_json
            FROM snow_outbound_lines
            WHERE month = ?
            ORDER BY row_number, id
            """,
            (month,),
        ).fetchall()
        return [
            json.loads(row["raw_json"] or "{}")
            for row in rows
            if not _parse_reversal_ticket(row["ticket_no"])
            and row["ticket_no"] not in reversed_tickets
        ]

    @staticmethod
    def _product_settlement_prices(
        connection: sqlite3.Connection,
        product_names: list[str],
    ) -> dict[str, float | None]:
        normalized_names = list(
            dict.fromkeys(name.strip() for name in product_names if name.strip())
        )
        if not normalized_names:
            return {}
        if connection.execute(
            """
            SELECT 1 FROM sqlite_master
            WHERE type = 'table' AND name = 'products'
            """
        ).fetchone() is None:
            return {}
        placeholders = ",".join("?" for _ in normalized_names)
        product_rows = connection.execute(
            f"""
            SELECT product_name, settlement_price, status, updated_at, id
            FROM products
            WHERE deleted_at = ''
              AND product_name IN ({placeholders})
            ORDER BY CASE status WHEN '正常' THEN 0 ELSE 1 END,
                     updated_at DESC, id DESC
            """,
            normalized_names,
        ).fetchall()
        prices: dict[str, float | None] = {}
        for product_row in product_rows:
            if product_row["product_name"] in prices:
                continue
            prices[product_row["product_name"]] = (
                float(product_row["settlement_price"])
                if product_row["settlement_price"] is not None
                else None
            )
        return prices

    def _attach_policy_reimbursement_metrics(
        self,
        connection: sqlite3.Connection,
        policies: list[dict[str, Any]],
    ) -> None:
        """批量附加政策核销箱数及金额，避免列表页逐标签查询数据库。"""

        rows_by_period: dict[str, list[dict[str, Any]]] = {}
        for policy in policies:
            period = f"{int(policy['year']):04d}-{int(policy['month']):02d}"
            if period not in rows_by_period:
                rows_by_period[period] = self._effective_outbound_rows(
                    connection,
                    period,
                )

        product_names = [
            str(row.get("product_name") or "").strip()
            for rows in rows_by_period.values()
            for row in rows
        ]
        prices = self._product_settlement_prices(connection, product_names)
        for policy in policies:
            outbound_code = str(policy.get("outbound_code") or "").strip()
            sale_type = str(policy.get("gift_type") or "").strip()
            period = f"{int(policy['year']):04d}-{int(policy['month']):02d}"
            quantity_total = 0.0
            amount_total = 0.0
            if outbound_code and sale_type:
                for row in rows_by_period.get(period, []):
                    try:
                        quantity = float(row.get("converted_boxes") or 0)
                    except (TypeError, ValueError):
                        continue
                    if quantity < 0:
                        continue
                    if outbound_code not in str(row.get("outbound_remark") or ""):
                        continue
                    if str(row.get("sale_type") or "").strip() != sale_type:
                        continue
                    quantity_total += quantity
                    price = prices.get(
                        str(row.get("product_name") or "").strip()
                    )
                    if price is not None:
                        amount_total += quantity * price
            policy["reimbursement_quantity"] = round(quantity_total, 6)
            policy["reimbursement_amount"] = round(amount_total, 2)

    def policy_export_rows(
        self,
        policy_id: str,
    ) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        """返回同时命中出库编码和售卖类型的有效出库明细。"""

        normalized_id = policy_id.strip()
        if not normalized_id:
            raise ValueError("雪花政策标签不存在")
        with self._connect() as connection:
            policy_row = connection.execute(
                """
                SELECT * FROM snow_policies
                WHERE id = ? AND deleted_at = ''
                """,
                (normalized_id,),
            ).fetchone()
            if policy_row is None:
                raise ValueError("雪花政策标签不存在")
            policy = self._hydrate_policy(connection, dict(policy_row))
            outbound_code = str(policy.get("outbound_code") or "").strip()
            sale_type = str(policy.get("gift_type") or "").strip()
            if not outbound_code or not sale_type:
                raise ValueError("请先补全标签的出库编码和售卖类型")

            period = f"{int(policy['year']):04d}-{int(policy['month']):02d}"
            matched_rows: list[dict[str, Any]] = []
            for row in self._effective_outbound_rows(connection, period):
                try:
                    quantity = float(row.get("converted_boxes") or 0)
                except (TypeError, ValueError):
                    continue
                if quantity < 0:
                    continue
                if outbound_code not in str(row.get("outbound_remark") or ""):
                    continue
                if str(row.get("sale_type") or "").strip() != sale_type:
                    continue
                matched_rows.append({**row, "quantity": quantity})

            product_names = [
                str(row.get("product_name") or "").strip()
                for row in matched_rows
            ]
            prices_by_name = self._product_settlement_prices(
                connection,
                product_names,
            )

            export_rows = []
            for row in matched_rows:
                source_product_name = str(row.get("product_name") or "").strip()
                export_rows.append(
                    {
                        "terminal_code": str(row.get("terminal_code") or ""),
                        "customer_name": str(row.get("customer_name") or ""),
                        "product_name": source_product_name,
                        "quantity": row["quantity"],
                        "settlement_price": prices_by_name.get(
                            source_product_name
                        ),
                    }
                )
            return policy, export_rows

    def _rebuild_month_policy_results(
        self,
        connection: sqlite3.Connection,
        month: str,
        *,
        policy_snapshot: list[dict[str, Any]],
        legacy_rules: bool,
        fallback_import_id: str,
    ) -> int:
        if not policy_snapshot:
            return 0
        if legacy_rules:
            normalized_snapshot = validate_rules(policy_snapshot)
            policy_ids: list[str] = []
        else:
            normalized_snapshot = policy_snapshot
            policy_ids = [
                policy["id"]
                for policy in normalized_snapshot
                if f"{int(policy['year']):04d}-{int(policy['month']):02d}" == month
            ]
            if not policy_ids:
                return 0

        if legacy_rules:
            connection.execute(
                "DELETE FROM customer_policy_tags WHERE month = ?",
                (month,),
            )
        else:
            placeholders = ",".join("?" for _ in policy_ids)
            connection.execute(
                f"""
                DELETE FROM customer_policy_tags
                WHERE month = ? AND policy_id IN ({placeholders})
                """,
                [month, *policy_ids],
            )
            connection.execute(
                f"""
                DELETE FROM snow_policy_reversal_matches
                WHERE month = ? AND policy_id IN ({placeholders})
                """,
                [month, *policy_ids],
            )

        import_row = connection.execute(
            """
            SELECT import_id FROM snow_outbound_tickets
            WHERE month = ?
            ORDER BY rowid DESC LIMIT 1
            """,
            (month,),
        ).fetchone()
        import_id = (
            import_row["import_id"] if import_row is not None else fallback_import_id
        )
        now = _now()
        effective_rows = self._effective_outbound_rows(connection, month)
        matches = (
            evaluate_policy_tags(effective_rows, normalized_snapshot)
            if legacy_rules
            else evaluate_policies(effective_rows, normalized_snapshot)
        )
        for match in matches.values():
            connection.execute(
                """
                INSERT INTO customer_policy_tags (
                    month, terminal_code, tag, matched_ticket_no,
                    matched_row_number, import_id, created_at,
                    policy_id, color, rule_snapshot_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    match["month"],
                    match["terminal_code"],
                    match["tag"],
                    match["ticket_no"],
                    match["row_number"],
                    import_id,
                    now,
                    match.get("policy_id", ""),
                    match.get("color", "gray"),
                    json.dumps(match.get("rule_snapshot", {}), ensure_ascii=False),
                ),
            )

        if not legacy_rules:
            reversal_rows = connection.execute(
                """
                SELECT original_ticket_no, reversal_ticket_no,
                       reversal_date, reason, import_id
                FROM snow_outbound_reversals
                WHERE original_month = ?
                ORDER BY reversal_date, reversal_ticket_no
                """,
                (month,),
            ).fetchall()
            reversals_by_original: dict[str, list[sqlite3.Row]] = defaultdict(list)
            for reversal in reversal_rows:
                reversals_by_original[reversal["original_ticket_no"]].append(
                    reversal
                )
            if reversals_by_original:
                placeholders = ",".join("?" for _ in reversals_by_original)
                original_rows = connection.execute(
                    f"""
                    SELECT ticket_no, raw_json
                    FROM snow_outbound_lines
                    WHERE month = ? AND ticket_no IN ({placeholders})
                    ORDER BY row_number, id
                    """,
                    [month, *reversals_by_original.keys()],
                ).fetchall()
                inserted: set[tuple[str, str, str, str]] = set()
                raw_rows_by_ticket: dict[str, list[dict[str, Any]]] = defaultdict(list)
                for original_row in original_rows:
                    raw_rows_by_ticket[original_row["ticket_no"]].append(
                        json.loads(original_row["raw_json"] or "{}")
                    )
                for original_ticket_no, ticket_rows in raw_rows_by_ticket.items():
                    for policy in normalized_snapshot:
                        matched_row = _policy_ticket_match_row(ticket_rows, policy)
                        if matched_row is None:
                            continue
                        for reversal in reversals_by_original[
                            original_ticket_no
                        ]:
                            key = (
                                policy["id"],
                                matched_row["terminal_code"],
                                original_ticket_no,
                                reversal["reversal_ticket_no"],
                            )
                            if key in inserted:
                                continue
                            inserted.add(key)
                            connection.execute(
                                """
                                INSERT INTO snow_policy_reversal_matches (
                                    month, policy_id, terminal_code,
                                    original_ticket_no, reversal_ticket_no,
                                    reversal_date, reason, import_id, created_at
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    month,
                                    policy["id"],
                                    matched_row["terminal_code"],
                                    original_ticket_no,
                                    reversal["reversal_ticket_no"],
                                    reversal["reversal_date"],
                                    reversal["reason"],
                                    reversal["import_id"],
                                    now,
                                ),
                            )
        return len(matches)

    def create_policy(
        self,
        payload: dict[str, Any],
        *,
        operator: str,
        operator_name: str,
    ) -> dict[str, Any]:
        policy = normalize_policy(payload)
        now = _now()
        policy_id = (
            f"POL-{policy['year']:04d}{policy['month']:02d}-"
            f"{uuid.uuid4().hex[:6].upper()}"
        )
        with self._connect() as connection:
            color_index = connection.execute(
                "SELECT COUNT(*) FROM snow_policies"
            ).fetchone()[0]
            color = POLICY_COLORS[color_index % len(POLICY_COLORS)]
            try:
                self._validate_conflict_policies(
                    connection,
                    policy_id,
                    policy["year"],
                    policy["month"],
                    policy["conflict_policy_ids"],
                )
                self._validate_policy_products(connection, policy)
                connection.execute(
                    """
                    INSERT INTO snow_policies (
                        id, name, display_name, outbound_code, explanation,
                        requires_photo, set_limit, month_target,
                        year, month, color, enabled, gift_type,
                        created_by, created_by_name, created_at,
                        updated_by, updated_by_name, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        policy_id,
                        policy["name"],
                        policy["display_name"],
                        policy["outbound_code"],
                        policy["explanation"],
                        int(policy["requires_photo"]),
                        policy["set_limit"],
                        policy["month_target"],
                        policy["year"],
                        policy["month"],
                        color,
                        policy["gift_type"],
                        operator,
                        operator_name,
                        now,
                        operator,
                        operator_name,
                        now,
                    ),
                )
                self._replace_policy_conditions(
                    connection, policy_id, policy["conditions"]
                )
                self._replace_policy_conflicts(
                    connection, policy_id, policy["conflict_policy_ids"]
                )
                self._replace_policy_products(connection, policy_id, policy)
            except sqlite3.IntegrityError as exc:
                raise self._policy_integrity_error(exc, policy) from exc
            self._insert_policy_log(
                connection,
                policy_id=policy_id,
                operator=operator,
                operator_name=operator_name,
                action_type="create",
                action_summary="新建雪花出库政策",
                changes={"after": policy},
                operated_at=now,
            )
            self._recompute_policy_alerts(
                connection,
                [f"{policy['year']:04d}-{policy['month']:02d}"],
            )
            row = connection.execute(
                "SELECT * FROM snow_policies WHERE id = ?", (policy_id,)
            ).fetchone()
            return self._hydrate_policy(connection, dict(row))

    def update_policy(
        self,
        policy_id: str,
        payload: dict[str, Any],
        *,
        operator: str,
        operator_name: str,
    ) -> dict[str, Any]:
        policy = normalize_policy(payload)
        now = _now()
        with self._connect() as connection:
            current_row = connection.execute(
                "SELECT * FROM snow_policies WHERE id = ? AND deleted_at = ''",
                (policy_id,),
            ).fetchone()
            if not current_row:
                raise ValueError("雪花出库政策不存在")
            before = self._hydrate_policy(connection, dict(current_row))
            try:
                self._validate_conflict_policies(
                    connection,
                    policy_id,
                    policy["year"],
                    policy["month"],
                    policy["conflict_policy_ids"],
                )
                self._validate_policy_products(connection, policy)
                connection.execute(
                    """
                    UPDATE snow_policies
                    SET name = ?, display_name = ?, outbound_code = ?,
                        explanation = ?, requires_photo = ?, set_limit = ?,
                        month_target = ?, year = ?, month = ?, gift_type = ?,
                        updated_by = ?,
                        updated_by_name = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        policy["name"],
                        policy["display_name"],
                        policy["outbound_code"],
                        policy["explanation"],
                        int(policy["requires_photo"]),
                        policy["set_limit"],
                        policy["month_target"],
                        policy["year"],
                        policy["month"],
                        policy["gift_type"],
                        operator,
                        operator_name,
                        now,
                        policy_id,
                    ),
                )
                self._replace_policy_conditions(
                    connection, policy_id, policy["conditions"]
                )
                self._replace_policy_conflicts(
                    connection, policy_id, policy["conflict_policy_ids"]
                )
                self._replace_policy_products(connection, policy_id, policy)
            except sqlite3.IntegrityError as exc:
                raise self._policy_integrity_error(exc, policy) from exc
            self._insert_policy_log(
                connection,
                policy_id=policy_id,
                operator=operator,
                operator_name=operator_name,
                action_type="update",
                action_summary="修改雪花出库政策",
                changes={"before": before, "after": policy},
                operated_at=now,
            )
            affected_months = {
                f"{int(before['year']):04d}-{int(before['month']):02d}",
                f"{policy['year']:04d}-{policy['month']:02d}",
            }
            self._recompute_policy_alerts(connection, sorted(affected_months))
            row = connection.execute(
                "SELECT * FROM snow_policies WHERE id = ?", (policy_id,)
            ).fetchone()
            return self._hydrate_policy(connection, dict(row))

    def set_policy_enabled(
        self,
        policy_id: str,
        enabled: bool,
        *,
        operator: str,
        operator_name: str,
    ) -> dict[str, Any]:
        now = _now()
        with self._connect() as connection:
            current = connection.execute(
                "SELECT * FROM snow_policies WHERE id = ? AND deleted_at = ''",
                (policy_id,),
            ).fetchone()
            if not current:
                raise ValueError("雪花出库政策不存在")
            hydrated = self._hydrate_policy(connection, dict(current))
            if enabled and not hydrated["required_fields_complete"]:
                raise ValueError("请先编辑并补全所有必填项后再启用")
            connection.execute(
                """
                UPDATE snow_policies
                SET enabled = ?, updated_by = ?, updated_by_name = ?, updated_at = ?
                WHERE id = ?
                """,
                (int(enabled), operator, operator_name, now, policy_id),
            )
            self._insert_policy_log(
                connection,
                policy_id=policy_id,
                operator=operator,
                operator_name=operator_name,
                action_type="enable" if enabled else "disable",
                action_summary="启用雪花出库政策" if enabled else "停用雪花出库政策",
                changes={"enabled": enabled},
                operated_at=now,
            )
            policy_month = connection.execute(
                "SELECT year, month FROM snow_policies WHERE id = ?",
                (policy_id,),
            ).fetchone()
            self._recompute_policy_alerts(
                connection,
                [f"{int(policy_month['year']):04d}-{int(policy_month['month']):02d}"],
            )
            row = connection.execute(
                "SELECT * FROM snow_policies WHERE id = ?", (policy_id,)
            ).fetchone()
            return self._hydrate_policy(connection, dict(row))

    def delete_policy(
        self,
        policy_id: str,
        *,
        operator: str,
        operator_name: str,
    ) -> None:
        now = _now()
        with self._connect() as connection:
            current = connection.execute(
                "SELECT id FROM snow_policies WHERE id = ? AND deleted_at = ''",
                (policy_id,),
            ).fetchone()
            if not current:
                raise ValueError("雪花出库政策不存在")
            connection.execute(
                """
                UPDATE snow_policies
                SET enabled = 0, deleted_by = ?, deleted_by_name = ?,
                    deleted_at = ?, updated_by = ?, updated_by_name = ?,
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    operator,
                    operator_name,
                    now,
                    operator,
                    operator_name,
                    now,
                    policy_id,
                ),
            )
            self._insert_policy_log(
                connection,
                policy_id=policy_id,
                operator=operator,
                operator_name=operator_name,
                action_type="delete",
                action_summary="删除雪花出库政策",
                changes={},
                operated_at=now,
            )
            connection.execute(
                """
                DELETE FROM snow_policy_conflicts
                WHERE policy_id = ? OR conflict_policy_id = ?
                """,
                (policy_id, policy_id),
            )
            deleted_policy = connection.execute(
                "SELECT year, month FROM snow_policies WHERE id = ?",
                (policy_id,),
            ).fetchone()
            self._recompute_policy_alerts(
                connection,
                [
                    f"{int(deleted_policy['year']):04d}-"
                    f"{int(deleted_policy['month']):02d}"
                ],
            )

    def list_policy_logs(self, policy_id: str) -> list[dict[str, Any]]:
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT id, operator, operator_name, action_type,
                       action_summary, changes_json, operated_at
                FROM snow_policy_logs
                WHERE policy_id = ?
                ORDER BY operated_at DESC, id DESC
                """,
                (policy_id,),
            ).fetchall()
        items = []
        for row in rows:
            item = dict(row)
            item["changes"] = json.loads(item.pop("changes_json") or "{}")
            items.append(item)
        return items

    @staticmethod
    def _policy_integrity_error(
        error: sqlite3.IntegrityError,
        policy: dict[str, Any],
    ) -> ValueError:
        message = str(error)
        if "outbound_code" in message:
            return ValueError(f"出库编码“{policy['outbound_code']}”已存在")
        if "year, snow_policies.month, snow_policies.name" in message:
            return ValueError(
                f"{policy['year']}年{policy['month']}月已存在标签“{policy['name']}”"
            )
        return ValueError("政策数据与已有记录冲突")

    @staticmethod
    def _replace_policy_conditions(
        connection: sqlite3.Connection,
        policy_id: str,
        conditions: list[dict[str, Any]],
    ) -> None:
        connection.execute(
            "DELETE FROM snow_policy_conditions WHERE policy_id = ?", (policy_id,)
        )
        for position, condition in enumerate(conditions, start=1):
            connection.execute(
                """
                INSERT INTO snow_policy_conditions (
                    policy_id, position, field, operator, value_text
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (
                    policy_id,
                    position,
                    condition["field"],
                    condition["operator"],
                    str(condition["value"]),
                ),
            )

    @staticmethod
    def _validate_policy_products(
        connection: sqlite3.Connection,
        policy: dict[str, Any],
    ) -> None:
        product_ids = list(
            dict.fromkeys(
                [
                    *policy["normal_sale_product_ids"],
                    *policy["gift_product_ids"],
                ]
            )
        )
        if not product_ids:
            return
        products_table = connection.execute(
            """
            SELECT 1 FROM sqlite_master
            WHERE type = 'table' AND name = 'products'
            """
        ).fetchone()
        if not products_table:
            raise ValueError("产品档案尚未初始化")
        placeholders = ",".join("?" for _ in product_ids)
        rows = connection.execute(
            f"""
            SELECT id, status, deleted_at
            FROM products
            WHERE id IN ({placeholders})
            """,
            product_ids,
        ).fetchall()
        valid_ids = {
            int(row["id"])
            for row in rows
            if not row["deleted_at"] and row["status"] == "正常"
        }
        invalid_ids = [
            product_id for product_id in product_ids if product_id not in valid_ids
        ]
        if invalid_ids:
            raise ValueError("所选产品不存在、已删除或状态不是正常")

    @staticmethod
    def _replace_policy_products(
        connection: sqlite3.Connection,
        policy_id: str,
        policy: dict[str, Any],
    ) -> None:
        products_table = connection.execute(
            """
            SELECT 1 FROM sqlite_master
            WHERE type = 'table' AND name = 'products'
            """
        ).fetchone()
        if not products_table:
            if policy["normal_sale_product_ids"] or policy["gift_product_ids"]:
                raise ValueError("产品档案尚未初始化")
            return
        connection.execute(
            "DELETE FROM snow_policy_products WHERE policy_id = ?",
            (policy_id,),
        )
        for position, product_id in enumerate(
            policy["normal_sale_product_ids"],
            start=1,
        ):
            connection.execute(
                """
                INSERT INTO snow_policy_products (
                    policy_id, product_id, role, position
                ) VALUES (?, ?, 'normal_sale', ?)
                """,
                (policy_id, product_id, position),
            )
        for position, product_id in enumerate(
            policy["gift_product_ids"],
            start=1,
        ):
            connection.execute(
                """
                INSERT INTO snow_policy_products (
                    policy_id, product_id, role, position
                ) VALUES (?, ?, 'gift', ?)
                """,
                (policy_id, product_id, position),
            )

    @staticmethod
    def _validate_conflict_policies(
        connection: sqlite3.Connection,
        policy_id: str,
        year: int,
        month: int,
        conflict_policy_ids: list[str],
    ) -> None:
        if policy_id in conflict_policy_ids:
            raise ValueError("冲突政策不能选择当前标签")
        if not conflict_policy_ids:
            return
        placeholders = ",".join("?" for _ in conflict_policy_ids)
        rows = connection.execute(
            f"""
            SELECT id, year, month
            FROM snow_policies
            WHERE id IN ({placeholders}) AND deleted_at = ''
            """,
            conflict_policy_ids,
        ).fetchall()
        by_id = {row["id"]: row for row in rows}
        missing = [
            conflict_id
            for conflict_id in conflict_policy_ids
            if conflict_id not in by_id
        ]
        if missing:
            raise ValueError("所选冲突政策不存在或已删除")
        if any(
            int(row["year"]) != year or int(row["month"]) != month
            for row in rows
        ):
            raise ValueError("冲突政策必须与当前标签属于同一月份")

    @staticmethod
    def _replace_policy_conflicts(
        connection: sqlite3.Connection,
        policy_id: str,
        conflict_policy_ids: list[str],
    ) -> None:
        connection.execute(
            "DELETE FROM snow_policy_conflicts WHERE policy_id = ?",
            (policy_id,),
        )
        for position, conflict_policy_id in enumerate(
            conflict_policy_ids,
            start=1,
        ):
            connection.execute(
                """
                INSERT INTO snow_policy_conflicts (
                    policy_id, conflict_policy_id, position
                ) VALUES (?, ?, ?)
                """,
                (policy_id, conflict_policy_id, position),
            )

    def _recompute_policy_alerts(
        self,
        connection: sqlite3.Connection,
        months: list[str],
    ) -> None:
        normalized_months = list(dict.fromkeys(months))
        if not normalized_months:
            return
        for period in normalized_months:
            year, month = int(period[:4]), int(period[5:7])
            connection.execute(
                "DELETE FROM snow_policy_alert_results WHERE month = ?",
                (period,),
            )
            policy_rows = connection.execute(
                """
                SELECT * FROM snow_policies
                WHERE deleted_at = '' AND enabled = 1
                  AND year = ? AND month = ?
                ORDER BY created_at, id
                """,
                (year, month),
            ).fetchall()
            policies = [
                self._hydrate_policy(connection, dict(row))
                for row in policy_rows
            ]
            if not policies:
                continue

            tag_rows = connection.execute(
                """
                SELECT policy_id, terminal_code
                FROM customer_policy_tags
                WHERE month = ? AND policy_id != ''
                GROUP BY policy_id, terminal_code
                """,
                (period,),
            ).fetchall()
            terminals_by_policy: dict[str, set[str]] = defaultdict(set)
            for row in tag_rows:
                terminals_by_policy[row["policy_id"]].add(row["terminal_code"])

            outbound_rows = [
                row
                for row in self._effective_outbound_rows(connection, period)
                if float(row.get("converted_boxes") or 0) >= 0
            ]
            outbound_by_ticket: dict[
                tuple[str, str], list[dict[str, Any]]
            ] = defaultdict(list)
            for row in outbound_rows:
                outbound_by_ticket[
                    (
                        str(row.get("terminal_code") or ""),
                        str(row.get("ticket_no") or ""),
                    )
                ].append(row)
            now = _now()

            for policy in policies:
                current_terminals = terminals_by_policy.get(policy["id"], set())
                if not current_terminals:
                    continue
                matched_tickets_by_terminal: dict[
                    str, list[list[dict[str, Any]]]
                ] = defaultdict(list)
                for (terminal_code, ticket_no), ticket_rows in (
                    outbound_by_ticket.items()
                ):
                    if (
                        terminal_code in current_terminals
                        and ticket_no
                        and _policy_ticket_match_row(ticket_rows, policy)
                    ):
                        matched_tickets_by_terminal[terminal_code].append(ticket_rows)
                conflict_ids = policy["conflict_policy_ids"]
                if conflict_ids:
                    for terminal_code in current_terminals:
                        matched_conflicts = [
                            conflict_id
                            for conflict_id in conflict_ids
                            if terminal_code
                            in terminals_by_policy.get(conflict_id, set())
                        ]
                        if matched_conflicts:
                            connection.execute(
                                """
                                INSERT INTO snow_policy_alert_results (
                                    month, policy_id, terminal_code, alert_type,
                                    conflict_policy_ids_json, ticket_count,
                                    set_limit, evaluated_at
                                ) VALUES (?, ?, ?, 'policy_conflict', ?, NULL, NULL, ?)
                                """,
                                (
                                    period,
                                    policy["id"],
                                    terminal_code,
                                    json.dumps(
                                        matched_conflicts,
                                        ensure_ascii=False,
                                    ),
                                    now,
                                ),
                            )

                if policy["set_limit"] is not None:
                    tickets_by_terminal: dict[str, set[str]] = defaultdict(set)
                    for terminal_code, ticket_groups in (
                        matched_tickets_by_terminal.items()
                    ):
                        tickets_by_terminal[terminal_code].update(
                            str(ticket_rows[0].get("ticket_no") or "")
                            for ticket_rows in ticket_groups
                            if ticket_rows
                        )
                    for terminal_code, ticket_numbers in tickets_by_terminal.items():
                        ticket_count = len(ticket_numbers)
                        if ticket_count > policy["set_limit"]:
                            connection.execute(
                                """
                                INSERT INTO snow_policy_alert_results (
                                    month, policy_id, terminal_code, alert_type,
                                    conflict_policy_ids_json, ticket_count,
                                    set_limit, evaluated_at
                                ) VALUES (?, ?, ?, 'duplicate_outbound', '[]', ?, ?, ?)
                                """,
                                (
                                    period,
                                    policy["id"],
                                    terminal_code,
                                    ticket_count,
                                    policy["set_limit"],
                                    now,
                                ),
                            )

                normal_product_names = {
                    str(product.get("product_name") or "").strip()
                    for product in policy.get("normal_sale_products", [])
                    if str(product.get("product_name") or "").strip()
                }
                gift_product_names = {
                    str(product.get("product_name") or "").strip()
                    for product in policy.get("gift_products", [])
                    if str(product.get("product_name") or "").strip()
                }
                expected_gift_type = str(policy.get("gift_type") or "").strip()
                if (
                    not normal_product_names
                    and not gift_product_names
                    and not expected_gift_type
                ):
                    continue

                for terminal_code, ticket_groups in (
                    matched_tickets_by_terminal.items()
                ):
                    details_by_type: dict[str, list[dict[str, str]]] = defaultdict(
                        list
                    )
                    for ticket_rows in ticket_groups:
                        for row in ticket_rows:
                            sale_type = str(row.get("sale_type") or "").strip()
                            product_name = str(
                                row.get("product_name") or ""
                            ).strip()
                            detail = {
                                "ticket_no": str(row.get("ticket_no") or ""),
                                "product_name": product_name,
                                "actual_sale_type": sale_type,
                                "row_number": str(row.get("row_number") or ""),
                            }
                            if sale_type == "正常销售":
                                if product_name not in normal_product_names:
                                    details_by_type[
                                        "normal_sale_product_error"
                                    ].append(
                                        {
                                            **detail,
                                            "reason": (
                                                f"商品“{product_name or '空'}”"
                                                "不属于标签设置的正常销售产品"
                                            ),
                                        }
                                    )
                                continue
                            if product_name not in gift_product_names:
                                details_by_type["gift_product_error"].append(
                                    {
                                        **detail,
                                        "reason": (
                                            f"商品“{product_name or '空'}”"
                                            "不属于标签设置的赠送产品"
                                        ),
                                    }
                                )
                            if sale_type != expected_gift_type:
                                details_by_type["sale_type_error"].append(
                                    {
                                        **detail,
                                        "reason": (
                                            f"售卖类型“{sale_type or '空'}”"
                                            f"不等于“{expected_gift_type}”"
                                        ),
                                    }
                                )
                    for alert_type, details in details_by_type.items():
                        if not details:
                            continue
                        connection.execute(
                            """
                            INSERT INTO snow_policy_alert_results (
                                month, policy_id, terminal_code, alert_type,
                                conflict_policy_ids_json, ticket_count,
                                set_limit, details_json, evaluated_at
                            ) VALUES (?, ?, ?, ?, '[]', NULL, NULL, ?, ?)
                            """,
                            (
                                period,
                                policy["id"],
                                terminal_code,
                                alert_type,
                                json.dumps(details, ensure_ascii=False),
                                now,
                            ),
                        )

    @staticmethod
    def _insert_policy_log(
        connection: sqlite3.Connection,
        *,
        policy_id: str,
        operator: str,
        operator_name: str,
        action_type: str,
        action_summary: str,
        changes: dict[str, Any],
        operated_at: str,
    ) -> None:
        connection.execute(
            """
            INSERT INTO snow_policy_logs (
                policy_id, operator, operator_name, action_type,
                action_summary, changes_json, operated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                policy_id,
                operator,
                operator_name,
                action_type,
                action_summary,
                json.dumps(changes, ensure_ascii=False),
                operated_at,
            ),
        )

    @staticmethod
    def _hydrate_policy(
        connection: sqlite3.Connection,
        item: dict[str, Any],
    ) -> dict[str, Any]:
        rows = connection.execute(
            """
            SELECT field, operator, value_text
            FROM snow_policy_conditions
            WHERE policy_id = ?
            ORDER BY position
            """,
            (item["id"],),
        ).fetchall()
        item["conditions"] = [
            {
                "field": row["field"],
                "operator": row["operator"],
                "value": (
                    float(row["value_text"])
                    if row["field"] == "converted_boxes"
                    else row["value_text"]
                ),
            }
            for row in rows
        ]
        conflict_rows = connection.execute(
            """
            SELECT conflict_policy_id
            FROM snow_policy_conflicts
            WHERE policy_id = ?
            ORDER BY position
            """,
            (item["id"],),
        ).fetchall()
        item["conflict_policy_ids"] = [
            row["conflict_policy_id"] for row in conflict_rows
        ]
        products_table = connection.execute(
            """
            SELECT 1 FROM sqlite_master
            WHERE type = 'table' AND name = 'products'
            """
        ).fetchone()
        product_rows = (
            connection.execute(
                """
                SELECT links.product_id, links.role, products.short_name,
                       products.product_name, products.status,
                       GROUP_CONCAT(product_codes.code, '、')
                           AS product_codes_text
                FROM snow_policy_products AS links
                LEFT JOIN products ON products.id = links.product_id
                LEFT JOIN product_codes
                  ON product_codes.product_id = links.product_id
                WHERE links.policy_id = ?
                GROUP BY links.product_id, links.role, links.position,
                         products.short_name, products.product_name,
                         products.status
                ORDER BY links.role, links.position
                """,
                (item["id"],),
            ).fetchall()
            if products_table
            else []
        )
        normal_products = [
            {
                "id": int(row["product_id"]),
                "short_name": row["short_name"] or "",
                "product_name": row["product_name"] or "",
                "status": row["status"] or "",
                "product_codes": (
                    str(row["product_codes_text"]).split("、")
                    if row["product_codes_text"]
                    else []
                ),
            }
            for row in product_rows
            if row["role"] == "normal_sale"
        ]
        gift_products = [
            {
                "id": int(row["product_id"]),
                "short_name": row["short_name"] or "",
                "product_name": row["product_name"] or "",
                "status": row["status"] or "",
                "product_codes": (
                    str(row["product_codes_text"]).split("、")
                    if row["product_codes_text"]
                    else []
                ),
            }
            for row in product_rows
            if row["role"] == "gift"
        ]
        item["normal_sale_product_ids"] = [
            product["id"] for product in normal_products
        ]
        item["normal_sale_products"] = normal_products
        item["gift_product_ids"] = [
            product["id"] for product in gift_products
        ]
        item["gift_products"] = gift_products
        # 保留旧字段供历史前端平滑升级，新逻辑统一使用多选字段。
        item["gift_product_id"] = (
            gift_products[0]["id"] if gift_products else None
        )
        item["gift_product"] = gift_products[0] if gift_products else None
        item["enabled"] = bool(item["enabled"])
        item["requires_photo"] = bool(item["requires_photo"])
        item["required_fields_complete"] = (
            SnowOutboundStore._policy_required_fields_complete(connection, item)
        )
        return item

    def list_templates(self) -> list[dict[str, Any]]:
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT id, name, rules_json, is_default, created_by,
                       created_by_name, created_at, updated_by_name, updated_at
                FROM snow_rule_templates
                ORDER BY is_default DESC, updated_at DESC, name ASC
                """
            ).fetchall()
        return [self._public_template(dict(row)) for row in rows]

    def save_template(
        self,
        *,
        template_id: str = "",
        name: str,
        rules: Any,
        is_default: bool,
        operator: str,
        operator_name: str,
        is_admin: bool = False,
    ) -> dict[str, Any]:
        name = str(name or "").strip()
        if not name:
            raise ValueError("模板名称不能为空")
        if len(name) > 100:
            raise ValueError("模板名称不能超过100个字符")
        normalized_rules = validate_rules(rules)
        now = _now()
        with self._connect() as connection:
            if is_default:
                connection.execute("UPDATE snow_rule_templates SET is_default = 0")
            if template_id:
                current = connection.execute(
                    "SELECT created_by FROM snow_rule_templates WHERE id = ?",
                    (template_id,),
                ).fetchone()
                if not current:
                    raise ValueError("规则模板不存在")
                if not is_admin and current["created_by"] != operator:
                    raise PermissionError("只能修改自己创建的规则模板")
                try:
                    connection.execute(
                        """
                        UPDATE snow_rule_templates
                        SET name = ?, rules_json = ?, is_default = ?,
                            updated_by = ?, updated_by_name = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            name,
                            json.dumps(normalized_rules, ensure_ascii=False),
                            int(is_default),
                            operator,
                            operator_name,
                            now,
                            template_id,
                        ),
                    )
                except sqlite3.IntegrityError as exc:
                    raise ValueError(f"模板名称“{name}”已存在") from exc
            else:
                template_id = uuid.uuid4().hex
                try:
                    connection.execute(
                        """
                        INSERT INTO snow_rule_templates (
                            id, name, rules_json, is_default,
                            created_by, created_by_name, created_at,
                            updated_by, updated_by_name, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            template_id,
                            name,
                            json.dumps(normalized_rules, ensure_ascii=False),
                            int(is_default),
                            operator,
                            operator_name,
                            now,
                            operator,
                            operator_name,
                            now,
                        ),
                    )
                except sqlite3.IntegrityError as exc:
                    raise ValueError(f"模板名称“{name}”已存在") from exc
            row = connection.execute(
                "SELECT * FROM snow_rule_templates WHERE id = ?",
                (template_id,),
            ).fetchone()
        return self._public_template(dict(row))

    def delete_template(
        self,
        template_id: str,
        *,
        operator: str,
        is_admin: bool,
    ) -> None:
        with self._connect() as connection:
            current = connection.execute(
                "SELECT created_by FROM snow_rule_templates WHERE id = ?",
                (template_id,),
            ).fetchone()
            if not current:
                raise ValueError("规则模板不存在")
            if not is_admin and current["created_by"] != operator:
                raise PermissionError("只能删除自己创建的规则模板")
            connection.execute(
                "DELETE FROM snow_rule_templates WHERE id = ?",
                (template_id,),
            )

    def create_preview(
        self,
        *,
        filename: str,
        operator: str,
        operator_name: str,
        rows: list[dict[str, Any]],
        rules: Any = None,
        update_policy: bool = True,
    ) -> dict[str, Any]:
        months = sorted({row["month"] for row in rows})
        tickets = {row["ticket_no"] for row in rows}
        terminals = {row["terminal_code"] for row in rows}
        effective_rows = _effective_preview_rows(rows)
        with self._connect() as connection:
            if rules is None:
                policy_snapshot = (
                    self.active_policies_for_months(connection, months)
                    if update_policy
                    else []
                )
                matches = (
                    evaluate_policies(effective_rows, policy_snapshot)
                    if update_policy
                    else {}
                )
                legacy_rules = False
            else:
                # 兼容旧预览调用；新页面不再提交上传时规则。
                policy_snapshot = validate_rules(rules)
                matches = evaluate_policy_tags(effective_rows, policy_snapshot)
                legacy_rules = True
        tag_counts: dict[str, int] = defaultdict(int)
        tag_colors: dict[str, str] = {}
        for match in matches.values():
            tag_counts[match["tag"]] += 1
            tag_colors[match["tag"]] = match.get("color", "gray")

        with self._connect() as connection:
            cutoff = (datetime.now() - timedelta(hours=24)).isoformat(timespec="seconds")
            connection.execute(
                "DELETE FROM snow_outbound_previews WHERE created_at < ?",
                (cutoff,),
            )
            placeholders = ",".join("?" for _ in terminals)
            existing_rows = connection.execute(
                f"""
                SELECT terminal_code, deleted_at
                FROM customers
                WHERE terminal_code IN ({placeholders})
                """,
                sorted(terminals),
            ).fetchall()
            active_codes = {
                row["terminal_code"] for row in existing_rows if not row["deleted_at"]
            }
            deleted_codes = {
                row["terminal_code"] for row in existing_rows if row["deleted_at"]
            }
            if deleted_codes:
                examples = "、".join(sorted(deleted_codes)[:10])
                raise ValueError(
                    f"发现{len(deleted_codes)}个终端编码属于已删除客户档案，"
                    f"请先处理后再导入：{examples}"
                )
            preview_id = uuid.uuid4().hex
            summary = {
                "months": months,
                "row_count": len(rows),
                "ticket_count": len(tickets),
                "terminal_count": len(terminals),
                "matched_terminal_count": len(
                    {(item["month"], item["terminal_code"]) for item in matches.values()}
                ),
                "tag_count": len(matches),
                "tag_counts": dict(tag_counts),
                "tag_colors": tag_colors,
                "policy_count": len(policy_snapshot),
                "update_policy": bool(update_policy),
                "legacy_rules": legacy_rules,
                "negative_row_count": sum(
                    1 for row in rows if float(row["converted_boxes"]) < 0
                ),
                "reversal_ticket_count": len(
                    {
                        row["ticket_no"]
                        for row in rows
                        if _parse_reversal_ticket(row["ticket_no"])
                    }
                ),
                "auto_customer_count": len(terminals - active_codes),
                "unknown_salespeople": sorted(
                    {
                        row["salesperson_raw"]
                        for row in rows
                        if row["salesperson_raw"]
                        and row["salesperson_raw"]
                        not in {*SALESPEOPLE, *SNOW_SALESPEOPLE}
                    }
                ),
            }
            connection.execute(
                """
                INSERT INTO snow_outbound_previews (
                    id, filename, operator, operator_name, rows_json,
                    rules_json, summary_json, created_at, update_policy
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    preview_id,
                    filename,
                    operator,
                    operator_name,
                    json.dumps(rows, ensure_ascii=False),
                    json.dumps(policy_snapshot, ensure_ascii=False),
                    json.dumps(summary, ensure_ascii=False),
                    _now(),
                    int(update_policy),
                ),
            )
        return {"preview_id": preview_id, **summary}

    def commit_preview(
        self,
        preview_id: str,
        *,
        operator: str,
        operator_name: str,
        is_admin: bool = False,
    ) -> dict[str, Any]:
        with self._connect() as connection:
            connection.execute("BEGIN IMMEDIATE")
            preview = connection.execute(
                "SELECT * FROM snow_outbound_previews WHERE id = ?",
                (preview_id,),
            ).fetchone()
            if not preview:
                raise ValueError("预览记录不存在或已过期")
            if not is_admin and preview["operator"] != operator:
                raise PermissionError("不能提交其他用户的导入预览")
            created_at = datetime.fromisoformat(preview["created_at"])
            if datetime.now() - created_at > timedelta(hours=24):
                raise ValueError("预览记录已过期，请重新上传文件")
            rows = json.loads(preview["rows_json"])
            summary = json.loads(preview["summary_json"])
            policy_snapshot = json.loads(preview["rules_json"])
            update_policy = bool(preview["update_policy"])
            legacy_rules = bool(summary.get("legacy_rules"))
            if not update_policy:
                matches = {}
            elif legacy_rules:
                policy_snapshot = validate_rules(policy_snapshot)
                matches = evaluate_policy_tags(
                    _effective_preview_rows(rows),
                    policy_snapshot,
                )
            else:
                matches = evaluate_policies(
                    _effective_preview_rows(rows),
                    policy_snapshot,
                )
            months = summary["months"]
            terminal_codes = sorted({row["terminal_code"] for row in rows})
            placeholders = ",".join("?" for _ in terminal_codes)
            existing = connection.execute(
                f"""
                SELECT terminal_code, deleted_at
                FROM customers
                WHERE terminal_code IN ({placeholders})
                """,
                terminal_codes,
            ).fetchall()
            active_codes = {
                row["terminal_code"] for row in existing if not row["deleted_at"]
            }
            deleted_codes = {
                row["terminal_code"] for row in existing if row["deleted_at"]
            }
            if deleted_codes:
                raise ValueError("部分终端编码属于已删除客户档案，导入已取消")

            auto_customers = self._build_auto_customers(rows, active_codes)
            for payload in auto_customers:
                self._insert_system_customer(connection, payload)

            import_id = uuid.uuid4().hex
            now = _now()
            connection.execute(
                """
                INSERT INTO snow_outbound_imports (
                    id, filename, operator, operator_name, rules_json,
                    months_json, row_count, ticket_count, terminal_count,
                    tag_count, auto_customer_count, created_at, update_policy
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    preview["filename"],
                    operator,
                    operator_name,
                    json.dumps(policy_snapshot, ensure_ascii=False),
                    json.dumps(months, ensure_ascii=False),
                    len(rows),
                    len({row["ticket_no"] for row in rows}),
                    len(terminal_codes),
                    len(matches),
                    len(auto_customers),
                    now,
                    int(update_policy),
                ),
            )
            affected_months: set[str] = set()
            for month in months:
                previous_reversals = connection.execute(
                    """
                    SELECT original_month
                    FROM snow_outbound_reversals
                    WHERE reversal_month = ?
                    """,
                    (month,),
                ).fetchall()
                affected_months.update(
                    row["original_month"]
                    for row in previous_reversals
                    if row["original_month"]
                )
                connection.execute(
                    "DELETE FROM snow_outbound_reversals WHERE reversal_month = ?",
                    (month,),
                )
                connection.execute(
                    "DELETE FROM snow_outbound_tickets WHERE month = ?",
                    (month,),
                )

            ticket_rows: dict[tuple[str, str], dict[str, Any]] = {}
            for row in rows:
                ticket_rows.setdefault((row["month"], row["ticket_no"]), row)
            for (month, ticket_no), row in ticket_rows.items():
                connection.execute(
                    """
                    INSERT INTO snow_outbound_tickets (
                        month, ticket_no, invoice_date, terminal_code,
                        customer_name, salesperson_raw, address, phone, import_id
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        month,
                        ticket_no,
                        row["invoice_date"],
                        row["terminal_code"],
                        row["customer_name"],
                        row["salesperson_raw"],
                        row["address"],
                        row["phone"],
                        import_id,
                    ),
                )
            for row in rows:
                connection.execute(
                    """
                    INSERT INTO snow_outbound_lines (
                        month, ticket_no, row_number, terminal_code,
                        converted_boxes, sale_type, outbound_remark,
                        salesperson_raw, raw_json, import_id
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row["month"],
                        row["ticket_no"],
                        row["row_number"],
                        row["terminal_code"],
                        row["converted_boxes"],
                        row["sale_type"],
                        row["outbound_remark"],
                        row["salesperson_raw"],
                        json.dumps(row, ensure_ascii=False),
                        import_id,
                    ),
                )

            reversal_groups: dict[tuple[str, str], list[dict[str, Any]]] = (
                defaultdict(list)
            )
            for row in rows:
                reversal = _parse_reversal_ticket(row["ticket_no"])
                if reversal:
                    reversal_groups[(row["month"], row["ticket_no"])].append(row)
            for (reversal_month, reversal_ticket_no), reversal_rows in (
                reversal_groups.items()
            ):
                _encoded_month, original_ticket_no = _parse_reversal_ticket(
                    reversal_ticket_no
                ) or ("", "")
                original = connection.execute(
                    """
                    SELECT month, terminal_code
                    FROM snow_outbound_tickets
                    WHERE ticket_no = ?
                    ORDER BY invoice_date DESC, rowid DESC
                    LIMIT 1
                    """,
                    (original_ticket_no,),
                ).fetchone()
                original_month = (
                    original["month"]
                    if original is not None
                    else _month_from_original_ticket(original_ticket_no)
                )
                terminal_code = (
                    original["terminal_code"]
                    if original is not None
                    else reversal_rows[0]["terminal_code"]
                )
                reasons = list(
                    dict.fromkeys(
                        str(row.get("outbound_remark") or "").strip()
                        for row in reversal_rows
                        if str(row.get("outbound_remark") or "").strip()
                    )
                )
                connection.execute(
                    """
                    INSERT INTO snow_outbound_reversals (
                        reversal_month, reversal_ticket_no,
                        original_ticket_no, original_month, terminal_code,
                        reversal_date, reason, import_id, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        reversal_month,
                        reversal_ticket_no,
                        original_ticket_no,
                        original_month,
                        terminal_code,
                        max(row["invoice_date"] for row in reversal_rows),
                        "；".join(reasons),
                        import_id,
                        now,
                    ),
                )
                if original_month:
                    affected_months.add(original_month)

            pending_for_imported_months = connection.execute(
                f"""
                SELECT DISTINCT original_month
                FROM snow_outbound_reversals
                WHERE original_month IN ({','.join('?' for _ in months)})
                """,
                months,
            ).fetchall()
            affected_months.update(
                row["original_month"] for row in pending_for_imported_months
            )
            if update_policy:
                affected_months.update(months)

            rebuilt_tag_count = 0
            for month in sorted(affected_months):
                if month in months and update_policy:
                    if legacy_rules:
                        month_snapshot = policy_snapshot
                        month_legacy = True
                    else:
                        month_snapshot = [
                            policy
                            for policy in policy_snapshot
                            if f"{int(policy['year']):04d}-{int(policy['month']):02d}"
                            == month
                        ]
                        month_legacy = False
                else:
                    month_snapshot, month_legacy = (
                        self._latest_policy_snapshot_for_month(connection, month)
                    )
                rebuilt_tag_count += self._rebuild_month_policy_results(
                    connection,
                    month,
                    policy_snapshot=month_snapshot,
                    legacy_rules=month_legacy,
                    fallback_import_id=import_id,
                )
            self._recompute_policy_alerts(connection, sorted(affected_months))
            connection.execute(
                "DELETE FROM snow_outbound_previews WHERE id = ?",
                (preview_id,),
            )
        return {
            "id": import_id,
            "months": months,
            "row_count": len(rows),
            "ticket_count": len(ticket_rows),
            "terminal_count": len(terminal_codes),
            "tag_count": rebuilt_tag_count,
            "auto_customer_count": len(auto_customers),
            "update_policy": update_policy,
        }

    def list_months(self) -> list[str]:
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT DISTINCT month FROM snow_outbound_tickets
                ORDER BY month DESC
                """
            ).fetchall()
        return [row["month"] for row in rows]

    def list_policy_tags(self, month: str) -> list[str]:
        month = str(month or "").strip()
        if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", month):
            raise ValueError("政策月份格式不正确")
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT DISTINCT tag
                FROM customer_policy_tags
                WHERE month = ?
                ORDER BY tag
                """,
                (month,),
            ).fetchall()
        return [row["tag"] for row in rows]

    def _build_auto_customers(
        self,
        rows: list[dict[str, Any]],
        active_codes: set[str],
    ) -> list[dict[str, str]]:
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            if row["terminal_code"] not in active_codes:
                grouped[row["terminal_code"]].append(row)
        customers = []
        for terminal_code, terminal_rows in grouped.items():
            terminal_rows.sort(
                key=lambda item: (item["invoice_date"], -item["row_number"]),
                reverse=True,
            )
            customer_name = next(
                item["customer_name"] for item in terminal_rows if item["customer_name"]
            )
            salesperson = next(
                (
                    item["salesperson_raw"]
                    for item in terminal_rows
                    if item["salesperson_raw"] in SALESPEOPLE
                ),
                "",
            )
            snow_salesperson = next(
                (
                    item["salesperson_raw"]
                    for item in terminal_rows
                    if item["salesperson_raw"] in SNOW_SALESPEOPLE
                ),
                "",
            )
            address = next(
                (item["address"] for item in terminal_rows if item["address"]),
                "",
            )
            phone = next(
                (item["phone"] for item in terminal_rows if item["phone"]),
                "",
            )
            customers.append(
                normalize_customer(
                    {
                        "terminal_code": terminal_code,
                        "customer_name": customer_name,
                        "status": "运营",
                        "route": "",
                        "salesperson": salesperson,
                        "snow_salesperson": snow_salesperson,
                        "contact": "",
                        "address": address,
                        "phone": phone,
                        "remark": "由系统创建",
                    }
                )
            )
        return customers

    @staticmethod
    def _insert_system_customer(
        connection: sqlite3.Connection,
        customer: dict[str, str],
    ) -> None:
        now = _now()
        cursor = connection.execute(
            f"""
            INSERT INTO customers (
                {", ".join(CUSTOMER_FIELDS)},
                created_by, created_by_name, created_at,
                updated_by, updated_by_name, updated_at
            ) VALUES ({", ".join("?" for _ in CUSTOMER_FIELDS)}, ?, ?, ?, ?, ?, ?)
            """,
            [
                *(customer[field] for field in CUSTOMER_FIELDS),
                "system",
                "系统",
                now,
                "system",
                "系统",
                now,
            ],
        )
        CustomerStore._insert_log(
            connection,
            customer_id=cursor.lastrowid,
            operator="system",
            operator_name="系统",
            action_type="create",
            action_summary="雪花出库上传自动创建客户档案",
            changes={},
            operated_at=now,
        )

    @staticmethod
    def _public_template(item: dict[str, Any]) -> dict[str, Any]:
        item["rules"] = json.loads(item.pop("rules_json") or "[]")
        item["is_default"] = bool(item["is_default"])
        return item
