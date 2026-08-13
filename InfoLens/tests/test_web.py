import importlib
import io
import json
import os
import tempfile
import time
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash

from infolens.extractor import ExtractResult, SavedImage
class WebSecurityTests(unittest.TestCase):
    def setUp(self):
        self.output = tempfile.TemporaryDirectory()
        self.environment = patch.dict(
            os.environ,
            {
                "INFOLENS_AUTH_MODE": "password",
                "INFOLENS_ENV": "production",
                "INFOLENS_OUTPUT_ROOT": self.output.name,
                "INFOLENS_USERNAME": "team",
                "INFOLENS_PASSWORD_HASH": generate_password_hash(
                    "correct horse",
                    method="pbkdf2:sha256",
                ),
                "INFOLENS_SESSION_SECRET": "a" * 64,
            },
            clear=False,
        )
        self.environment.start()
        import web

        self.web = importlib.reload(web)
        self.client = self.web.app.test_client()

    def tearDown(self):
        self.environment.stop()
        self.output.cleanup()

    def login(self):
        self.client.get("/login")
        with self.client.session_transaction() as session:
            csrf_token = session["csrf_token"]
        return self.client.post(
            "/login",
            data={
                "username": "team",
                "password": "correct horse",
                "csrf_token": csrf_token,
            },
        )

    def test_protected_routes_require_login(self):
        self.assertEqual(self.client.get("/").status_code, 302)
        self.assertEqual(self.client.get("/api/results").status_code, 401)
        self.assertEqual(self.client.get("/api/distributions").status_code, 401)
        self.assertEqual(self.client.get("/api/extraction-records").status_code, 401)
        self.assertEqual(self.client.delete("/api/distributions").status_code, 401)
        self.assertEqual(self.client.post("/api/batch-extract").status_code, 401)
        self.assertEqual(
            self.client.get("/api/batch-extract/unknown").status_code,
            401,
        )
        self.assertEqual(self.client.get("/output/private.jpg").status_code, 401)
        self.assertEqual(self.client.get("/healthz").status_code, 200)

    def test_login_session_and_csrf(self):
        login_page = self.client.get("/login")
        self.assertEqual(login_page.status_code, 200)
        with self.client.session_transaction() as session:
            login_csrf = session["csrf_token"]

        bad = self.client.post(
            "/login",
            data={"username": "team", "password": "wrong", "csrf_token": login_csrf},
        )
        self.assertIn("账号或密码不正确", bad.get_data(as_text=True))

        good = self.client.post(
            "/login",
            data={"username": "team", "password": "correct horse", "csrf_token": login_csrf},
        )
        self.assertEqual(good.status_code, 302)

        session_response = self.client.get("/api/session")
        session_data = session_response.get_json()
        self.assertEqual(session_data["user"], "team")
        self.assertEqual(session_data["role"], "admin")
        self.assertTrue(session_data["is_admin"])
        self.assertTrue(session_data["csrf_token"])

        missing_csrf = self.client.post("/api/extract", json={"url": "x"})
        self.assertEqual(missing_csrf.status_code, 403)

        empty_url = self.client.post(
            "/api/extract",
            json={"url": ""},
            headers={"X-CSRF-Token": session_data["csrf_token"]},
        )
        self.assertEqual(empty_url.status_code, 400)

    def test_admin_can_manage_users(self):
        good = self.login()
        self.assertEqual(good.status_code, 302)
        session_data = self.client.get("/api/session").get_json()
        csrf_token = session_data["csrf_token"]

        users = self.client.get("/api/users")
        self.assertEqual(users.status_code, 200)
        initial_items = users.get_json()["items"]
        self.assertEqual(initial_items[0]["username"], "team")
        self.assertTrue(initial_items[0]["is_super_admin"])

        created = self.client.post(
            "/api/users",
            json={
                "username": "worker",
                "display_name": "普通用户",
                "password": "secret1",
                "role": "user",
                "status": "enabled",
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(created.status_code, 201)
        worker = created.get_json()
        self.assertEqual(worker["username"], "worker")

        updated = self.client.patch(
            f"/api/users/{worker['id']}",
            json={
                "display_name": "普通用户2",
                "role": "admin",
                "status": "disabled",
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(updated.status_code, 200)
        updated_data = updated.get_json()
        self.assertEqual(updated_data["role"], "admin")
        self.assertEqual(updated_data["status"], "disabled")

        forbidden_delete = self.client.delete(
            f"/api/users/{initial_items[0]['id']}",
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(forbidden_delete.status_code, 400)

        deleted = self.client.delete(
            f"/api/users/{worker['id']}",
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(deleted.status_code, 200)

    def test_customer_crud_intersection_query_import_and_permissions(self):
        self.login()
        session_data = self.client.get("/api/session").get_json()
        csrf_token = session_data["csrf_token"]
        base = {
            "status": "运营",
            "route": "一号线路",
            "salesperson": "黄春梅",
            "snow_salesperson": "陈家利",
            "contact": "",
            "address": "",
            "phone": "",
            "remark": "",
        }
        first = self.client.post(
            "/api/customers",
            json={
                **base,
                "terminal_code": "1000000001",
                "customer_name": "甲客户",
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(first.status_code, 201)
        first_data = first.get_json()
        second = self.client.post(
            "/api/customers",
            json={
                **base,
                "terminal_code": "1000000002",
                "customer_name": "乙客户",
                "route": "二号线路",
                "salesperson": "罗伟",
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(second.status_code, 201)

        intersection = self.client.get(
            "/api/customers?terminal_code=1000000001&salesperson=罗伟"
        ).get_json()
        self.assertEqual(intersection["total"], 0)
        matching_intersection = self.client.get(
            "/api/customers?route=二号线路&salesperson=罗伟"
        ).get_json()
        self.assertEqual(matching_intersection["total"], 1)
        self.assertEqual(
            matching_intersection["items"][0]["terminal_code"],
            "1000000002",
        )
        route_filtered = self.client.get(
            "/api/customers?route=二号线路"
        ).get_json()
        self.assertEqual(route_filtered["total"], 1)
        self.assertEqual(
            route_filtered["items"][0]["terminal_code"],
            "1000000002",
        )
        options = self.client.get("/api/customers/options").get_json()
        self.assertEqual(options["routes"], ["一号线路", "二号线路"])
        updated = self.client.patch(
            f"/api/customers/{first_data['id']}",
            json={
                **base,
                "terminal_code": "1000000001",
                "customer_name": "甲客户（修改）",
                "version": first_data["version"],
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(updated.status_code, 200)
        logs = self.client.get(
            f"/api/customers/{first_data['id']}/logs"
        ).get_json()["items"]
        self.assertEqual([item["action_type"] for item in logs], ["update", "create"])

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "终端编码",
                "客户全名",
                "终端业态",
                "状态",
                "线路归属",
                "业务员",
                "雪花业务员",
                "客户联系人",
                "客户地址",
                "客户手机",
                "备注",
            ]
        )
        worksheet.append(
            ["1000000003", "丙客户", "便利店", "", "二号线路", "韦春云", "陈俊杰", "", "", "", ""]
        )
        worksheet.append(
            ["1000000004", "错误客户", "商超", "运营", "二号线路", "名单外", "陈俊杰", "", "", "", ""]
        )
        worksheet.append(
            ["1000000003", "重复客户", "便利店", "运营", "二号线路", "韦春云", "陈俊杰", "", "", "", ""]
        )
        excel = io.BytesIO()
        workbook.save(excel)
        excel.seek(0)
        imported = self.client.post(
            "/api/customers/import",
            data={"file": (excel, "customers.xlsx")},
            headers={"X-CSRF-Token": csrf_token},
            content_type="multipart/form-data",
        )
        self.assertEqual(imported.status_code, 200)
        imported_data = imported.get_json()
        self.assertEqual(imported_data["success_count"], 1)
        self.assertEqual(imported_data["failed_count"], 2)
        report = self.client.get(imported_data["error_report_url"])
        self.assertEqual(report.status_code, 200)

        with self.client.session_transaction() as current_session:
            current_session["role"] = "user"
        forbidden_delete = self.client.delete(
            f"/api/customers/{first_data['id']}",
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(forbidden_delete.status_code, 403)
        forbidden_import = self.client.post(
            "/api/customers/import",
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(forbidden_import.status_code, 403)

    def test_regular_user_can_preview_and_import_snow_outbound(self):
        self.login()
        session_data = self.client.get("/api/session").get_json()
        csrf_token = session_data["csrf_token"]
        with self.client.session_transaction() as current_session:
            current_session["role"] = "user"

        policy = self.client.post(
            "/api/snow-outbound/policies",
            json={
                "name": "旺季套餐",
                "outbound_code": "PLX260001001939",
                "explanation": "旺季套餐陈列政策",
                "requires_photo": True,
                "set_limit": 10,
                "month_target": 30,
                "year": 2026,
                "month": 7,
                "conditions": [
                    {
                        "field": "outbound_remark",
                        "operator": "contains",
                        "value": "PLX260001001939",
                    }
                ],
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(policy.status_code, 201)
        policy_data = policy.get_json()
        self.assertEqual(policy_data["display_name"], "7月-旺季套餐")
        self.assertEqual(policy_data["month_target"], 30)
        forbidden_delete = self.client.delete(
            f"/api/snow-outbound/policies/{policy_data['id']}",
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(forbidden_delete.status_code, 403)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "票号",
                "开票日期",
                "业务员",
                "对象编码",
                "对象名称",
                "地址",
                "电话号码",
                "折合箱数",
                "售卖类型",
                "出库单备注",
            ]
        )
        worksheet.append(
            [
                "XH-1",
                "2026-07-20",
                "陈俊杰",
                "1000000099",
                "自动创建客户",
                "测试地址",
                "13800000099",
                8,
                "正常",
                "前缀PLX260001001939后缀",
            ]
        )
        excel = io.BytesIO()
        workbook.save(excel)
        excel.seek(0)
        preview = self.client.post(
            "/api/snow-outbound/preview",
            data={
                "file": (excel, "snow-outbound.xlsx"),
                "update_policy": "true",
            },
            headers={"X-CSRF-Token": csrf_token},
            content_type="multipart/form-data",
        )
        self.assertEqual(preview.status_code, 200)
        preview_data = preview.get_json()
        self.assertEqual(preview_data["auto_customer_count"], 1)
        self.assertEqual(preview_data["tag_count"], 1)

        imported = self.client.post(
            "/api/snow-outbound/import",
            json={"preview_id": preview_data["preview_id"]},
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(imported.status_code, 200)
        self.assertEqual(imported.get_json()["auto_customer_count"], 1)

        customer_data = self.client.get(
            "/api/customers?terminal_code=1000000099&policy_month=2026-07"
        ).get_json()
        self.assertEqual(customer_data["total"], 1)
        customer = customer_data["items"][0]
        self.assertEqual(customer["salesperson"], "")
        self.assertEqual(customer["snow_salesperson"], "陈俊杰")
        self.assertEqual(customer["remark"], "由系统创建")
        self.assertEqual(customer["policy_tags"], ["7月-旺季套餐"])
        self.assertEqual(
            customer["policy_tag_details"][0]["policy_id"],
            policy_data["id"],
        )
        policy_options = self.client.get(
            "/api/customers/policy-options?month=2026-07"
        )
        self.assertEqual(policy_options.status_code, 200)
        self.assertIn(
            "7月-旺季套餐",
            policy_options.get_json()["items"],
        )
        filtered_customers = self.client.get(
            "/api/customers?policy_month=2026-07"
            "&policy_tag=7月-旺季套餐"
        ).get_json()
        self.assertEqual(filtered_customers["total"], 1)
        self.assertEqual(
            filtered_customers["items"][0]["terminal_code"],
            "1000000099",
        )
        empty_policy_filter = self.client.get(
            "/api/customers?policy_month=2026-07&policy_tag=不存在"
        ).get_json()
        self.assertEqual(empty_policy_filter["total"], 0)

    def test_snow_policy_can_export_reimbursement_workbook(self):
        self.login()
        csrf_token = self.client.get("/api/session").get_json()["csrf_token"]

        def create_product(code, short_name, full_name, settlement_price):
            response = self.client.post(
                "/api/products",
                json={
                    "product_codes": [code],
                    "short_name": short_name,
                    "product_name": full_name,
                    "snow_inventory": 10,
                    "housekeeper_codes": [f"GJ-{code}"],
                    "specification": 12,
                    "auxiliary_unit": "瓶",
                    "settlement_price": settlement_price,
                },
                headers={"X-CSRF-Token": csrf_token},
            )
            self.assertEqual(response.status_code, 201)
            return response.get_json()

        priced_product = create_product(
            "EXPORT-001",
            "核销产品",
            "核销产品商品全名",
            12.5,
        )
        blank_price_product = create_product(
            "EXPORT-002",
            "无结算价产品",
            "无结算价产品商品全名",
            None,
        )
        policy_response = self.client.post(
            "/api/snow-outbound/policies",
            json={
                "name": "导出测试",
                "outbound_code": "PLX-EXPORT-001",
                "explanation": "核销明细导出测试",
                "requires_photo": True,
                "set_limit": 10,
                "year": 2026,
                "month": 7,
                "gift_type": "陈列赠酒",
                "normal_sale_product_ids": [priced_product["id"]],
                "gift_product_ids": [
                    priced_product["id"],
                    blank_price_product["id"],
                ],
                "conditions": [
                    {
                        "field": "outbound_remark",
                        "operator": "contains",
                        "value": "PLX-EXPORT-001",
                    }
                ],
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(policy_response.status_code, 201)
        policy = policy_response.get_json()
        empty_policy_response = self.client.post(
            "/api/snow-outbound/policies",
            json={
                "name": "无核销数据",
                "outbound_code": "PLX-EXPORT-NONE",
                "explanation": "用于验证核销排序",
                "requires_photo": True,
                "set_limit": 10,
                "year": 2026,
                "month": 7,
                "gift_type": "陈列赠酒",
                "normal_sale_product_ids": [priced_product["id"]],
                "gift_product_ids": [priced_product["id"]],
                "conditions": [
                    {
                        "field": "outbound_remark",
                        "operator": "contains",
                        "value": "PLX-EXPORT-NONE",
                    }
                ],
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(empty_policy_response.status_code, 201)
        empty_policy = empty_policy_response.get_json()

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "票号",
                "开票日期",
                "业务员",
                "对象编码",
                "对象名称",
                "折合箱数",
                "售卖类型",
                "出库单备注",
                "商品简称",
            ]
        )
        sheet.append(
            [
                "CK20260701-EXPORT-1",
                "2026-07-01",
                "陈俊杰",
                "1000001001",
                "核销客户甲",
                2,
                "陈列赠酒",
                "前缀 PLX-EXPORT-001 后缀",
                "核销产品商品全名",
            ]
        )
        sheet.append(
            [
                "CK20260701-EXPORT-2",
                "2026-07-01",
                "陈俊杰",
                "1000001002",
                "核销客户乙",
                3,
                "陈列赠酒",
                "PLX-EXPORT-001",
                "无结算价产品商品全名",
            ]
        )
        sheet.append(
            [
                "CK20260701-WRONG-TYPE",
                "2026-07-01",
                "陈俊杰",
                "1000001003",
                "售卖类型不匹配",
                4,
                "促销赠酒-临时搭赠",
                "PLX-EXPORT-001",
                "核销产品商品全名",
            ]
        )
        sheet.append(
            [
                "CK20260701-NEGATIVE",
                "2026-07-01",
                "陈俊杰",
                "1000001004",
                "负数行",
                -1,
                "陈列赠酒",
                "PLX-EXPORT-001",
                "核销产品商品全名",
            ]
        )
        source = io.BytesIO()
        workbook.save(source)
        source.seek(0)
        preview = self.client.post(
            "/api/snow-outbound/preview",
            data={
                "file": (source, "policy-export-source.xlsx"),
                "update_policy": "true",
            },
            headers={"X-CSRF-Token": csrf_token},
            content_type="multipart/form-data",
        )
        self.assertEqual(preview.status_code, 200)
        imported = self.client.post(
            "/api/snow-outbound/import",
            json={"preview_id": preview.get_json()["preview_id"]},
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(imported.status_code, 200)

        listed_policies = self.client.get(
            "/api/snow-outbound/policies?year=2026&month=7"
        ).get_json()["items"]
        listed_policy = next(
            item for item in listed_policies if item["id"] == policy["id"]
        )
        self.assertEqual(listed_policy["reimbursement_quantity"], 5)
        self.assertEqual(listed_policy["reimbursement_amount"], 25)
        amount_desc = self.client.get(
            "/api/snow-outbound/policies?year=2026&month=7"
            "&sort_by=reimbursement_amount&sort_order=desc"
        ).get_json()["items"]
        self.assertEqual(amount_desc[0]["id"], policy["id"])
        amount_asc = self.client.get(
            "/api/snow-outbound/policies?year=2026&month=7"
            "&sort_by=reimbursement_amount&sort_order=asc"
        ).get_json()["items"]
        self.assertEqual(amount_asc[0]["id"], empty_policy["id"])
        shipped_desc = self.client.get(
            "/api/snow-outbound/policies?year=2026&month=7"
            "&sort_by=shipped_count&sort_order=desc"
        ).get_json()["items"]
        self.assertEqual(shipped_desc[0]["id"], policy["id"])
        pending_sort = self.client.get(
            "/api/snow-outbound/policies?year=2026&month=7"
            "&sort_by=pending_outbound_count&sort_order=desc"
        )
        self.assertEqual(pending_sort.status_code, 200)
        invalid_sort = self.client.get(
            "/api/snow-outbound/policies?sort_by=unknown"
        )
        self.assertEqual(invalid_sort.status_code, 400)

        exported = self.client.post(
            f"/api/snow-outbound/policies/{policy['id']}/export",
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(exported.status_code, 200)
        self.assertIn(
            "spreadsheetml.sheet",
            exported.headers["Content-Type"],
        )
        exported_book = load_workbook(io.BytesIO(exported.data), data_only=False)
        exported_sheet = exported_book["Sheet1"]
        self.assertEqual(
            [exported_sheet.cell(2, column).value for column in range(1, 9)],
            [
                "序号",
                "对象编码",
                "对象名称",
                "产品",
                "数量",
                "单价",
                "核销金额",
                "是否达标",
            ],
        )
        self.assertEqual(exported_sheet["B3"].value, "1000001001")
        self.assertEqual(exported_sheet["D3"].value, "核销产品商品全名")
        self.assertEqual(exported_sheet["F3"].value, 12.5)
        self.assertEqual(exported_sheet["G3"].value, "=E3*F3")
        self.assertEqual(exported_sheet["B4"].value, "1000001002")
        self.assertIsNone(exported_sheet["F4"].value)
        self.assertIsNone(exported_sheet["G4"].value)
        self.assertEqual(exported_sheet["H4"].value, "是")
        self.assertEqual(exported_sheet["D5"].value, "合计")
        self.assertEqual(exported_sheet["E5"].value, "=SUM(E3:E4)")
        self.assertEqual(exported_sheet["A6"].value, "业务员签字：")
        self.assertEqual(exported_sheet["D6"].value, "业务部经理签字：")

    def test_single_extract_creates_success_and_failure_records(self):
        with self.client.session_transaction() as current_session:
            current_session["user"] = "team"
            current_session["display_name"] = "测试管理员"
            current_session["csrf_token"] = "test-token"

        output_dir = Path(self.output.name) / "单链接终端"
        output_dir.mkdir(parents=True, exist_ok=True)
        image_path = output_dir / "1000000001_单链接终端_业务员_01.jpg"
        image_path.write_bytes(b"image")
        result = ExtractResult(
            visit_id="SINGLE-VISIT-1",
            terminal_name="单链接终端",
            partner_name="业务员",
            output_dir=str(output_dir),
            images=[
                SavedImage(
                    index=1,
                    photoid=(
                        "private/TCOS/Z0019/O50002488/20260710/"
                        "1000000001/source.jpeg"
                    ),
                    filename=image_path.name,
                    url="",
                    size_bytes=5,
                )
            ],
            metadata_file=str(output_dir / "metadata.json"),
            visit_in_time="1783660800000",
        )
        with patch.object(
            self.web,
            "extract_images",
            return_value=result,
        ), patch.object(
            self.web.IMAGE_LIBRARY,
            "add_result",
            return_value=1,
        ):
            response = self.client.post(
                "/api/extract",
                json={"url": "https://crm.example/visit?id=1"},
                headers={"X-CSRF-Token": "test-token"},
            )
        self.assertEqual(response.status_code, 200)
        success = self.client.get("/api/extraction-records").get_json()["items"][0]
        self.assertEqual(success["owner_display_name"], "测试管理员")
        self.assertEqual(success["method"], "single_link")
        self.assertEqual(success["status"], "success")
        self.assertEqual(success["image_count"], 1)
        self.assertEqual(success["terminal_count"], 1)

        with patch.object(
            self.web,
            "extract_images",
            side_effect=self.web.CrmApiError(
                "接口失败 https://crm.example/visit?token=secret"
            ),
        ):
            failed_response = self.client.post(
                "/api/extract",
                json={"url": "https://crm.example/visit?id=2"},
                headers={"X-CSRF-Token": "test-token"},
            )
        self.assertEqual(failed_response.status_code, 400)
        failed = self.client.get("/api/extraction-records").get_json()["items"][0]
        self.assertEqual(failed["status"], "failed")
        self.assertEqual(failed["image_count"], 0)
        self.assertEqual(failed["terminal_count"], 0)
        self.assertNotIn("secret", failed["error_information"])
        self.assertIn("[链接已隐藏]", failed["error_information"])

    def test_user_management_requires_admin_role(self):
        with self.client.session_transaction() as current_session:
            current_session["user"] = "worker"
            current_session["role"] = "user"
            current_session["csrf_token"] = "test-token"

        self.assertEqual(self.client.get("/api/users").status_code, 403)

    def test_security_headers(self):
        response = self.client.get("/healthz")
        self.assertEqual(response.headers["X-Frame-Options"], "DENY")
        self.assertEqual(response.headers["X-Content-Type-Options"], "nosniff")
        self.assertIn("default-src 'self'", response.headers["Content-Security-Policy"])

    def test_batch_extract_builds_downloadable_zip(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["链接"])
        link = (
            "https://crm.example/visitDetail"
            "?appuser=u&id=954187FD1234&process_type=p"
        )
        worksheet.append([link])
        worksheet.append([link])
        worksheet.append([link.replace("954187FD1234", "A343379C1234")])
        worksheet.append([link.replace("954187FD1234", "B453379C1234")])
        worksheet.append(["https://example.com/not-a-crm-link"])
        excel = io.BytesIO()
        workbook.save(excel)
        excel.seek(0)

        with self.client.session_transaction() as current_session:
            current_session["user"] = "team"
            current_session["csrf_token"] = "test-token"

        def fake_extract(_url, output_root):
            field = (
                "2045678901"
                if "B453379C1234" in _url
                else "1023275022"
            )
            output_dir = Path(output_root) / "测试终端_954187FD"
            output_dir.mkdir(parents=True, exist_ok=True)
            filename = f"{field}_测试终端_测试业务员_01.jpg"
            (output_dir / filename).write_bytes(b"image")
            return ExtractResult(
                visit_id="954187FD1234",
                terminal_name="测试终端",
                partner_name="测试业务员",
                output_dir=str(output_dir),
                images=[
                    SavedImage(
                        index=1,
                        photoid=(
                            "private/TCOS/Z0019/O50002488/20260610/"
                            f"{field}/source.jpeg"
                        ),
                        filename=filename,
                        url="",
                        size_bytes=5,
                    )
                ],
                metadata_file=str(output_dir / "metadata.json"),
                visit_in_time="1782714405357",
            )

        with patch.object(
            self.web,
            "extract_images",
            side_effect=fake_extract,
        ) as extract_images, patch.object(
            self.web.IMAGE_LIBRARY,
            "add_result",
            side_effect=[1, 0, 1],
        ):
            response = self.client.post(
                "/api/batch-extract",
                data={"file": (excel, "links.xlsx")},
                headers={"X-CSRF-Token": "test-token"},
                content_type="multipart/form-data",
            )
            self.assertEqual(response.status_code, 202)
            started = response.get_json()
            self.assertEqual(started["status"], "queued")
            self.assertEqual(started["total"], 3)
            self.assertEqual(started["input_count"], 5)
            self.assertEqual(started["duplicate_count"], 1)
            self.assertEqual(started["invalid_count"], 1)
            self.assertEqual(started["rejected_count"], 2)
            self.assertEqual(started["chunk_index"], 1)
            self.assertEqual(started["chunk_count"], 1)

            for _attempt in range(200):
                status_response = self.client.get(
                    f"/api/batch-extract/{started['job_id']}"
                )
                self.assertEqual(status_response.status_code, 200)
                job = status_response.get_json()
                if job["status"] in {"completed", "failed"}:
                    break
                time.sleep(0.01)
            else:
                self.fail("批量任务未在预期时间内完成")

        self.assertEqual(job["status"], "completed")
        self.assertEqual(job["processed"], 3)
        data = job["result"]
        self.assertEqual(extract_images.call_count, 3)
        self.assertEqual(data["total"], 3)
        self.assertEqual(data["input_count"], 5)
        self.assertEqual(data["duplicate_count"], 1)
        self.assertEqual(data["invalid_count"], 1)
        self.assertEqual(data["rejected_count"], 2)
        self.assertEqual(data["retry_count"], 0)
        self.assertEqual(data["succeeded"], 3)
        self.assertEqual(data["image_count"], 3)
        self.assertRegex(
            data["archive_name"],
            r"^\d{8}_测试业务员_2\.zip$",
        )
        self.assertEqual(
            data["field_rows"],
            [
                {"row": 2, "field": "1023275022"},
                {"row": 5, "field": "2045678901"},
            ],
        )
        archive_path = Path(self.output.name) / "_batches" / data["archive_name"]
        self.assertTrue(archive_path.is_file())
        with zipfile.ZipFile(archive_path) as archive:
            names = archive.namelist()
        self.assertIn("提取结果.json", names)
        self.assertIn("01_1023275022_测试终端/01.jpg", names)
        self.assertIn("01_1023275022_测试终端/02.jpg", names)
        self.assertIn("02_2045678901_测试终端/01.jpg", names)
        self.assertEqual(
            {
                name.rsplit("/", 1)[0]
                for name in names
                if name.endswith(".jpg")
            },
            {
                "01_1023275022_测试终端",
                "02_2045678901_测试终端",
            },
        )
        checkpoint = self.web._batch_checkpoint_path(started["job_id"])
        self.assertTrue(checkpoint.is_file())
        checkpoint_data = json.loads(checkpoint.read_text(encoding="utf-8"))
        self.assertEqual(checkpoint_data["status"], "completed")
        self.assertEqual(checkpoint_data["processed"], 3)
        extraction_record = self.client.get("/api/extraction-records").get_json()[
            "items"
        ][0]
        self.assertEqual(extraction_record["method"], "batch")
        self.assertEqual(extraction_record["status"], "partial_success")
        self.assertEqual(extraction_record["image_count"], 2)
        self.assertEqual(extraction_record["terminal_count"], 2)
        self.assertIn("无效链接 1 条", extraction_record["error_information"])

    def test_batch_excel_accepts_500_links_and_rejects_501(self):
        def build_excel(count):
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["链接"])
            for index in range(count):
                worksheet.append(
                    [
                        "https://crm.example/visitDetail"
                        f"?appuser=u&id={index:012d}&process_type=p"
                    ]
                )
            excel = io.BytesIO()
            workbook.save(excel)
            excel.seek(0)
            return excel

        with patch.object(self.web, "MAX_BATCH_LINKS", 500):
            links, stats = self.web._parse_excel_links(build_excel(500))
            self.assertEqual(len(links), 500)
            self.assertEqual(stats["input_count"], 500)
            with self.assertRaisesRegex(ValueError, "单次最多处理 500 条链接"):
                self.web._parse_excel_links(build_excel(501))

    def test_batch_link_retries_and_records_retry_count(self):
        image_path = Path(self.output.name) / "retry-image.jpg"
        image_path.write_bytes(b"image")
        job_id = "retry-job"
        now = time.time()
        self.web._register_batch_job(
            job_id,
            {
                "owner": "team",
                "status": "queued",
                "processed": 0,
                "total": 1,
                "succeeded": 0,
                "failed": 0,
                "image_count": 0,
                "retry_count": 0,
                "links": [[2, "https://crm.example/visitDetail?appuser=u&id=1&process_type=p"]],
                "completed_records": [],
                "errors": [],
                "input_count": 1,
                "duplicate_count": 0,
                "invalid_count": 0,
                "rejected_count": 0,
                "created_at": now,
                "updated_at": now,
            },
        )
        record = {
            "row": 2,
            "terminal_name": "重试终端",
            "partner_name": "测试业务员",
            "images": [{"field": "1000000001", "source": str(image_path)}],
        }
        with patch.object(self.web, "BATCH_LINK_ATTEMPTS", 2), patch.object(
            self.web,
            "_extract_batch_record",
            side_effect=[self.web.CrmApiError("临时失败"), record],
        ) as extract_record:
            should_continue = self.web._run_batch_job_chunk(self.web.app, job_id)

        self.assertFalse(should_continue)
        self.assertEqual(extract_record.call_count, 2)
        job = self.web.BATCH_JOBS[job_id]
        self.assertEqual(job["status"], "completed")
        self.assertEqual(job["retry_count"], 1)
        self.assertEqual(job["result"]["retry_count"], 1)

    def test_system_checkpoints_are_not_downloadable(self):
        with self.client.session_transaction() as current_session:
            current_session["user"] = "team"
        response = self.client.get("/output/_system/batch_jobs/private.json")
        self.assertEqual(response.status_code, 404)

    def test_batch_extract_rejects_wrong_header(self):
        workbook = Workbook()
        workbook.active.append(["网址"])
        workbook.active.append(["https://example.com"])
        excel = io.BytesIO()
        workbook.save(excel)
        excel.seek(0)

        with self.client.session_transaction() as current_session:
            current_session["user"] = "team"
            current_session["csrf_token"] = "test-token"
        response = self.client.post(
            "/api/batch-extract",
            data={"file": (excel, "links.xlsx")},
            headers={"X-CSRF-Token": "test-token"},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("字段名为“链接”", response.get_json()["error"])
        record = self.client.get("/api/extraction-records").get_json()["items"][0]
        self.assertEqual(record["status"], "failed")
        self.assertIn("字段名为“链接”", record["error_information"])

    def test_distribution_summary_and_business_archive(self):
        output_dir = (
            Path(self.output.name)
            / "测试业务员"
            / "测试终端_VISIT001"
        )
        output_dir.mkdir(parents=True)
        photoid = (
            "private/TCOS/Z0019/O50002488/20260610/"
            "1023275022/source.jpeg"
        )
        filename = "1023275022_测试终端_测试业务员_01.jpeg"
        (output_dir / filename).write_bytes(b"image")
        (output_dir / "metadata.json").write_text(
            json.dumps(
                {
                    "visit_id": "VISIT001",
                    "terminal_name": "测试终端",
                    "partner_name": "测试业务员",
                    "images": [
                        {
                            "photoid": photoid,
                            "filename": filename,
                            "size_bytes": 5,
                        }
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        job, _duplicate = self.web.DISTRIBUTION_STORE.enqueue(
            job_id="job-1",
            url=(
                "https://crm.example/visitDetail"
                "?appuser=u&id=VISIT001&process_type=p"
            ),
        )
        self.web.DISTRIBUTION_STORE.complete(
            job.id,
            ExtractResult(
                visit_id="VISIT001",
                terminal_name="测试终端",
                partner_name="测试业务员",
                output_dir=str(output_dir),
                images=[
                    SavedImage(
                        index=1,
                        photoid=photoid,
                        filename=filename,
                        url="",
                        size_bytes=5,
                    )
                ],
                metadata_file=str(output_dir / "metadata.json"),
            ),
        )

        with self.client.session_transaction() as current_session:
            current_session["user"] = "team"
            current_session["csrf_token"] = "test-token"

        summary = self.client.get("/api/distributions")
        self.assertEqual(summary.status_code, 200)
        item = summary.get_json()["items"][0]
        self.assertEqual(item["business"], "测试业务员")
        self.assertEqual(item["quantity"], 1)
        self.assertEqual(item["field_values"], ["1023275022"])
        self.assertEqual(item["distributed_count"], 1)
        self.assertEqual(item["pending_download_count"], 1)

        missing_csrf = self.client.post(
            "/api/distributions/测试业务员/archive"
        )
        self.assertEqual(missing_csrf.status_code, 403)
        archive_response = self.client.post(
            "/api/distributions/测试业务员/archive",
            headers={"X-CSRF-Token": "test-token"},
        )
        self.assertEqual(archive_response.status_code, 200)
        archive_data = archive_response.get_json()
        archive_path = (
            Path(self.output.name)
            / "_distribution_downloads"
            / archive_data["archive_name"]
        )
        self.assertTrue(archive_path.is_file())
        with zipfile.ZipFile(archive_path) as archive:
            names = archive.namelist()
        self.assertIn("分发提取结果.json", names)
        self.assertIn(
            "01_1023275022_测试终端/01.jpeg",
            names,
        )
        refreshed = self.client.get("/api/distributions").get_json()
        self.assertEqual(
            refreshed["items"][0]["pending_download_count"],
            0,
        )

        missing_clear_csrf = self.client.delete("/api/distributions")
        self.assertEqual(missing_clear_csrf.status_code, 403)
        clear_response = self.client.delete(
            "/api/distributions",
            headers={"X-CSRF-Token": "test-token"},
        )
        self.assertEqual(clear_response.status_code, 200)
        self.assertEqual(clear_response.get_json()["deleted_count"], 1)
        self.assertEqual(
            self.client.get("/api/distributions").get_json()["items"],
            [],
        )

    def test_image_library_search_archive_and_policy_export(self):
        output_dir = Path(self.output.name) / "测试终端_VISITLIB"
        output_dir.mkdir(parents=True)
        photoid = (
            "private/TCOS/Z0019/O50002488/20260610/"
            "1023275022/source.jpeg"
        )
        filename = "1023275022_测试终端_测试业务员_01.jpeg"
        (output_dir / filename).write_bytes(b"image")
        self.web.IMAGE_LIBRARY.add_result(
            ExtractResult(
                visit_id="VISITLIB",
                terminal_name="测试终端",
                partner_name="测试业务员",
                output_dir=str(output_dir),
                images=[
                    SavedImage(
                        index=1,
                        photoid=photoid,
                        filename=filename,
                        url="",
                        size_bytes=5,
                    )
                ],
                metadata_file=str(output_dir / "metadata.json"),
                visit_in_time="1782714405357",
            ),
            created_at="2026-07-07T09:00:00",
        )

        with self.client.session_transaction() as current_session:
            current_session["user"] = "team"
            current_session["csrf_token"] = "test-token"

        search = self.client.post(
            "/api/image-library/search",
            json={"fields": "1023275022\n9999999999", "month": "2026-06"},
        )
        self.assertEqual(search.status_code, 200)
        data = search.get_json()
        self.assertEqual(data["field_count"], 1)
        self.assertEqual(data["image_count"], 1)
        self.assertEqual(data["missing_fields"], ["9999999999"])
        self.assertEqual(data["pagination"]["page"], 1)
        self.assertEqual(data["pagination"]["total_groups"], 1)
        image_data = data["items"][0]["images"][0]
        image_id = image_data["id"]
        thumbnail = self.client.get(image_data["thumbnail_url"])
        self.assertEqual(thumbnail.status_code, 200)
        self.assertIn("private", thumbnail.headers["Cache-Control"])
        self.assertIn("immutable", thumbnail.headers["Cache-Control"])
        thumbnail.close()

        with patch.object(self.web, "X_ACCEL_ENABLED", True):
            accelerated_thumbnail = self.client.get(image_data["thumbnail_url"])
            self.assertEqual(accelerated_thumbnail.status_code, 200)
            self.assertTrue(
                accelerated_thumbnail.headers["X-Accel-Redirect"].startswith(
                    "/_protected_media/"
                )
            )
            self.assertEqual(
                accelerated_thumbnail.headers["X-Accel-Expires"],
                str(self.web.IMAGE_CACHE_SECONDS),
            )
            self.assertEqual(accelerated_thumbnail.get_data(), b"")

            accelerated_original = self.client.get(image_data["url"])
            self.assertEqual(accelerated_original.status_code, 200)
            self.assertTrue(
                accelerated_original.headers["X-Accel-Redirect"].startswith(
                    "/_protected_media/_image_library/"
                )
            )
            self.assertIn("private", accelerated_original.headers["Cache-Control"])

        bad_pagination = self.client.get("/api/image-library?page=invalid")
        self.assertEqual(bad_pagination.status_code, 400)

        policy = self.web.SNOW_OUTBOUND_STORE.create_policy(
            {
                "name": "测试陈列",
                "outbound_code": "PLX260000000001",
                "explanation": "测试照片归档",
                "requires_photo": True,
                "set_limit": 1,
                "month_target": 2,
                "year": 2026,
                "month": 6,
                "conditions": [
                    {
                        "field": "outbound_remark",
                        "operator": "equals",
                        "value": "PLX260000000001",
                    }
                ],
            },
            operator="team",
            operator_name="测试用户",
        )
        customer_payload = {
            "status": "运营",
            "route": "",
            "salesperson": "",
            "snow_salesperson": "",
            "contact": "",
            "address": "",
            "phone": "",
            "remark": "",
        }
        self.web.CUSTOMER_STORE.create_customer(
            {
                **customer_payload,
                "terminal_code": "1023275022",
                "customer_name": "测试终端",
            },
            operator="team",
            operator_name="测试用户",
        )
        self.web.CUSTOMER_STORE.create_customer(
            {
                **customer_payload,
                "terminal_code": "1023275023",
                "customer_name": "缺失终端",
                "salesperson": "韦春云",
            },
            operator="team",
            operator_name="测试用户",
        )
        with self.web.SNOW_OUTBOUND_STORE._connect() as connection:
            connection.execute(
                """
                INSERT INTO snow_outbound_imports (
                    id, filename, operator, operator_name, rules_json,
                    months_json, row_count, ticket_count, terminal_count,
                    tag_count, auto_customer_count, created_at
                ) VALUES (
                    'IMPORT-ARCHIVE', 'archive.xlsx', 'team', '测试用户', '{}',
                    '["2026-06"]', 2, 2, 2, 2, 0, '2026-06-30T10:00:00'
                )
                """
            )
            for terminal_code in ("1023275022", "1023275023"):
                connection.execute(
                    """
                    INSERT INTO customer_policy_tags (
                        month, terminal_code, tag, matched_ticket_no,
                        matched_row_number, import_id, created_at,
                        policy_id, color, rule_snapshot_json
                    ) VALUES (
                        '2026-06', ?, '测试陈列', ?, 1, 'IMPORT-ARCHIVE',
                        '2026-06-30T10:00:00', ?, 'blue', '{}'
                    )
                    """,
                    (terminal_code, f"T-{terminal_code}", policy["id"]),
                )

        pending_filename = "1023275099_待出库终端_测试业务员_01.jpeg"
        (output_dir / pending_filename).write_bytes(b"pending-image")
        self.web.IMAGE_LIBRARY.add_result(
            ExtractResult(
                visit_id="VISIT-PENDING",
                terminal_name="待出库终端",
                partner_name="测试业务员",
                output_dir=str(output_dir),
                images=[
                    SavedImage(
                        index=1,
                        photoid=(
                            "private/TCOS/Z0019/O50002488/20260610/"
                            "1023275099/source.jpeg"
                        ),
                        filename=pending_filename,
                        url="",
                        size_bytes=13,
                    )
                ],
                metadata_file=str(output_dir / "pending-metadata.json"),
                visit_in_time="1782714405357",
            ),
            created_at="2026-07-07T09:30:00",
        )
        pending_image_id = self.web.IMAGE_LIBRARY.query(
            fields=["1023275099"],
            month="2026-06",
        )["items"][0]["images"][0]["id"]

        no_photo_policy = self.web.SNOW_OUTBOUND_STORE.create_policy(
            {
                "name": "无需拍照政策",
                "outbound_code": "PLX260000000002",
                "explanation": "不应出现在照片归档枚举中",
                "requires_photo": False,
                "set_limit": 1,
                "month_target": 1,
                "year": 2026,
                "month": 6,
                "conditions": [
                    {
                        "field": "outbound_remark",
                        "operator": "equals",
                        "value": "PLX260000000002",
                    }
                ],
            },
            operator="team",
            operator_name="测试用户",
        )
        options = self.client.get(
            "/api/photo-archive/options?month=2026-06"
        ).get_json()["items"]
        self.assertEqual([item["id"] for item in options], [policy["id"]])

        rejected_no_photo_archive = self.client.post(
            "/api/photo-archive",
            json={
                "image_ids": [pending_image_id],
                "policy_id": no_photo_policy["id"],
                "month": "2026-06",
            },
            headers={"X-CSRF-Token": "test-token"},
        )
        self.assertEqual(rejected_no_photo_archive.status_code, 400)
        self.assertIn(
            "无需拍照",
            rejected_no_photo_archive.get_json()["error"],
        )

        policy_search = self.client.post(
            "/api/image-library/search",
            json={
                "month": "2026-06",
                "policy_ids": [policy["id"]],
                "businesses": ["测试业务员", "不存在的业务员"],
            },
        )
        self.assertEqual(policy_search.status_code, 200)
        policy_search_data = policy_search.get_json()
        self.assertEqual(policy_search_data["image_count"], 1)
        tagged_item = next(
            item
            for item in policy_search_data["items"]
            if item["field"] == "1023275022"
        )
        self.assertEqual(
            tagged_item["policy_tags"],
            [
                {
                    "color": "blue",
                    "policy_id": policy["id"],
                    "tag": "测试陈列",
                }
            ],
        )
        self.assertEqual(
            [item["id"] for item in policy_search_data["policy_options"]],
            [policy["id"], no_photo_policy["id"]],
        )
        excluded_policy_search = self.client.post(
            "/api/image-library/search",
            json={
                "month": "2026-06",
                "policy_ids": [policy["id"]],
                "policy_match": "exclude",
            },
        )
        self.assertEqual(excluded_policy_search.status_code, 200)
        excluded_data = excluded_policy_search.get_json()
        self.assertEqual(excluded_data["image_count"], 1)
        self.assertEqual(
            [item["field"] for item in excluded_data["items"]],
            ["1023275099"],
        )
        invalid_policy_match = self.client.post(
            "/api/image-library/search",
            json={
                "month": "2026-06",
                "policy_ids": [policy["id"]],
                "policy_match": "invalid",
            },
        )
        self.assertEqual(invalid_policy_match.status_code, 400)
        invalid_policy_search = self.client.post(
            "/api/image-library/search",
            json={"month": "2026-06", "policy_ids": ["POL-NOT-FOUND"]},
        )
        self.assertEqual(invalid_policy_search.status_code, 400)

        missing_csrf = self.client.post(
            "/api/photo-archive",
            json={
                "image_ids": [image_id],
                "policy_id": policy["id"],
                "month": "2026-06",
            },
        )
        self.assertEqual(missing_csrf.status_code, 403)

        archived = self.client.post(
            "/api/photo-archive",
            json={
                "image_ids": [image_id],
                "policy_id": policy["id"],
                "month": "2026-06",
            },
            headers={"X-CSRF-Token": "test-token"},
        )
        self.assertEqual(archived.status_code, 200)
        self.assertEqual(archived.get_json()["archived_count"], 1)
        self.assertEqual(archived.get_json()["skipped_count"], 0)

        pending_archived = self.client.post(
            "/api/photo-archive",
            json={
                "image_ids": [pending_image_id],
                "policy_id": policy["id"],
                "month": "2026-06",
            },
            headers={"X-CSRF-Token": "test-token"},
        )
        self.assertEqual(pending_archived.status_code, 200)
        self.assertEqual(pending_archived.get_json()["archived_count"], 1)

        archived_search = self.client.post(
            "/api/image-library/search",
            json={
                "fields": ["1023275022", "1023275099"],
                "month": "2026-06",
            },
        )
        self.assertEqual(archived_search.status_code, 200)
        archived_images = {
            item["field"]: item["images"][0]
            for item in archived_search.get_json()["items"]
        }
        expected_archive_tag = [
            {
                "color": policy["color"],
                "policy_id": policy["id"],
                "tag": "6月-测试陈列",
            }
        ]
        self.assertEqual(
            archived_images["1023275022"]["archive_tags"],
            expected_archive_tag,
        )
        self.assertEqual(
            archived_images["1023275099"]["archive_tags"],
            expected_archive_tag,
        )

        duplicate = self.client.post(
            "/api/photo-archive",
            json={
                "image_ids": [image_id],
                "policy_id": policy["id"],
                "month": "2026-06",
            },
            headers={"X-CSRF-Token": "test-token"},
        )
        self.assertEqual(duplicate.status_code, 200)
        self.assertEqual(duplicate.get_json()["archived_count"], 0)
        self.assertEqual(duplicate.get_json()["skipped_count"], 1)

        archives = self.client.get(
            "/api/photo-archive/policies?month=2026-06&page=1&page_size=20"
        )
        self.assertEqual(archives.status_code, 200)
        archive_data = archives.get_json()
        self.assertEqual(archive_data["total"], 1)
        archive_item = archive_data["items"][0]
        self.assertEqual(archive_item["shipped_count"], 2)
        self.assertEqual(archive_item["photographed_count"], 2)
        self.assertEqual(archive_item["missing_count"], 1)
        self.assertEqual(archive_item["photo_count"], 2)
        self.assertEqual(archive_item["latest_operation"]["action_type"], "archive")

        policy_list = self.client.get(
            "/api/snow-outbound/policies?year=2026&month=6"
        )
        self.assertEqual(policy_list.status_code, 200)
        self.assertEqual(
            policy_list.get_json()["items"][0]["photographed_count"],
            2,
        )
        self.assertEqual(
            policy_list.get_json()["items"][0]["pending_outbound_count"],
            1,
        )
        shipped = self.client.get(
            f"/api/snow-outbound/policies/{policy['id']}/shipped-terminals"
        )
        self.assertEqual(shipped.status_code, 200)
        self.assertEqual(shipped.get_json()["total"], 2)
        self.assertEqual(
            {
                item["terminal_code"]
                for item in shipped.get_json()["items"]
            },
            {"1023275022", "1023275023"},
        )
        photographed = self.client.get(
            f"/api/snow-outbound/policies/{policy['id']}/photographed-terminals"
        )
        self.assertEqual(photographed.status_code, 200)
        self.assertEqual(photographed.get_json()["total"], 2)
        self.assertEqual(
            {
                item["terminal_code"]
                for item in photographed.get_json()["items"]
            },
            {"1023275022", "1023275099"},
        )
        pending = self.client.get(
            f"/api/snow-outbound/policies/{policy['id']}/pending-outbound"
        )
        self.assertEqual(pending.status_code, 200)
        self.assertEqual(
            pending.get_json()["items"],
            [
                {
                    "customer_name": "待出库终端",
                    "salesperson": "测试业务员",
                    "terminal_code": "1023275099",
                }
            ],
        )

        missing = self.client.get(
            f"/api/photo-archive/policies/{policy['id']}/missing"
        )
        self.assertEqual(missing.status_code, 200)
        self.assertEqual(
            missing.get_json()["items"],
            [
                {
                    "customer_name": "缺失终端",
                    "salesperson": "韦春云",
                    "terminal_code": "1023275023",
                }
            ],
        )

        first_export = self.client.post(
            f"/api/photo-archive/policies/{policy['id']}/export",
            headers={"X-CSRF-Token": "test-token"},
        )
        self.assertEqual(first_export.status_code, 200)
        self.assertEqual(first_export.mimetype, "application/zip")
        with zipfile.ZipFile(io.BytesIO(first_export.data)) as archive:
            names = archive.namelist()
            detail_bytes = archive.read("照片档案明细.xlsx")
        first_export.close()
        self.assertTrue(
            any(name.startswith("1023275022_测试终端/") for name in names)
        )
        self.assertGreater(len(detail_bytes), 100)
        detail_workbook = load_workbook(io.BytesIO(detail_bytes), read_only=True)
        detail_headers = [
            cell.value for cell in next(detail_workbook.active.iter_rows())
        ]
        self.assertIn("终端照片数量", detail_headers)
        detail_workbook.close()

        second_export = self.client.post(
            f"/api/photo-archive/policies/{policy['id']}/export",
            headers={"X-CSRF-Token": "test-token"},
        )
        self.assertEqual(second_export.status_code, 200)
        second_export.close()
        self.assertEqual(self.client.get("/api/export-records").status_code, 404)
        self.assertEqual(
            self.client.post(
                "/api/image-library/export",
                headers={"X-CSRF-Token": "test-token"},
            ).status_code,
            404,
        )

        refreshed_archive = self.client.get(
            "/api/photo-archive/policies?month=2026-06"
        ).get_json()["items"][0]
        self.assertEqual(
            refreshed_archive["latest_operation"]["action_type"],
            "export",
        )

        refreshed = self.client.get(
            "/api/image-library?fields=1023275022&month=2026-06"
        ).get_json()
        self.assertEqual(refreshed["image_count"], 1)

        remove_url = (
            f"/api/photo-archive/images/{image_id}/policies/{policy['id']}"
        )
        self.assertEqual(self.client.delete(remove_url).status_code, 403)
        removed = self.client.delete(
            remove_url,
            headers={"X-CSRF-Token": "test-token"},
        )
        self.assertEqual(removed.status_code, 200)
        self.assertEqual(removed.get_json()["removed_count"], 1)
        after_remove = self.client.post(
            "/api/image-library/search",
            json={
                "fields": ["1023275022", "1023275099"],
                "month": "2026-06",
            },
        ).get_json()
        images_after_remove = {
            item["field"]: item["images"][0]
            for item in after_remove["items"]
        }
        self.assertEqual(
            images_after_remove["1023275022"]["archive_tags"],
            [],
        )
        self.assertEqual(
            images_after_remove["1023275099"]["archive_tags"],
            expected_archive_tag,
        )
        archive_after_remove = self.client.get(
            "/api/photo-archive/policies?month=2026-06"
        ).get_json()["items"][0]
        self.assertEqual(archive_after_remove["photo_count"], 1)
        self.assertEqual(archive_after_remove["photographed_count"], 1)
        self.assertEqual(
            archive_after_remove["latest_operation"]["action_type"],
            "unarchive",
        )
        self.assertEqual(
            self.client.delete(
                remove_url,
                headers={"X-CSRF-Token": "test-token"},
            ).status_code,
            404,
        )

    def test_product_crud_and_stock_upload(self):
        self.login()
        csrf_token = self.client.get("/api/session").get_json()["csrf_token"]
        created = self.client.post(
            "/api/products",
            json={
                "product_codes": ["P001", "P001-A"],
                "short_name": "蓝听",
                "product_name": "雪花蓝听500ml听1*12",
                "snow_inventory": 10.5,
                "housekeeper_codes": ["G001"],
                "specification": 12,
                "auxiliary_unit": "听",
                "settlement_price": 50,
            },
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(created.status_code, 201)
        product = created.get_json()
        self.assertEqual(product["status"], "正常")

        listed = self.client.get("/api/products?product_code=P001")
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(listed.get_json()["total"], 1)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            ["年月", "商品编号", "商品名称", "单位", "入库千升数", "可用（箱）"]
        )
        sheet.append(
            [
                "202607",
                "P001",
                "雪花蓝听500ml听1*12纸箱",
                "箱",
                1.5,
                20.25,
            ]
        )
        sheet.append(["202607", "PROMO", "太阳伞", "顶", 0, 5])
        stream = io.BytesIO()
        workbook.save(stream)
        preview = self.client.post(
            "/api/products/import/preview",
            data={"file": (io.BytesIO(stream.getvalue()), "stock.xlsx")},
            headers={"X-CSRF-Token": csrf_token},
            content_type="multipart/form-data",
        )
        self.assertEqual(preview.status_code, 200)
        preview_data = preview.get_json()
        self.assertEqual(preview_data["updated_count"], 1)
        self.assertEqual(preview_data["skipped_count"], 1)

        committed = self.client.post(
            "/api/products/import",
            json={"preview_id": preview_data["preview_id"]},
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(committed.status_code, 200)
        updated = self.client.get("/api/products?product_code=P001").get_json()
        self.assertEqual(updated["items"][0]["snow_inventory"], 20.25)
        self.assertTrue(updated["latest_upload_at"])
        self.assertEqual(updated["monthly_inbound_tons"], 1.5)
        self.assertEqual(updated["snow_inventory_boxes"], 20.25)

        deleted = self.client.delete(
            f"/api/products/{product['id']}",
            headers={"X-CSRF-Token": csrf_token},
        )
        self.assertEqual(deleted.status_code, 200)


if __name__ == "__main__":
    unittest.main()
