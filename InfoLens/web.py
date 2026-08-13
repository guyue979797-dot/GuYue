#!/usr/bin/env python3
"""InfoLens Web 应用。

本地开发:
    INFOLENS_AUTH_MODE=off python web.py

生产环境:
    gunicorn --bind 0.0.0.0:8000 --workers 1 --threads 8 web:app
"""

from __future__ import annotations

import hmac
import io
import json
import mimetypes
import os
import queue
import re
import secrets
import shutil
import threading
import time
import urllib.parse
import zipfile
from collections import defaultdict, deque
from datetime import datetime
from functools import wraps
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException
from flask import (
    Flask,
    Response,
    abort,
    jsonify,
    redirect,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from authlib.integrations.flask_client import OAuth
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import generate_password_hash

from infolens.crm_client import CrmApiError
from infolens.customers import (
    EXCEL_HEADERS,
    HEADER_TO_FIELD,
    SALESPEOPLE,
    SNOW_SALESPEOPLE,
    CustomerStore,
)
from infolens.distribution import DistributionStore
from infolens.extraction_records import ExtractionRecordStore
from infolens.extractor import (
    ExtractResult,
    build_image_filename,
    extract_images,
    parse_visit_url,
    photoid_name_field,
)
from infolens.image_library import ImageLibraryStore
from infolens.products import ProductStore, parse_stock_workbook
from infolens.policy_export import build_policy_reimbursement_workbook
from infolens.snow_outbound import (
    POLICY_GIFT_TYPES,
    POLICY_TAGS,
    RULE_FIELDS,
    RULE_OPERATOR_LABELS,
    SnowOutboundStore,
    parse_outbound_workbook,
)
from infolens.users import UserStore
from infolens.wecom_bot import (
    MessageDeduplicator,
    WecomBotCrypto,
    WecomBotError,
    extract_crm_urls,
    message_text,
    send_response_url,
    stream_reply,
    text_reply,
    validate_callback_timestamp,
)


ROOT = Path(__file__).resolve().parent
WEB_ROOT = ROOT / "web"
OUTPUT_ROOT = Path(os.environ.get("INFOLENS_OUTPUT_ROOT", ROOT / "output")).resolve()
POLICY_EXPORT_DISTRIBUTOR = os.environ.get(
    "INFOLENS_DISTRIBUTOR_NAME",
    "贵州鑫向晨商贸有限公司",
).strip()
AUTH_MODE = os.environ.get("INFOLENS_AUTH_MODE", "off").strip().lower()
EXTRACT_LOCK = threading.Lock()
RATE_LOCK = threading.Lock()
RATE_BUCKETS: dict[str, deque[float]] = defaultdict(deque)
BATCH_JOBS_LOCK = threading.Lock()
BATCH_JOBS: dict[str, dict] = {}
BATCH_QUEUE: queue.Queue[str] = queue.Queue()
BATCH_WORKER_LOCK = threading.Lock()
BATCH_WORKER_STARTED = False
BATCH_JOB_TTL_SECONDS = 6 * 60 * 60
MAX_BATCH_LINKS = int(os.environ.get("INFOLENS_MAX_BATCH_LINKS", "500"))
BATCH_CHUNK_SIZE = max(1, int(os.environ.get("INFOLENS_BATCH_CHUNK_SIZE", "50")))
BATCH_LINK_ATTEMPTS = max(1, int(os.environ.get("INFOLENS_BATCH_LINK_ATTEMPTS", "3")))
MAX_UPLOAD_BYTES = int(os.environ.get("INFOLENS_MAX_UPLOAD_BYTES", str(4 * 1024 * 1024)))
IMAGE_LIBRARY_PAGE_SIZE = max(
    1,
    min(50, int(os.environ.get("INFOLENS_IMAGE_LIBRARY_PAGE_SIZE", "12"))),
)
IMAGE_CACHE_SECONDS = 30 * 24 * 60 * 60
X_ACCEL_ENABLED = os.environ.get("INFOLENS_X_ACCEL_ENABLED", "false").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}
X_ACCEL_PREFIX = "/_protected_media"
WECOM_BOT_ENABLED = os.environ.get("WECOM_BOT_ENABLED", "false").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}
WECOM_BOT_MODE = os.environ.get("WECOM_BOT_MODE", "callback").strip().lower()
WECOM_BOT_CALLBACK_ENABLED = WECOM_BOT_ENABLED and WECOM_BOT_MODE == "callback"
WECOM_BOT_MAX_LINKS = int(os.environ.get("WECOM_BOT_MAX_LINKS", "10"))
WECOM_DEDUPLICATOR = MessageDeduplicator()
DISTRIBUTION_STORE = DistributionStore(
    OUTPUT_ROOT / "_system" / "distributions.sqlite3"
)
IMAGE_LIBRARY = ImageLibraryStore(
    OUTPUT_ROOT / "_system" / "image_library.sqlite3",
    OUTPUT_ROOT,
)
USER_STORE = UserStore(OUTPUT_ROOT / "_system" / "users.sqlite3")
CUSTOMER_STORE = CustomerStore(
    OUTPUT_ROOT / "_system" / "customer_profiles.sqlite3"
)
PRODUCT_STORE = ProductStore(
    OUTPUT_ROOT / "_system" / "customer_profiles.sqlite3"
)
SNOW_OUTBOUND_STORE = SnowOutboundStore(
    OUTPUT_ROOT / "_system" / "customer_profiles.sqlite3"
)
EXTRACTION_RECORD_STORE = ExtractionRecordStore(
    OUTPUT_ROOT / "_system" / "extraction_records.sqlite3"
)


def _safe_record_error(message: str, *, max_length: int = 4000) -> str:
    text = re.sub(r"https?://\S+", "[链接已隐藏]", str(message or ""))
    text = re.sub(
        r"(?i)(token|access_token|authorization|cookie)=([^\s&]+)",
        r"\1=[已隐藏]",
        text,
    )
    text = " ".join(text.split()) if "\n" not in text else "\n".join(
        " ".join(line.split()) for line in text.splitlines() if line.strip()
    )
    return text[:max_length]


def _record_owner() -> tuple[str, str]:
    username = _current_user() or ""
    display_name = str(session.get("display_name") or username)
    return username, display_name


def _complete_extraction_record(
    application: Flask,
    record_id: str,
    *,
    status: str,
    image_count: int = 0,
    terminal_count: int = 0,
    error_information: str = "",
) -> None:
    try:
        EXTRACTION_RECORD_STORE.complete_record(
            record_id,
            status=status,
            image_count=image_count,
            terminal_count=terminal_count,
            error_information=_safe_record_error(error_information),
        )
    except ValueError as exc:
        if str(exc) == "新增记录不存在":
            return
        application.logger.exception("更新新增记录失败：%s", record_id)
    except Exception:
        application.logger.exception("更新新增记录失败：%s", record_id)


def _require_production_config() -> None:
    if AUTH_MODE not in {"off", "password", "oidc", "proxy"}:
        raise RuntimeError("INFOLENS_AUTH_MODE 必须为 off、password、oidc 或 proxy")
    if os.environ.get("INFOLENS_ENV") == "production" and AUTH_MODE == "off":
        raise RuntimeError("生产环境禁止关闭鉴权")
    if AUTH_MODE == "password":
        has_super_admin_secret = bool(
            os.environ.get("INFOLENS_SUPER_ADMIN_PASSWORD_HASH")
            or os.environ.get("INFOLENS_SUPER_ADMIN_PASSWORD")
            or (
                os.environ.get("INFOLENS_USERNAME")
                and os.environ.get("INFOLENS_PASSWORD_HASH")
            )
        )
        required = ("INFOLENS_SESSION_SECRET",)
        missing = [name for name in required if not os.environ.get(name)]
        if not has_super_admin_secret:
            missing.append("INFOLENS_SUPER_ADMIN_PASSWORD_HASH")
        if missing:
            raise RuntimeError(f"密码登录缺少环境变量: {', '.join(missing)}")
    if AUTH_MODE == "oidc":
        required = (
            "INFOLENS_OIDC_METADATA_URL",
            "INFOLENS_OIDC_CLIENT_ID",
            "INFOLENS_OIDC_CLIENT_SECRET",
            "INFOLENS_SESSION_SECRET",
        )
        missing = [name for name in required if not os.environ.get(name)]
        if missing:
            raise RuntimeError(f"OIDC 登录缺少环境变量: {', '.join(missing)}")
    if WECOM_BOT_MODE not in {"callback", "long_connection"}:
        raise RuntimeError("WECOM_BOT_MODE 必须为 callback 或 long_connection")
    if WECOM_BOT_CALLBACK_ENABLED:
        required = (
            "WECOM_BOT_TOKEN",
            "WECOM_BOT_ENCODING_AES_KEY",
            "WECOM_BOT_ID",
        )
        missing = [name for name in required if not os.environ.get(name)]
        if missing:
            raise RuntimeError(f"企业微信机器人缺少环境变量: {', '.join(missing)}")


def _image_url(folder: str, filename: str) -> str:
    return "/output/" + "/".join(
        urllib.parse.quote(part) for part in (folder, filename)
    )


def _serialize_result(result: ExtractResult) -> dict:
    try:
        folder = str(Path(result.output_dir).relative_to(OUTPUT_ROOT))
    except ValueError:
        folder = Path(result.output_dir).name
    return {
        "visit_id": result.visit_id,
        "terminal_name": result.terminal_name,
        "partner_name": result.partner_name,
        "images": [
            {
                "filename": image.filename,
                "size_bytes": image.size_bytes,
                "url": _image_url(folder, image.filename),
            }
            for image in result.images
        ],
    }


def _parse_excel_links(file_stream) -> tuple[list[tuple[int, str]], dict[str, int]]:
    """读取首个工作表中唯一的“链接”列。"""
    try:
        payload = file_stream.read(MAX_UPLOAD_BYTES + 1)
        if len(payload) > MAX_UPLOAD_BYTES:
            raise ValueError("Excel 文件超过上传大小限制")
        excel_buffer = io.BytesIO(payload)
        with zipfile.ZipFile(excel_buffer) as archive:
            expanded_size = sum(item.file_size for item in archive.infolist())
            if expanded_size > 32 * 1024 * 1024:
                raise ValueError("Excel 文件解压后的内容过大")
        excel_buffer.seek(0)
        workbook = load_workbook(excel_buffer, read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError, zipfile.BadZipFile) as exc:
        if isinstance(exc, ValueError) and str(exc).startswith("Excel 文件"):
            raise
        raise ValueError("无法读取 Excel 文件，请确认文件为有效的 .xlsx 格式") from exc

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError("Excel 文件为空")

        populated_headers = [
            str(value).strip() for value in header if value is not None and str(value).strip()
        ]
        if populated_headers != ["链接"]:
            raise ValueError('Excel 第一行必须只有一个字段，字段名为“链接”')

        links: list[tuple[int, str]] = []
        seen: set[str] = set()
        duplicate_count = 0
        invalid_count = 0
        input_count = 0
        for row_number, row in enumerate(rows, start=2):
            populated = [
                value for value in row[1:] if value is not None and str(value).strip()
            ]
            if populated:
                raise ValueError(f"Excel 第 {row_number} 行包含“链接”列之外的数据")

            value = row[0] if row else None
            if value is None or not str(value).strip():
                continue
            input_count += 1
            if input_count > MAX_BATCH_LINKS:
                raise ValueError(f"单次最多处理 {MAX_BATCH_LINKS} 条链接")
            link = str(value).strip()
            if link in seen:
                duplicate_count += 1
                continue
            seen.add(link)
            try:
                parse_visit_url(link)
            except ValueError:
                invalid_count += 1
                continue
            links.append((row_number, link))

        if not links:
            raise ValueError("Excel 中没有格式有效且可处理的 CRM 链接")
        return links, {
            "input_count": input_count,
            "duplicate_count": duplicate_count,
            "invalid_count": invalid_count,
            "rejected_count": duplicate_count + invalid_count,
        }
    finally:
        workbook.close()


def _build_batch_archive(
    completed_records: list[dict],
    errors: list[dict],
    total: int,
    input_stats: dict[str, int] | None = None,
    retry_count: int = 0,
) -> dict:
    input_stats = input_stats or {}
    batch_dir = OUTPUT_ROOT / "_batches"
    batch_dir.mkdir(parents=True, exist_ok=True)
    batch_key = f"{datetime.now():%Y%m%d_%H%M%S}_{secrets.token_hex(4)}"
    archive_path = batch_dir / f".batch_{batch_key}.zip"
    completed: list[dict] = []
    field_rows: list[dict] = []
    seen_fields: set[str] = set()
    image_groups: dict[tuple[str, str], dict[str, str | int]] = {}
    image_count = 0

    with zipfile.ZipFile(
        archive_path,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=6,
    ) as archive:
        for record in completed_records:
            archived_for_visit = 0
            row_number = int(record["row"])
            terminal_name = str(record["terminal_name"])
            for image in record.get("images", []):
                field = str(image["field"])
                image_key = (field, terminal_name)
                group = image_groups.get(image_key)
                if group is None:
                    safe_field = re.sub(r'[\\/:*?"<>|]', "_", field)
                    safe_terminal = re.sub(
                        r'[\\/:*?"<>|]',
                        "_",
                        terminal_name.strip(),
                    ) or "未知终端"
                    group = {
                        "folder": (
                            f"{len(image_groups) + 1:02d}_"
                            f"{safe_field}_{safe_terminal}"
                        ),
                        "image_count": 0,
                    }
                    image_groups[image_key] = group

                source = Path(str(image["source"]))
                if not source.is_file():
                    raise ValueError(f"第 {row_number} 行的已提取图片不存在，无法恢复归档")
                extension = source.suffix.lower() or ".jpg"
                group["image_count"] = int(group["image_count"]) + 1
                archive_filename = (
                    f"{group['folder']}/"
                    f"{int(group['image_count']):02d}{extension}"
                )
                archive.write(source, archive_filename)
                image_count += 1
                archived_for_visit += 1
                if field not in seen_fields:
                    seen_fields.add(field)
                    field_rows.append({"row": row_number, "field": field})
            completed.append(
                {
                    "row": row_number,
                    "terminal_name": terminal_name,
                    "partner_name": record["partner_name"],
                    "image_count": archived_for_visit,
                }
            )

        report = {
            "total": total,
            "input_count": input_stats.get("input_count", total),
            "duplicate_count": input_stats.get("duplicate_count", 0),
            "invalid_count": input_stats.get("invalid_count", 0),
            "rejected_count": input_stats.get("rejected_count", 0),
            "retry_count": retry_count,
            "succeeded": len(completed),
            "failed": len(errors),
            "image_count": image_count,
            "field_rows": field_rows,
            "completed": completed,
            "errors": errors,
        }
        archive.writestr(
            "提取结果.json",
            json.dumps(report, ensure_ascii=False, indent=2),
        )

    if not completed:
        archive_path.unlink(missing_ok=True)
        first_error = errors[0]["error"] if errors else "没有成功提取任何图片"
        raise ValueError(f"批量提取失败：{first_error}")

    safe_partner = re.sub(
        r'[\\/:*?"<>|]',
        "_",
        str(completed[0]["partner_name"]).strip(),
    ) or "未知业务员"
    archive_stem = f"{datetime.now():%Y%m%d}_{safe_partner}_{len(seen_fields)}"
    archive_name = f"{archive_stem}.zip"
    final_archive_path = batch_dir / archive_name
    sequence = 2
    while final_archive_path.exists():
        archive_name = f"{archive_stem}_{sequence:02d}.zip"
        final_archive_path = batch_dir / archive_name
        sequence += 1
    archive_path.replace(final_archive_path)
    return {
        **report,
        "archive_name": archive_name,
        "archive_url": _image_url("_batches", archive_name),
    }


def _batch_checkpoint_dir() -> Path:
    path = OUTPUT_ROOT / "_system" / "batch_jobs"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _batch_checkpoint_path(job_id: str) -> Path:
    return _batch_checkpoint_dir() / f"{job_id}.json"


def _write_batch_checkpoint(job_id: str, job: dict) -> None:
    path = _batch_checkpoint_path(job_id)
    temp_path = path.with_suffix(".tmp")
    temp_path.write_text(
        json.dumps(job, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    temp_path.replace(path)


def _update_batch_job(job_id: str, **values) -> None:
    with BATCH_JOBS_LOCK:
        job = BATCH_JOBS.get(job_id)
        if job is None:
            return
        job.update(values)
        job["updated_at"] = time.time()
        checkpoint = dict(job)
    _write_batch_checkpoint(job_id, checkpoint)


def _register_batch_job(job_id: str, job: dict) -> None:
    with BATCH_JOBS_LOCK:
        BATCH_JOBS[job_id] = job
        checkpoint = dict(job)
    _write_batch_checkpoint(job_id, checkpoint)


def _extract_batch_record(row_number: int, link: str) -> dict:
    result = extract_images(link, OUTPUT_ROOT)
    added_image_count = IMAGE_LIBRARY.add_result(result, source_url=link)
    images: list[dict] = []
    fields: set[str] = set()
    for image in result.images:
        try:
            field = photoid_name_field(image.photoid)
        except ValueError:
            continue
        fields.add(field)
        images.append(
            {
                "field": field,
                "source": str(Path(result.output_dir) / image.filename),
            }
        )
    return {
        "row": row_number,
        "terminal_name": result.terminal_name,
        "partner_name": result.partner_name,
        "images": images,
        "added_image_count": added_image_count,
        "added_fields": sorted(fields) if added_image_count else [],
    }


def _batch_record_metrics(records: list[dict]) -> tuple[int, int]:
    image_count = sum(int(record.get("added_image_count", 0)) for record in records)
    fields = {
        str(field)
        for record in records
        for field in record.get("added_fields", [])
        if str(field)
    }
    return image_count, len(fields)


def _batch_error_information(job: dict) -> str:
    messages = [
        f"第 {item.get('row', '-')} 行：{item.get('error', '处理失败')}"
        for item in job.get("errors", [])
    ]
    invalid_count = int(job.get("invalid_count", 0))
    if invalid_count:
        messages.insert(0, f"无效链接 {invalid_count} 条")
    return "\n".join(messages)


def _run_batch_job_chunk(application: Flask, job_id: str) -> bool:
    with BATCH_JOBS_LOCK:
        current = BATCH_JOBS.get(job_id)
        if current is None:
            return False
        job = dict(current)
    links = [(int(row), str(link)) for row, link in job.get("links", [])]
    total = len(links)
    start = int(job.get("processed", 0))
    end = min(start + BATCH_CHUNK_SIZE, total)
    chunk_count = max(1, (total + BATCH_CHUNK_SIZE - 1) // BATCH_CHUNK_SIZE)
    _update_batch_job(
        job_id,
        status="running",
        chunk_index=min(start // BATCH_CHUNK_SIZE + 1, chunk_count),
        chunk_count=chunk_count,
    )
    try:
        for index in range(start, end):
            row_number, link = links[index]
            record = None
            error_message = "处理失败，请联系管理员"
            attempts_used = 0
            for attempt in range(1, BATCH_LINK_ATTEMPTS + 1):
                attempts_used = attempt
                try:
                    with EXTRACT_LOCK:
                        record = _extract_batch_record(row_number, link)
                    break
                except (ValueError, CrmApiError) as exc:
                    error_message = str(exc)
                except Exception:
                    application.logger.exception(
                        "批量任务 %s 第 %s 行第 %s 次提取失败",
                        job_id,
                        row_number,
                        attempt,
                    )
                    error_message = "处理失败，请联系管理员"
                if attempt < BATCH_LINK_ATTEMPTS:
                    time.sleep(min(0.5 * attempt, 1.5))

            with BATCH_JOBS_LOCK:
                live_job = BATCH_JOBS[job_id]
                completed_records = list(live_job.get("completed_records", []))
                errors = list(live_job.get("errors", []))
                retry_count = int(live_job.get("retry_count", 0)) + max(
                    attempts_used - 1,
                    0,
                )
            if record is not None:
                completed_records.append(record)
            else:
                errors.append(
                    {
                        "row": row_number,
                        "error": error_message,
                        "attempts": attempts_used,
                    }
                )
            _update_batch_job(
                job_id,
                processed=index + 1,
                current_row=row_number,
                succeeded=len(completed_records),
                failed=len(errors),
                image_count=sum(
                    len(item.get("images", [])) for item in completed_records
                ),
                completed_records=completed_records,
                errors=errors,
                retry_count=retry_count,
            )

        if end < total:
            _update_batch_job(job_id, status="queued")
            return True

        with BATCH_JOBS_LOCK:
            finished_job = dict(BATCH_JOBS[job_id])
        input_stats = {
            key: int(finished_job.get(key, 0))
            for key in (
                "input_count",
                "duplicate_count",
                "invalid_count",
                "rejected_count",
            )
        }
        result = _build_batch_archive(
            list(finished_job.get("completed_records", [])),
            list(finished_job.get("errors", [])),
            total,
            input_stats,
            int(finished_job.get("retry_count", 0)),
        )
        _update_batch_job(
            job_id,
            status="completed",
            processed=total,
            result=result,
        )
        record_image_count, record_terminal_count = _batch_record_metrics(
            list(finished_job.get("completed_records", []))
        )
        has_errors = bool(finished_job.get("errors")) or int(
            finished_job.get("invalid_count", 0)
        ) > 0
        _complete_extraction_record(
            application,
            str(finished_job.get("record_id") or job_id),
            status="partial_success" if has_errors else "success",
            image_count=record_image_count,
            terminal_count=record_terminal_count,
            error_information=_batch_error_information(finished_job),
        )
        return False
    except ValueError as exc:
        _update_batch_job(job_id, status="failed", error=str(exc))
        with BATCH_JOBS_LOCK:
            failed_job = dict(BATCH_JOBS.get(job_id) or job)
        record_image_count, record_terminal_count = _batch_record_metrics(
            list(failed_job.get("completed_records", []))
        )
        _complete_extraction_record(
            application,
            str(failed_job.get("record_id") or job_id),
            status="failed",
            image_count=record_image_count,
            terminal_count=record_terminal_count,
            error_information="\n".join(
                item
                for item in (_batch_error_information(failed_job), str(exc))
                if item
            ),
        )
        return False
    except Exception:
        application.logger.exception("批量提取图片失败")
        _update_batch_job(
            job_id,
            status="failed",
            error="批量提取失败，请联系管理员查看服务日志",
        )
        with BATCH_JOBS_LOCK:
            failed_job = dict(BATCH_JOBS.get(job_id) or job)
        record_image_count, record_terminal_count = _batch_record_metrics(
            list(failed_job.get("completed_records", []))
        )
        _complete_extraction_record(
            application,
            str(failed_job.get("record_id") or job_id),
            status="failed",
            image_count=record_image_count,
            terminal_count=record_terminal_count,
            error_information="批量提取失败，请联系管理员查看服务日志",
        )
        return False


def _batch_worker(application: Flask) -> None:
    while True:
        job_id = BATCH_QUEUE.get()
        try:
            if _run_batch_job_chunk(application, job_id):
                BATCH_QUEUE.put(job_id)
        finally:
            BATCH_QUEUE.task_done()


def _restore_batch_jobs() -> None:
    for path in _batch_checkpoint_dir().glob("*.json"):
        try:
            job = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if str(job.get("status", "")) not in {"queued", "running"}:
            continue
        job_id = path.stem
        job["status"] = "queued"
        job["resumed"] = True
        with BATCH_JOBS_LOCK:
            BATCH_JOBS[job_id] = job
        _write_batch_checkpoint(job_id, job)
        BATCH_QUEUE.put(job_id)


def _start_batch_worker(application: Flask) -> None:
    global BATCH_WORKER_STARTED
    with BATCH_WORKER_LOCK:
        if BATCH_WORKER_STARTED:
            return
        _restore_batch_jobs()
        threading.Thread(
            target=_batch_worker,
            args=(application,),
            daemon=True,
            name="batch-worker",
        ).start()
        BATCH_WORKER_STARTED = True


def _prune_batch_jobs() -> None:
    cutoff = time.time() - BATCH_JOB_TTL_SECONDS
    with BATCH_JOBS_LOCK:
        expired = [
            job_id
            for job_id, job in BATCH_JOBS.items()
            if job.get("status") in {"completed", "failed"}
            and job.get("updated_at", 0) < cutoff
        ]
        for job_id in expired:
            BATCH_JOBS.pop(job_id, None)
            _batch_checkpoint_path(job_id).unlink(missing_ok=True)


def _public_batch_job(job: dict) -> dict:
    fields = (
        "status",
        "processed",
        "total",
        "current_row",
        "succeeded",
        "failed",
        "image_count",
        "input_count",
        "duplicate_count",
        "invalid_count",
        "rejected_count",
        "retry_count",
        "chunk_index",
        "chunk_count",
        "resumed",
        "result",
        "error",
    )
    return {field: job[field] for field in fields if field in job}


def _load_saved_results() -> list[dict]:
    results: list[dict] = []
    if not OUTPUT_ROOT.exists():
        return results

    metadata_files = sorted(
        OUTPUT_ROOT.glob("**/metadata.json"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    for metadata_file in metadata_files:
        try:
            data = json.loads(metadata_file.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue

        folder = str(metadata_file.parent.relative_to(OUTPUT_ROOT))
        images = []
        private_images = [
            item
            for item in data.get("images", [])
            if str(item.get("photoid") or "").startswith("private")
        ]
        for index, item in enumerate(private_images, start=1):
            filename = item.get("filename", "")
            image_file = metadata_file.parent / filename
            if filename and image_file.is_file():
                try:
                    display_filename = build_image_filename(
                        item["photoid"],
                        data.get("terminal_name") or "未知终端",
                        data.get("partner_name") or "未知业务员",
                        index,
                        image_file.suffix,
                    )
                except (KeyError, ValueError):
                    display_filename = filename
                images.append(
                    {
                        "filename": filename,
                        "display_filename": display_filename,
                        "size_bytes": item.get("size_bytes", image_file.stat().st_size),
                        "url": _image_url(folder, filename),
                    }
                )

        results.append(
            {
                "visit_id": data.get("visit_id", ""),
                "terminal_name": data.get("terminal_name", "未知终端"),
                "partner_name": data.get("partner_name", "未知业务员"),
                "extracted_at": data.get("extracted_at", ""),
                "images": images,
            }
        )
    return results


def _current_user() -> str | None:
    if AUTH_MODE == "off":
        return "本地用户"
    if AUTH_MODE in {"password", "oidc"}:
        return session.get("user")

    header = os.environ.get(
        "INFOLENS_PROXY_USER_HEADER",
        "Cf-Access-Authenticated-User-Email",
    )
    user = request.headers.get(header, "").strip().lower()
    if not user:
        return None

    allowed_domain = os.environ.get("INFOLENS_ALLOWED_EMAIL_DOMAIN", "").lower()
    allowed_emails = {
        email.strip().lower()
        for email in os.environ.get("INFOLENS_ALLOWED_EMAILS", "").split(",")
        if email.strip()
    }
    if allowed_domain and not user.endswith(f"@{allowed_domain}"):
        return None
    if allowed_emails and user not in allowed_emails:
        return None
    return user


def _current_role() -> str:
    if AUTH_MODE == "off":
        return "admin"
    return str(session.get("role") or "user")


def _is_admin() -> bool:
    return _current_role() == "admin"


def _identity_allowed(user: str) -> bool:
    normalized = user.strip().lower()
    allowed_domain = os.environ.get("INFOLENS_ALLOWED_EMAIL_DOMAIN", "").lower()
    allowed_emails = {
        email.strip().lower()
        for email in os.environ.get("INFOLENS_ALLOWED_EMAILS", "").split(",")
        if email.strip()
    }
    if allowed_domain and not normalized.endswith(f"@{allowed_domain}"):
        return False
    if allowed_emails and normalized not in allowed_emails:
        return False
    return True


def _login_required(view: Callable):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if _current_user():
            return view(*args, **kwargs)
        if request.path.startswith("/api/") or request.path.startswith("/output/"):
            return jsonify({"error": "请先登录"}), 401
        return redirect(url_for("login", next=request.path))

    return wrapped


def _admin_required(view: Callable):
    @wraps(view)
    @_login_required
    def wrapped(*args, **kwargs):
        if _is_admin():
            return view(*args, **kwargs)
        return jsonify({"error": "没有权限访问用户管理"}), 403

    return wrapped


def _super_admin_config() -> tuple[str, str, str]:
    username = (
        os.environ.get("INFOLENS_SUPER_ADMIN_USERNAME")
        or os.environ.get("INFOLENS_USERNAME")
        or "admin"
    )
    display_name = os.environ.get("INFOLENS_SUPER_ADMIN_DISPLAY_NAME", "超级管理员")
    password_hash = os.environ.get("INFOLENS_SUPER_ADMIN_PASSWORD_HASH")
    if not password_hash:
        password = os.environ.get("INFOLENS_SUPER_ADMIN_PASSWORD")
        if password:
            password_hash = generate_password_hash(password, method="pbkdf2:sha256")
        else:
            password_hash = os.environ.get("INFOLENS_PASSWORD_HASH", "")
    return username, password_hash, display_name


def _ensure_super_admin() -> None:
    if AUTH_MODE != "password":
        return
    username, password_hash, display_name = _super_admin_config()
    if not password_hash:
        return
    USER_STORE.ensure_super_admin(
        username=username,
        password_hash=password_hash,
        display_name=display_name,
    )


def _csrf_token() -> str:
    token = session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["csrf_token"] = token
    return token


def _check_csrf() -> None:
    expected = session.get("csrf_token", "")
    supplied = request.headers.get("X-CSRF-Token", "")
    if not supplied and request.is_json:
        payload = request.get_json(silent=True) or {}
        supplied = str(payload.get("csrf_token") or "")
    if not supplied:
        supplied = str(request.form.get("csrf_token") or "")
    if not expected or not hmac.compare_digest(expected, supplied):
        abort(403, description="安全令牌无效，请刷新页面后重试")


def _check_rate_limit() -> None:
    limit = int(os.environ.get("INFOLENS_RATE_LIMIT", "10"))
    window = int(os.environ.get("INFOLENS_RATE_WINDOW_SECONDS", "600"))
    identity = _current_user() or request.remote_addr or "unknown"
    now = time.monotonic()
    with RATE_LOCK:
        bucket = RATE_BUCKETS[identity]
        while bucket and now - bucket[0] > window:
            bucket.popleft()
        if len(bucket) >= limit:
            abort(429, description="请求过于频繁，请稍后再试")
        bucket.append(now)


def _customer_operator() -> tuple[str, str]:
    username = str(_current_user() or "")
    return username, str(session.get("display_name") or username)


def _excel_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _safe_excel_cell(value) -> str:
    text = _excel_text(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _parse_customer_workbook(file_stream) -> list[dict]:
    try:
        workbook = load_workbook(file_stream, read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise ValueError("无法读取 Excel 文件，请确认文件为有效的 .xlsx 格式") from exc
    try:
        worksheet = workbook.active
        worksheet.reset_dimensions()
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [_excel_text(value) for value in first_row]
        while headers and not headers[-1]:
            headers.pop()
        if not headers:
            raise ValueError("Excel 表头不能为空")
        if any(not header for header in headers):
            raise ValueError("Excel 表头中不能包含空白列")
        duplicates = sorted({header for header in headers if headers.count(header) > 1})
        if duplicates:
            raise ValueError(f"Excel 表头存在重复字段：{'、'.join(duplicates)}")
        missing = [header for header in EXCEL_HEADERS if header not in headers]
        unknown = [header for header in headers if header not in EXCEL_HEADERS]
        if missing:
            raise ValueError(f"Excel 缺少表头：{'、'.join(missing)}")
        if unknown:
            raise ValueError(f"Excel 存在未知表头：{'、'.join(unknown)}")

        rows = []
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            raw = {
                header: _excel_text(values[index] if index < len(values) else None)
                for index, header in enumerate(headers)
            }
            if not any(raw.values()):
                continue
            rows.append(
                {
                    "row_number": row_number,
                    "raw": raw,
                    "payload": {
                        HEADER_TO_FIELD[header]: raw[header]
                        for header in EXCEL_HEADERS
                    },
                }
            )
        return rows
    finally:
        workbook.close()


def _customer_workbook_response(*, errors: list[dict] | None = None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "客户档案"
    headers = [*EXCEL_HEADERS, *(["失败原因"] if errors is not None else [])]
    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="165DFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{worksheet.cell(1, len(headers)).coordinate}"
    widths = [16, 28, 18, 12, 18, 14, 16, 16, 36, 18, 36]
    if errors is not None:
        widths.append(50)
        for error in errors:
            raw = error.get("data") or {}
            worksheet.append(
                [_safe_excel_cell(raw.get(header, "")) for header in EXCEL_HEADERS]
                + [_safe_excel_cell(error.get("error", ""))]
            )
    else:
        worksheet.append(
            [
                "1000000001",
                "示例客户（请删除本行）",
                "示例业态",
                "运营",
                "示例线路",
                SALESPEOPLE[0],
                SNOW_SALESPEOPLE[0],
                "",
                "",
                "",
                "",
            ]
        )
        status_validation = DataValidation(
            type="list",
            formula1='"运营,停用"',
            allow_blank=True,
        )
        salesperson_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(SALESPEOPLE)}"',
        )
        snow_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(SNOW_SALESPEOPLE)}"',
            allow_blank=True,
        )
        for validation, cells in (
            (status_validation, "D2:D1000"),
            (salesperson_validation, "F2:F1000"),
            (snow_validation, "G2:G1000"),
        ):
            worksheet.add_data_validation(validation)
            validation.add(cells)
        for row in range(2, 1001):
            worksheet.cell(row, 1).number_format = "@"
            worksheet.cell(row, 10).number_format = "@"
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _run_wecom_extract_job(
    application: Flask,
    crypto: WecomBotCrypto,
    task_id: str,
    message: dict,
    links: list[str],
) -> None:
    succeeded: list[ExtractResult] = []
    errors: list[str] = []
    response_url = str(message.get("response_url") or "")

    with EXTRACT_LOCK:
        for position, link in enumerate(links, start=1):
            try:
                result = extract_images(
                    link,
                    OUTPUT_ROOT,
                    group_by_partner=True,
                )
                IMAGE_LIBRARY.add_result(result, source_url=link)
                succeeded.append(result)
                audit = {
                    "task_id": task_id,
                    "wecom_message_id": message.get("msgid"),
                    "wecom_user_id": (message.get("from") or {}).get("userid"),
                    "wecom_chat_id": message.get("chatid"),
                    "received_at": datetime.now().isoformat(timespec="seconds"),
                }
                (Path(result.output_dir) / "wecom_submission.json").write_text(
                    json.dumps(audit, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
            except (ValueError, CrmApiError) as exc:
                errors.append(f"第 {position} 条：{exc}")
            except Exception:
                application.logger.exception(
                    "企业微信任务 %s 的第 %s 条链接处理失败",
                    task_id,
                    position,
                )
                errors.append(f"第 {position} 条：处理失败，请联系管理员")

    lines = [
        f"**任务 {task_id} 处理完成**",
        f"> 成功：{len(succeeded)} 条",
        f"> 失败：{len(errors)} 条",
    ]
    for result in succeeded:
        lines.append(
            f"- {result.partner_name}｜{result.terminal_name}｜{len(result.images)} 张图片"
        )
    if errors:
        lines.append("\n**失败明细**")
        lines.extend(f"- {error}" for error in errors[:5])
        if len(errors) > 5:
            lines.append(f"- 另有 {len(errors) - 5} 条失败")

    allowed_hosts = {
        item.strip().lower()
        for item in os.environ.get(
            "WECOM_BOT_RESPONSE_HOSTS",
            "qyapi.weixin.qq.com",
        ).split(",")
        if item.strip()
    }
    try:
        send_response_url(
            response_url,
            stream_reply(f"{task_id}-result", "\n".join(lines)),
            crypto,
            allowed_hosts=allowed_hosts,
        )
    except WecomBotError:
        application.logger.exception("企业微信任务 %s 的结果通知失败", task_id)


def _create_distribution_archive(business: str) -> dict:
    jobs = DISTRIBUTION_STORE.completed_for_business(business)
    if not jobs:
        raise ValueError("该业务暂无可下载的提取内容")

    archive_root = OUTPUT_ROOT / "_distribution_downloads"
    archive_root.mkdir(parents=True, exist_ok=True)
    safe_business = re.sub(r'[\\/:*?"<>|]', "_", business.strip()) or "未知业务"
    archive_key = f"{datetime.now():%Y%m%d_%H%M%S}_{secrets.token_hex(3)}"
    temp_path = archive_root / f".{archive_key}.zip"
    groups: dict[tuple[str, str], dict[str, str | int]] = {}
    archived_images = 0
    fields: set[str] = set()

    with zipfile.ZipFile(
        temp_path,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=6,
    ) as archive:
        for job in jobs:
            output_dir = Path(job.output_dir).resolve()
            try:
                output_dir.relative_to(OUTPUT_ROOT)
            except ValueError:
                continue
            metadata_file = output_dir / "metadata.json"
            try:
                metadata = json.loads(metadata_file.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue

            for item in metadata.get("images") or []:
                filename = str(item.get("filename") or "")
                image_file = (output_dir / filename).resolve()
                if not filename or not image_file.is_file():
                    continue
                try:
                    image_file.relative_to(output_dir)
                    field = photoid_name_field(str(item.get("photoid") or ""))
                except ValueError:
                    continue
                terminal = str(
                    metadata.get("terminal_name")
                    or job.terminal_name
                    or "未知终端"
                )
                group_key = (field, terminal)
                group = groups.get(group_key)
                if group is None:
                    safe_field = re.sub(r'[\\/:*?"<>|]', "_", field)
                    safe_terminal = re.sub(r'[\\/:*?"<>|]', "_", terminal)
                    group = {
                        "folder": (
                            f"{len(groups) + 1:02d}_{safe_field}_{safe_terminal}"
                        ),
                        "image_count": 0,
                    }
                    groups[group_key] = group
                group["image_count"] = int(group["image_count"]) + 1
                extension = image_file.suffix.lower() or ".jpg"
                archive.write(
                    image_file,
                    (
                        f"{group['folder']}/"
                        f"{int(group['image_count']):02d}{extension}"
                    ),
                )
                fields.add(field)
                archived_images += 1

        report = {
            "business": business,
            "field_count": len(fields),
            "fields": sorted(fields),
            "distributed_count": len(jobs),
            "image_count": archived_images,
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
        archive.writestr(
            "分发提取结果.json",
            json.dumps(report, ensure_ascii=False, indent=2),
        )

    if not archived_images:
        temp_path.unlink(missing_ok=True)
        raise ValueError("该业务没有可打包的图片文件")

    archive_name = (
        f"{datetime.now():%Y%m%d}_{safe_business}_{len(fields)}个字段.zip"
    )
    archive_path = archive_root / archive_name
    sequence = 2
    while archive_path.exists():
        archive_name = (
            f"{datetime.now():%Y%m%d}_{safe_business}_"
            f"{len(fields)}个字段_{sequence:02d}.zip"
        )
        archive_path = archive_root / archive_name
        sequence += 1
    temp_path.replace(archive_path)
    DISTRIBUTION_STORE.mark_downloaded(business)
    return {
        "archive_name": archive_name,
        "archive_url": _image_url("_distribution_downloads", archive_name),
        **report,
    }


def _parse_field_lines(value: str) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for item in re.split(r"[\s,，;；]+", value):
        field = item.strip()
        if not field or field in seen:
            continue
        seen.add(field)
        fields.append(field)
    return fields


def _parse_string_values(value) -> list[str]:
    if isinstance(value, list):
        source = value
    elif value is None:
        source = []
    else:
        source = [value]
    return list(
        dict.fromkeys(
            str(item).strip()
            for item in source
            if str(item).strip()
        )
    )


def _parse_pagination(page_value, page_size_value) -> tuple[int, int]:
    try:
        page = int(page_value or 1)
        page_size = int(page_size_value or IMAGE_LIBRARY_PAGE_SIZE)
    except (TypeError, ValueError) as exc:
        raise ValueError("分页参数必须是整数") from exc
    if page < 1 or page_size < 1:
        raise ValueError("分页参数必须大于 0")
    return page, min(page_size, 50)


def _attach_image_archive_tags(result: dict) -> None:
    images = [
        image
        for item in result.get("items", [])
        for image in item.get("images", [])
    ]
    archive_map = IMAGE_LIBRARY.archived_policy_ids_by_image(
        [str(image.get("id") or "") for image in images]
    )
    policy_ids = list(
        dict.fromkeys(
            policy_id
            for policy_ids in archive_map.values()
            for policy_id in policy_ids
        )
    )
    policy_map = SNOW_OUTBOUND_STORE.policy_summaries(
        policy_ids,
        include_deleted=True,
    )
    for image in images:
        image["archive_tags"] = [
            {
                "policy_id": policy_map[policy_id]["policy_id"],
                "tag": policy_map[policy_id]["tag"],
                "color": policy_map[policy_id]["color"],
                "enabled": policy_map[policy_id]["enabled"],
                "deleted": policy_map[policy_id]["deleted"],
            }
            for policy_id in archive_map.get(str(image.get("id") or ""), [])
            if policy_id in policy_map
        ]


def _archive_policy_options(month: str) -> list[dict]:
    if not month:
        return []
    policy_ids = IMAGE_LIBRARY.archived_policy_ids(month)
    policy_map = SNOW_OUTBOUND_STORE.policy_summaries(
        policy_ids,
        include_deleted=True,
    )
    return [
        {
            "id": policy_id,
            "display_name": policy_map[policy_id]["tag"],
            "color": policy_map[policy_id]["color"],
            "deleted": policy_map[policy_id]["deleted"],
        }
        for policy_id in policy_ids
        if policy_id in policy_map
    ]


def _private_image_cache(response):
    response.cache_control.public = False
    response.cache_control.private = True
    response.cache_control.max_age = IMAGE_CACHE_SECONDS
    response.cache_control.immutable = True
    return response


def _serve_output_file(
    path: Path,
    *,
    cache: bool = False,
    as_attachment: bool = False,
    download_name: str | None = None,
    mimetype: str | None = None,
):
    resolved = path.resolve()
    try:
        relative = resolved.relative_to(OUTPUT_ROOT)
    except ValueError:
        abort(404)
    if not resolved.is_file():
        abort(404)

    guessed_type = mimetype or mimetypes.guess_type(resolved.name)[0]
    if X_ACCEL_ENABLED:
        response = Response(status=200, mimetype=guessed_type or "application/octet-stream")
        encoded_path = "/".join(
            urllib.parse.quote(part, safe="") for part in relative.parts
        )
        response.headers["X-Accel-Redirect"] = f"{X_ACCEL_PREFIX}/{encoded_path}"
        response.headers["X-Accel-Expires"] = (
            str(IMAGE_CACHE_SECONDS) if cache else "0"
        )
        if as_attachment:
            safe_download_name = download_name or resolved.name
            response.headers.set(
                "Content-Disposition",
                "attachment",
                filename=safe_download_name,
            )
    else:
        response = send_file(
            resolved,
            conditional=True,
            max_age=IMAGE_CACHE_SECONDS if cache else None,
            as_attachment=as_attachment,
            download_name=download_name,
            mimetype=guessed_type,
        )
    return _private_image_cache(response) if cache else response


def _safe_archive_part(value: str, fallback: str = "未知") -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]', "_", str(value or "").strip())
    return cleaned[:100] or fallback


def _create_photo_archive(policy: dict) -> tuple[Path, str, int, int]:
    archived_items = IMAGE_LIBRARY.archived_images(policy["id"])
    valid_items: list[tuple[dict, Path]] = []
    for item in archived_items:
        image = item["image"]
        source = (OUTPUT_ROOT / image.file_path).resolve()
        try:
            source.relative_to(OUTPUT_ROOT)
        except ValueError:
            continue
        if source.is_file():
            valid_items.append((item, source))
    if not valid_items:
        raise ValueError("该政策标签暂无可导出的归档照片")

    export_root = OUTPUT_ROOT / "_photo_archive_exports"
    export_root.mkdir(parents=True, exist_ok=True)
    archive_name = (
        f"{datetime.now():%Y%m%d_%H%M%S}_"
        f"{_safe_archive_part(policy['display_name'], '雪花政策')}_照片档案.zip"
    )
    archive_path = export_root / f"{secrets.token_hex(8)}.zip"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "照片档案明细"
    headers = [
        "政策标签",
        "政策年月",
        "终端编码",
        "客户全名",
        "终端照片数量",
        "照片文件名",
        "归档人",
        "归档时间",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="165DFF")
        cell.alignment = Alignment(horizontal="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:H{len(valid_items) + 1}"
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 16
    worksheet.column_dimensions["D"].width = 30
    worksheet.column_dimensions["E"].width = 14
    worksheet.column_dimensions["F"].width = 38
    worksheet.column_dimensions["G"].width = 16
    worksheet.column_dimensions["H"].width = 22

    folder_sequences: dict[str, int] = defaultdict(int)
    terminal_photo_counts: dict[str, int] = defaultdict(int)
    for item, _source in valid_items:
        terminal_photo_counts[item["image"].field] += 1
    terminal_codes: set[str] = set()
    with zipfile.ZipFile(
        archive_path,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=6,
    ) as archive:
        for item, source in valid_items:
            image = item["image"]
            terminal_codes.add(image.field)
            folder = (
                f"{_safe_archive_part(image.field, '未知终端')}_"
                f"{_safe_archive_part(image.customer_name, '未知客户')}"
            )
            folder_sequences[folder] += 1
            extension = source.suffix.lower() or ".jpg"
            archived_filename = (
                f"{folder_sequences[folder]:03d}_"
                f"{_safe_archive_part(image.filename, '照片')}"
            )
            if not Path(archived_filename).suffix:
                archived_filename += extension
            archive.write(source, f"{folder}/{archived_filename}")
            worksheet.append(
                [
                    _safe_excel_cell(policy["display_name"]),
                    f"{int(policy['year']):04d}-{int(policy['month']):02d}",
                    _safe_excel_cell(image.field),
                    _safe_excel_cell(image.customer_name),
                    terminal_photo_counts[image.field],
                    _safe_excel_cell(archived_filename),
                    _safe_excel_cell(
                        item["archived_by_name"] or item["archived_by"]
                    ),
                    item["archived_at"],
                ]
            )
        workbook_stream = io.BytesIO()
        workbook.save(workbook_stream)
        archive.writestr("照片档案明细.xlsx", workbook_stream.getvalue())
    return archive_path, archive_name, len(valid_items), len(terminal_codes)


def _cleanup_legacy_image_exports() -> None:
    legacy_export_root = OUTPUT_ROOT / "_image_exports"
    if legacy_export_root.is_dir():
        shutil.rmtree(legacy_export_root)
    database = OUTPUT_ROOT / "_system" / "export_records.sqlite3"
    for candidate in (database, Path(f"{database}-wal"), Path(f"{database}-shm")):
        candidate.unlink(missing_ok=True)
    transient_root = OUTPUT_ROOT / "_photo_archive_exports"
    if transient_root.is_dir():
        shutil.rmtree(transient_root)


def create_app() -> Flask:
    _require_production_config()
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    _cleanup_legacy_image_exports()
    EXTRACTION_RECORD_STORE.prune_expired()
    _ensure_super_admin()
    if os.environ.get("INFOLENS_DISTRIBUTION_IMPORT_EXISTING", "").lower() in {
        "1",
        "true",
        "yes",
        "on",
    }:
        DISTRIBUTION_STORE.import_existing_outputs(OUTPUT_ROOT)
    if os.environ.get("INFOLENS_IMAGE_LIBRARY_IMPORT_EXISTING", "").lower() in {
        "1",
        "true",
        "yes",
        "on",
    }:
        IMAGE_LIBRARY.import_existing_outputs()

    application = Flask(__name__, static_folder=None)
    application.wsgi_app = ProxyFix(
        application.wsgi_app,
        x_for=1,
        x_proto=1,
        x_host=1,
    )
    application.config.update(
        MAX_CONTENT_LENGTH=MAX_UPLOAD_BYTES,
        SECRET_KEY=os.environ.get("INFOLENS_SESSION_SECRET") or secrets.token_hex(32),
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SAMESITE="Lax",
        SESSION_COOKIE_SECURE=os.environ.get("INFOLENS_ENV") == "production",
        PERMANENT_SESSION_LIFETIME=8 * 60 * 60,
    )
    oauth = OAuth(application)
    wecom_crypto = None
    if WECOM_BOT_CALLBACK_ENABLED:
        wecom_crypto = WecomBotCrypto(
            os.environ["WECOM_BOT_TOKEN"],
            os.environ["WECOM_BOT_ENCODING_AES_KEY"],
        )
    if AUTH_MODE == "oidc":
        oauth.register(
            name="company",
            server_metadata_url=os.environ["INFOLENS_OIDC_METADATA_URL"],
            client_id=os.environ["INFOLENS_OIDC_CLIENT_ID"],
            client_secret=os.environ["INFOLENS_OIDC_CLIENT_SECRET"],
            client_kwargs={"scope": "openid email profile"},
        )

    @application.after_request
    def security_headers(response):
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "no-referrer"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "img-src 'self' data:; "
            "style-src 'self' 'unsafe-inline'; "
            "script-src 'self' 'unsafe-inline'; "
            "connect-src 'self'; "
            "frame-ancestors 'none'; "
            "base-uri 'none'; "
            "form-action 'self'"
        )
        if not (
            response.cache_control.private
            and response.cache_control.max_age == IMAGE_CACHE_SECONDS
        ):
            response.headers["Cache-Control"] = "no-store"
        return response

    @application.get("/healthz")
    def healthz():
        return jsonify({"status": "ok"})

    @application.route("/api/wecom/bot/callback", methods=["GET", "POST"])
    def wecom_bot_callback():
        if not WECOM_BOT_CALLBACK_ENABLED or wecom_crypto is None:
            return jsonify({"error": "企业微信智能机器人未启用"}), 503

        msg_signature = request.args.get("msg_signature", "")
        timestamp = request.args.get("timestamp", "")
        nonce = request.args.get("nonce", "")
        if not msg_signature or not timestamp or not nonce:
            return jsonify({"error": "企业微信回调参数不完整"}), 400

        try:
            validate_callback_timestamp(
                timestamp,
                int(os.environ.get("WECOM_BOT_CALLBACK_MAX_AGE_SECONDS", "600")),
            )
            if request.method == "GET":
                echo_str = request.args.get("echostr", "")
                if not echo_str:
                    return jsonify({"error": "企业微信回调缺少 echostr"}), 400
                return wecom_crypto.verify_url(
                    msg_signature,
                    timestamp,
                    nonce,
                    echo_str,
                )

            if request.content_length and request.content_length > 1024 * 1024:
                return jsonify({"error": "企业微信回调正文过大"}), 413
            message = wecom_crypto.decrypt(
                request.get_data(cache=False),
                msg_signature,
                timestamp,
                nonce,
            )
            if message.get("aibotid") != os.environ["WECOM_BOT_ID"]:
                raise WecomBotError("企业微信机器人 ID 不匹配")

            if message.get("msgtype") == "event":
                event_type = (message.get("event") or {}).get("eventtype")
                payload = (
                    text_reply(
                        "发送 CRM 拜访详情链接，我会自动提取图片并按业务员归档。"
                    )
                    if event_type == "enter_chat"
                    else {}
                )
                return wecom_crypto.encrypt(payload, nonce)

            if message.get("msgtype") == "stream":
                return wecom_crypto.encrypt({}, nonce)

            links = extract_crm_urls(
                message_text(message),
                max_links=WECOM_BOT_MAX_LINKS,
            )
            if not links:
                return wecom_crypto.encrypt(
                    stream_reply(
                        f"help-{secrets.token_hex(6)}",
                        "没有识别到 CRM 拜访链接。\n"
                        "请发送包含 `visitDetail` 或 `workCirclevisit` 的链接。",
                    ),
                    nonce,
                )
            if not message.get("response_url"):
                return wecom_crypto.encrypt(
                    stream_reply(
                        f"error-{secrets.token_hex(6)}",
                        "消息缺少结果回传地址，请重新发送链接。",
                    ),
                    nonce,
                )

            proposed_task_id = (
                f"IL{datetime.now():%Y%m%d%H%M%S}{secrets.token_hex(2).upper()}"
            )
            duplicate, task_id = WECOM_DEDUPLICATOR.remember(
                str(message.get("msgid") or ""),
                proposed_task_id,
            )
            if not duplicate:
                threading.Thread(
                    target=_run_wecom_extract_job,
                    args=(application, wecom_crypto, task_id, message, links),
                    daemon=True,
                    name=f"wecom-{task_id}",
                ).start()

            duplicate_note = "（重复消息，未再次执行）" if duplicate else ""
            return wecom_crypto.encrypt(
                stream_reply(
                    f"{task_id}-accepted",
                    f"已接收 {len(links)} 条链接，任务号：`{task_id}`{duplicate_note}\n"
                    "图片正在后台提取，完成后会自动回复。",
                ),
                nonce,
            )
        except WecomBotError as exc:
            application.logger.warning("企业微信回调被拒绝：%s", exc)
            return jsonify({"error": "企业微信回调验证失败"}), 403
        except Exception:
            application.logger.exception("企业微信智能机器人回调处理失败")
            return jsonify({"error": "企业微信回调处理失败"}), 500

    @application.route("/login", methods=["GET", "POST"])
    def login():
        if AUTH_MODE == "off":
            return redirect(url_for("index"))
        if AUTH_MODE == "proxy":
            if _current_user():
                return redirect(url_for("index"))
            return "身份验证失败，请通过公司的登录入口访问。", 401
        if AUTH_MODE == "oidc":
            callback = url_for("oidc_callback", _external=True)
            return oauth.company.authorize_redirect(callback)

        error = ""
        if request.method == "POST":
            _check_csrf()
            username = request.form.get("username", "")
            password = request.form.get("password", "")
            user = USER_STORE.authenticate(username, password)
            if user:
                session.clear()
                session["user"] = user["username"]
                session["user_id"] = user["id"]
                session["role"] = user["role"]
                session["display_name"] = user["display_name"]
                session["is_super_admin"] = user["is_super_admin"]
                session.permanent = True
                destination = request.args.get("next", "/")
                if not destination.startswith("/") or destination.startswith("//"):
                    destination = "/"
                return redirect(destination)
            error = "账号或密码不正确"

        csrf = _csrf_token()
        return (
            (WEB_ROOT / "login.html")
            .read_text(encoding="utf-8")
            .replace("{{ERROR}}", error)
            .replace("{{CSRF_TOKEN}}", csrf)
        )

    @application.get("/auth/callback")
    def oidc_callback():
        if AUTH_MODE != "oidc":
            abort(404)
        token = oauth.company.authorize_access_token()
        userinfo = token.get("userinfo") or oauth.company.userinfo()
        user = str(userinfo.get("email") or userinfo.get("preferred_username") or "")
        if not user or not _identity_allowed(user):
            return "该公司账号没有 InfoLens 访问权限。", 403
        session.clear()
        session["user"] = user.lower()
        session["role"] = "user"
        session.permanent = True
        return redirect(url_for("index"))

    @application.get("/logout")
    def logout():
        session.clear()
        return redirect(url_for("login"))

    @application.get("/")
    @_login_required
    def index():
        return send_from_directory(WEB_ROOT, "index.html")

    @application.get("/assets/<path:filename>")
    @_login_required
    def frontend_assets(filename: str):
        response = send_from_directory(WEB_ROOT / "assets", filename)
        response.headers["Access-Control-Allow-Origin"] = "*"
        return response

    @application.get("/api/session")
    @_login_required
    def session_info():
        return jsonify(
            {
                "user": _current_user(),
                "display_name": session.get("display_name") or _current_user(),
                "role": _current_role(),
                "is_admin": _is_admin(),
                "csrf_token": _csrf_token(),
            }
        )

    @application.get("/api/users")
    @_admin_required
    def list_users():
        return jsonify({"items": USER_STORE.list_users()})

    @application.post("/api/users")
    @_admin_required
    def create_user():
        _check_csrf()
        payload = request.get_json(silent=True) or {}
        try:
            return (
                jsonify(
                    USER_STORE.create_user(
                        username=str(payload.get("username") or ""),
                        display_name=str(payload.get("display_name") or ""),
                        password=str(payload.get("password") or ""),
                        role=str(payload.get("role") or "user"),
                        status=str(payload.get("status") or "enabled"),
                    )
                ),
                201,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.patch("/api/users/<int:user_id>")
    @_admin_required
    def update_user(user_id: int):
        _check_csrf()
        payload = request.get_json(silent=True) or {}
        try:
            return jsonify(
                USER_STORE.update_user(
                    user_id,
                    display_name=str(payload.get("display_name") or ""),
                    role=str(payload.get("role") or "user"),
                    status=str(payload.get("status") or "enabled"),
                    password=str(payload.get("password") or ""),
                )
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.delete("/api/users/<int:user_id>")
    @_admin_required
    def delete_user(user_id: int):
        _check_csrf()
        try:
            USER_STORE.delete_user(user_id)
            return jsonify({"message": "用户已删除"})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.get("/api/customers/options")
    @_login_required
    def customer_options():
        return jsonify(
            {
                "statuses": ["运营", "停用"],
                "routes": CUSTOMER_STORE.list_routes(),
                "salespeople": list(SALESPEOPLE),
                "snow_salespeople": list(SNOW_SALESPEOPLE),
                "page_sizes": [20, 50, 100],
            }
        )

    @application.get("/api/customers")
    @_login_required
    def list_customers():
        try:
            return jsonify(
                CUSTOMER_STORE.list_customers(
                    terminal_code=request.args.get("terminal_code", ""),
                    customer_name=request.args.get("customer_name", ""),
                    routes=request.args.getlist("route"),
                    people=request.args.getlist("person"),
                    salesperson=request.args.get("salesperson", ""),
                    snow_salesperson=request.args.get("snow_salesperson", ""),
                    policy_month=(
                        request.args.get("policy_month")
                        or datetime.now().strftime("%Y-%m")
                    ),
                    policy_tag=request.args.get("policy_tag", ""),
                    page=int(request.args.get("page", "1")),
                    page_size=int(request.args.get("page_size", "20")),
                )
            )
        except (TypeError, ValueError) as exc:
            return jsonify({"error": str(exc) or "分页参数不正确"}), 400

    @application.post("/api/customers")
    @_login_required
    def create_customer():
        _check_csrf()
        payload = request.get_json(silent=True) or {}
        operator, operator_name = _customer_operator()
        try:
            return (
                jsonify(
                    CUSTOMER_STORE.create_customer(
                        payload,
                        operator=operator,
                        operator_name=operator_name,
                    )
                ),
                201,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.get("/api/customers/<int:customer_id>")
    @_login_required
    def get_customer(customer_id: int):
        try:
            return jsonify(CUSTOMER_STORE.get_customer(customer_id))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404

    @application.patch("/api/customers/<int:customer_id>")
    @_login_required
    def update_customer(customer_id: int):
        _check_csrf()
        payload = request.get_json(silent=True) or {}
        operator, operator_name = _customer_operator()
        try:
            return jsonify(
                CUSTOMER_STORE.update_customer(
                    customer_id,
                    payload,
                    operator=operator,
                    operator_name=operator_name,
                )
            )
        except ValueError as exc:
            message = str(exc)
            status = 409 if "其他人修改" in message else 400
            return jsonify({"error": message}), status

    @application.delete("/api/customers/<int:customer_id>")
    @_admin_required
    def delete_customer(customer_id: int):
        _check_csrf()
        operator, operator_name = _customer_operator()
        try:
            CUSTOMER_STORE.delete_customer(
                customer_id,
                operator=operator,
                operator_name=operator_name,
            )
            return jsonify({"message": "客户档案已删除"})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.get("/api/customers/<int:customer_id>/logs")
    @_login_required
    def customer_logs(customer_id: int):
        try:
            return jsonify({"items": CUSTOMER_STORE.list_logs(customer_id)})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404

    @application.get("/api/customers/import-template")
    @_login_required
    def customer_import_template():
        return send_file(
            _customer_workbook_response(),
            as_attachment=True,
            download_name="客户档案导入模板.xlsx",
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    @application.post("/api/customers/import")
    @_admin_required
    def import_customers():
        _check_csrf()
        upload = request.files.get("file")
        if not upload or not upload.filename:
            return jsonify({"error": "请选择需要导入的 Excel 文件"}), 400
        if Path(upload.filename).suffix.lower() != ".xlsx":
            return jsonify({"error": "仅支持 .xlsx 格式的 Excel 文件"}), 400
        try:
            rows = _parse_customer_workbook(io.BytesIO(upload.read()))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        if not rows:
            return jsonify({"error": "Excel 中没有可导入的客户数据"}), 400

        operator, operator_name = _customer_operator()
        success_count = 0
        errors = []
        seen_codes: set[str] = set()
        for row in rows:
            raw_code = str(row["payload"].get("terminal_code") or "")
            if raw_code in seen_codes:
                errors.append(
                    {
                        "row_number": row["row_number"],
                        "data": row["raw"],
                        "error": f"Excel 内终端编码 {raw_code} 重复",
                    }
                )
                continue
            if raw_code:
                seen_codes.add(raw_code)
            try:
                CUSTOMER_STORE.create_customer(
                    row["payload"],
                    operator=operator,
                    operator_name=operator_name,
                    source="batch",
                )
                success_count += 1
            except ValueError as exc:
                errors.append(
                    {
                        "row_number": row["row_number"],
                        "data": row["raw"],
                        "error": str(exc),
                    }
                )
        import_id = CUSTOMER_STORE.record_import(
            operator=operator,
            operator_name=operator_name,
            filename=Path(upload.filename).name,
            total_count=len(rows),
            success_count=success_count,
            errors=errors,
        )
        return jsonify(
            {
                "id": import_id,
                "total_count": len(rows),
                "success_count": success_count,
                "failed_count": len(errors),
                "errors": errors[:100],
                "error_report_url": (
                    f"/api/customers/imports/{import_id}/errors"
                    if errors
                    else ""
                ),
            }
        )

    @application.get("/api/customers/imports/<import_id>/errors")
    @_admin_required
    def customer_import_errors(import_id: str):
        try:
            import_job = CUSTOMER_STORE.get_import(import_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404
        if not import_job["errors"]:
            return jsonify({"error": "本次导入没有失败数据"}), 404
        return send_file(
            _customer_workbook_response(errors=import_job["errors"]),
            as_attachment=True,
            download_name="客户档案导入失败明细.xlsx",
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    @application.get("/api/snow-outbound/options")
    @_login_required
    def snow_outbound_options():
        return jsonify(
            {
                "tags": list(POLICY_TAGS),
                "fields": [
                    {
                        "value": field,
                        "label": config["label"],
                        "operators": sorted(config["operators"]),
                    }
                    for field, config in RULE_FIELDS.items()
                ],
                "operators": RULE_OPERATOR_LABELS,
                "gift_types": list(POLICY_GIFT_TYPES),
                "months": SNOW_OUTBOUND_STORE.list_months(),
                "current_month": datetime.now().strftime("%Y-%m"),
                "years": list(range(2026, 2036)),
            }
        )

    @application.get("/api/products")
    @_login_required
    def list_products():
        try:
            return jsonify(
                PRODUCT_STORE.list_products(
                    name=request.args.get("name", ""),
                    product_code=request.args.get("product_code", ""),
                    housekeeper_code=request.args.get("housekeeper_code", ""),
                    status=request.args.get("status", ""),
                    inventory_sort=request.args.get("inventory_sort", ""),
                    summary_month=request.args.get("summary_month", ""),
                    page=int(request.args.get("page", "1")),
                    page_size=int(request.args.get("page_size", "20")),
                )
            )
        except (TypeError, ValueError) as exc:
            return jsonify({"error": str(exc) or "分页参数不正确"}), 400

    @application.get("/api/products/options")
    @_login_required
    def product_options():
        return jsonify({"items": PRODUCT_STORE.normal_product_options()})

    @application.post("/api/products")
    @_login_required
    def create_product():
        _check_csrf()
        operator, operator_name = _customer_operator()
        try:
            return (
                jsonify(
                    PRODUCT_STORE.create_product(
                        request.get_json(silent=True) or {},
                        operator=operator,
                        operator_name=operator_name,
                    )
                ),
                201,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.patch("/api/products/<int:product_id>")
    @_login_required
    def update_product(product_id: int):
        _check_csrf()
        operator, operator_name = _customer_operator()
        try:
            return jsonify(
                PRODUCT_STORE.update_product(
                    product_id,
                    request.get_json(silent=True) or {},
                    operator=operator,
                    operator_name=operator_name,
                )
            )
        except ValueError as exc:
            message = str(exc)
            return jsonify({"error": message}), (
                409 if "其他人修改" in message else 400
            )

    @application.delete("/api/products/<int:product_id>")
    @_login_required
    def delete_product(product_id: int):
        _check_csrf()
        operator, operator_name = _customer_operator()
        try:
            PRODUCT_STORE.delete_product(
                product_id,
                operator=operator,
                operator_name=operator_name,
            )
            return jsonify({"message": "产品档案已删除"})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.post("/api/products/import/preview")
    @_login_required
    def preview_product_import():
        _check_csrf()
        upload = request.files.get("file")
        if not upload or not upload.filename:
            return jsonify({"error": "请选择雪花库存Excel文件"}), 400
        if Path(upload.filename).suffix.lower() != ".xlsx":
            return jsonify({"error": "仅支持.xlsx格式的Excel文件"}), 400
        operator, operator_name = _customer_operator()
        try:
            rows = parse_stock_workbook(io.BytesIO(upload.read()))
            if not rows:
                return jsonify({"error": "Excel中没有可导入的商品数据"}), 400
            return jsonify(
                PRODUCT_STORE.create_import_preview(
                    filename=Path(upload.filename).name,
                    operator=operator,
                    operator_name=operator_name,
                    rows=rows,
                )
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.post("/api/products/import")
    @_login_required
    def commit_product_import():
        _check_csrf()
        preview_id = str(
            (request.get_json(silent=True) or {}).get("preview_id") or ""
        )
        if not preview_id:
            return jsonify({"error": "缺少导入预览标识"}), 400
        operator, operator_name = _customer_operator()
        try:
            return jsonify(
                PRODUCT_STORE.commit_import_preview(
                    preview_id,
                    operator=operator,
                    operator_name=operator_name,
                )
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.get("/api/customers/policy-options")
    @_login_required
    def customer_policy_options():
        month = request.args.get("month", "")
        try:
            return jsonify(
                {
                    "month": month,
                    "items": SNOW_OUTBOUND_STORE.list_policy_tags(month),
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.get("/api/snow-outbound/policies")
    @_login_required
    def list_snow_policies():
        try:
            result = SNOW_OUTBOUND_STORE.list_policies(
                year=request.args.get("year", ""),
                month=request.args.get("month", ""),
                outbound_code=request.args.get("outbound_code", ""),
                name=request.args.get("name", ""),
                enabled=request.args.get("enabled", ""),
                sort_by=request.args.get("sort_by", ""),
                sort_order=request.args.get("sort_order", "desc"),
                page=int(request.args.get("page", "1")),
                page_size=int(request.args.get("page_size", "20")),
            )
            policy_ids = [item["id"] for item in result["items"]]
            archived_by_policy = IMAGE_LIBRARY.archived_terminal_codes_by_policy(
                policy_ids
            )
            shipped_by_policy = (
                SNOW_OUTBOUND_STORE.shipped_terminal_codes_by_policy(policy_ids)
            )
            for item in result["items"]:
                archived_codes = archived_by_policy.get(item["id"], set())
                shipped_codes = shipped_by_policy.get(item["id"], set())
                item["photographed_count"] = len(archived_codes)
                item["pending_outbound_count"] = len(
                    archived_codes - shipped_codes
                )
            return jsonify(result)
        except (TypeError, ValueError) as exc:
            return jsonify({"error": str(exc)}), 400

    @application.get("/api/snow-outbound/policies/options")
    @_login_required
    def snow_policy_options():
        try:
            return jsonify(
                {
                    "items": SNOW_OUTBOUND_STORE.policy_options(
                        year=int(request.args.get("year", "2026")),
                        month=int(request.args.get("month", "0")),
                        exclude_id=request.args.get("exclude_id", ""),
                    )
                }
            )
        except (TypeError, ValueError) as exc:
            return jsonify({"error": str(exc)}), 400

    @application.get("/api/snow-outbound/policies/<policy_id>/alerts")
    @_login_required
    def snow_policy_alerts(policy_id: str):
        try:
            items = SNOW_OUTBOUND_STORE.policy_alert_terminals(policy_id)
            return jsonify(
                {
                    "items": items,
                    "total": len(items),
                    "policy_id": policy_id,
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404

    @application.post("/api/snow-outbound/policies/<policy_id>/export")
    @_login_required
    def export_snow_policy(policy_id: str):
        _check_csrf()
        _operator, operator_name = _customer_operator()
        try:
            policy, rows = SNOW_OUTBOUND_STORE.policy_export_rows(policy_id)
            if not rows:
                return (
                    jsonify(
                        {
                            "error": (
                                "当前标签没有同时满足出库编码和售卖类型的"
                                "有效出库数据"
                            )
                        }
                    ),
                    400,
                )
            output = build_policy_reimbursement_workbook(
                policy,
                rows,
                operator_name=operator_name,
                distributor_name=POLICY_EXPORT_DISTRIBUTOR,
            )
            filename = (
                f"{int(policy['year']):04d}年{int(policy['month'])}月-"
                f"{_safe_archive_part(policy['name'])}-核销明细-"
                f"{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.get("/api/snow-outbound/policies/<policy_id>/pending-outbound")
    @_login_required
    def pending_outbound_terminals(policy_id: str):
        policy = SNOW_OUTBOUND_STORE.get_policy(policy_id)
        if not policy:
            return jsonify({"error": "雪花政策标签不存在"}), 404
        shipped_codes = {
            item["terminal_code"]
            for item in SNOW_OUTBOUND_STORE.shipped_terminals(policy_id)
        }
        pending = [
            terminal
            for terminal in IMAGE_LIBRARY.archived_terminals(policy_id)
            if terminal["terminal_code"] not in shipped_codes
        ]
        return jsonify(
            {
                "items": pending,
                "total": len(pending),
                "policy_id": policy_id,
                "policy_name": policy["display_name"],
            }
        )

    @application.get("/api/snow-outbound/policies/<policy_id>/shipped-terminals")
    @_login_required
    def shipped_policy_terminals(policy_id: str):
        policy = SNOW_OUTBOUND_STORE.get_policy(policy_id)
        if not policy:
            return jsonify({"error": "雪花政策标签不存在"}), 404
        items = SNOW_OUTBOUND_STORE.shipped_terminals(policy_id)
        return jsonify(
            {
                "items": items,
                "total": len(items),
                "policy_id": policy_id,
                "policy_name": policy["display_name"],
            }
        )

    @application.get("/api/snow-outbound/policies/<policy_id>/reversed-terminals")
    @_login_required
    def reversed_policy_terminals(policy_id: str):
        try:
            policy = SNOW_OUTBOUND_STORE.get_policy(policy_id)
            if not policy:
                return jsonify({"error": "雪花政策标签不存在"}), 404
            items = SNOW_OUTBOUND_STORE.reversed_terminals(policy_id)
            return jsonify(
                {
                    "items": items,
                    "total": len(items),
                    "policy_id": policy_id,
                    "policy_name": policy["display_name"],
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404

    @application.get(
        "/api/snow-outbound/policies/<policy_id>/photographed-terminals"
    )
    @_login_required
    def photographed_policy_terminals(policy_id: str):
        policy = SNOW_OUTBOUND_STORE.get_policy(policy_id)
        if not policy:
            return jsonify({"error": "雪花政策标签不存在"}), 404
        items = IMAGE_LIBRARY.archived_terminals(policy_id)
        return jsonify(
            {
                "items": items,
                "total": len(items),
                "policy_id": policy_id,
                "policy_name": policy["display_name"],
            }
        )

    @application.post("/api/snow-outbound/policies")
    @_login_required
    def create_snow_policy():
        _check_csrf()
        operator, operator_name = _customer_operator()
        try:
            return (
                jsonify(
                    SNOW_OUTBOUND_STORE.create_policy(
                        request.get_json(silent=True) or {},
                        operator=operator,
                        operator_name=operator_name,
                    )
                ),
                201,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.patch("/api/snow-outbound/policies/<policy_id>")
    @_login_required
    def update_snow_policy(policy_id: str):
        _check_csrf()
        operator, operator_name = _customer_operator()
        try:
            return jsonify(
                SNOW_OUTBOUND_STORE.update_policy(
                    policy_id,
                    request.get_json(silent=True) or {},
                    operator=operator,
                    operator_name=operator_name,
                )
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.post("/api/snow-outbound/policies/<policy_id>/status")
    @_login_required
    def set_snow_policy_status(policy_id: str):
        _check_csrf()
        payload = request.get_json(silent=True) or {}
        if not isinstance(payload.get("enabled"), bool):
            return jsonify({"error": "启用状态格式不正确"}), 400
        operator, operator_name = _customer_operator()
        try:
            return jsonify(
                SNOW_OUTBOUND_STORE.set_policy_enabled(
                    policy_id,
                    payload["enabled"],
                    operator=operator,
                    operator_name=operator_name,
                )
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.delete("/api/snow-outbound/policies/<policy_id>")
    @_admin_required
    def delete_snow_policy(policy_id: str):
        _check_csrf()
        operator, operator_name = _customer_operator()
        try:
            SNOW_OUTBOUND_STORE.delete_policy(
                policy_id,
                operator=operator,
                operator_name=operator_name,
            )
            return jsonify({"message": "雪花出库政策已删除"})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.get("/api/snow-outbound/policies/<policy_id>/logs")
    @_login_required
    def snow_policy_logs(policy_id: str):
        return jsonify(
            {"items": SNOW_OUTBOUND_STORE.list_policy_logs(policy_id)}
        )

    @application.get("/api/snow-outbound/templates")
    @_login_required
    def list_snow_rule_templates():
        return jsonify({"items": SNOW_OUTBOUND_STORE.list_templates()})

    @application.post("/api/snow-outbound/templates")
    @_login_required
    def create_snow_rule_template():
        _check_csrf()
        payload = request.get_json(silent=True) or {}
        operator, operator_name = _customer_operator()
        try:
            return (
                jsonify(
                    SNOW_OUTBOUND_STORE.save_template(
                        name=str(payload.get("name") or ""),
                        rules=payload.get("rules"),
                        is_default=bool(payload.get("is_default")),
                        operator=operator,
                        operator_name=operator_name,
                        is_admin=_is_admin(),
                    )
                ),
                201,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.patch("/api/snow-outbound/templates/<template_id>")
    @_login_required
    def update_snow_rule_template(template_id: str):
        _check_csrf()
        payload = request.get_json(silent=True) or {}
        operator, operator_name = _customer_operator()
        try:
            return jsonify(
                SNOW_OUTBOUND_STORE.save_template(
                    template_id=template_id,
                    name=str(payload.get("name") or ""),
                    rules=payload.get("rules"),
                    is_default=bool(payload.get("is_default")),
                    operator=operator,
                    operator_name=operator_name,
                    is_admin=_is_admin(),
                )
            )
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.delete("/api/snow-outbound/templates/<template_id>")
    @_login_required
    def delete_snow_rule_template(template_id: str):
        _check_csrf()
        operator, _operator_name = _customer_operator()
        try:
            SNOW_OUTBOUND_STORE.delete_template(
                template_id,
                operator=operator,
                is_admin=_is_admin(),
            )
            return jsonify({"message": "规则模板已删除"})
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.post("/api/snow-outbound/preview")
    @_login_required
    def preview_snow_outbound():
        _check_csrf()
        upload = request.files.get("file")
        if not upload or not upload.filename:
            return jsonify({"error": "请选择雪花出库Excel文件"}), 400
        if Path(upload.filename).suffix.lower() != ".xlsx":
            return jsonify({"error": "仅支持.xlsx格式的Excel文件"}), 400
        operator, operator_name = _customer_operator()
        try:
            # Werkzeug在部分Python版本中提供的临时上传流不完整实现
            # openpyxl所需的seekable接口；请求体已有全局大小限制，可安全转为内存流。
            rows = parse_outbound_workbook(io.BytesIO(upload.read()))
            update_policy = str(
                request.form.get("update_policy", "true")
            ).strip().lower() in {"1", "true", "yes", "on"}
            return jsonify(
                SNOW_OUTBOUND_STORE.create_preview(
                    filename=Path(upload.filename).name,
                    operator=operator,
                    operator_name=operator_name,
                    rows=rows,
                    update_policy=update_policy,
                )
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.post("/api/snow-outbound/import")
    @_login_required
    def commit_snow_outbound():
        _check_csrf()
        payload = request.get_json(silent=True) or {}
        preview_id = str(payload.get("preview_id") or "")
        if not preview_id:
            return jsonify({"error": "缺少导入预览标识"}), 400
        operator, operator_name = _customer_operator()
        try:
            return jsonify(
                SNOW_OUTBOUND_STORE.commit_preview(
                    preview_id,
                    operator=operator,
                    operator_name=operator_name,
                    is_admin=_is_admin(),
                )
            )
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @application.get("/api/results")
    @_login_required
    def saved_results():
        return jsonify(_load_saved_results())

    @application.get("/api/distributions")
    @_login_required
    def distribution_summary():
        items = DISTRIBUTION_STORE.summaries()
        return jsonify(
            {
                "items": items,
                "totals": {
                    "business_count": len(
                        [
                            item
                            for item in items
                            if item["business"] != "待识别"
                        ]
                    ),
                    "quantity": sum(item["quantity"] for item in items),
                    "distributed_count": sum(
                        item["distributed_count"] for item in items
                    ),
                    "pending_download_count": sum(
                        item["pending_download_count"] for item in items
                    ),
                },
            }
        )

    @application.delete("/api/distributions")
    @_login_required
    def clear_distributions():
        _check_csrf()
        deleted_count = DISTRIBUTION_STORE.clear_all()
        return jsonify(
            {
                "deleted_count": deleted_count,
                "message": f"已清空 {deleted_count} 条分发记录",
            }
        )

    @application.post("/api/distributions/<path:business>/archive")
    @_login_required
    def distribution_archive(business: str):
        _check_csrf()
        try:
            return jsonify(_create_distribution_archive(business))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception:
            application.logger.exception("生成业务分发压缩包失败")
            return jsonify({"error": "生成压缩包失败，请联系管理员"}), 500

    @application.get("/api/image-library")
    @_login_required
    def image_library():
        month = request.args.get("month", "").strip()
        businesses = _parse_string_values(
            request.args.getlist("business") or request.args.get("business", "")
        )
        customer_name = request.args.get("customer_name", "").strip()
        policy_ids = _parse_string_values(
            request.args.getlist("policy_id") or request.args.get("policy_id", "")
        )
        policy_match = request.args.get("policy_match", "include").strip()
        if policy_match not in {"include", "exclude"}:
            return jsonify({"error": "终端政策条件必须为包含或不包含"}), 400
        archive_policy_ids = _parse_string_values(
            request.args.getlist("archive_policy_id")
            or request.args.get("archive_policy_id", "")
        )
        archive_policy_match = request.args.get(
            "archive_policy_match", "archived"
        ).strip()
        if archive_policy_match not in {"archived", "unarchived"}:
            return jsonify({"error": "归档条件必须为已归档或未归档"}), 400
        fields = _parse_field_lines(request.args.get("fields", ""))
        try:
            page, page_size = _parse_pagination(
                request.args.get("page"),
                request.args.get("page_size"),
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        try:
            policy_options = (
                SNOW_OUTBOUND_STORE.active_policy_options(month) if month else []
            )
        except ValueError:
            policy_options = []
        archive_policy_options = _archive_policy_options(month)
        valid_archive_policy_ids = {
            item["id"] for item in archive_policy_options
        }
        if any(
            policy_id not in valid_archive_policy_ids
            for policy_id in archive_policy_ids
        ):
            return jsonify({"error": "所选归档标签在当前月份不存在归档记录"}), 400
        terminal_codes = None
        if policy_ids:
            valid_policy_ids = {item["id"] for item in policy_options}
            if any(policy_id not in valid_policy_ids for policy_id in policy_ids):
                return jsonify({"error": "所选政策标签不属于当前月份或未启用"}), 400
            terminal_codes = sorted(
                {
                    item["terminal_code"]
                    for policy_id in policy_ids
                    for item in SNOW_OUTBOUND_STORE.shipped_terminals(policy_id)
                }
            )
        result = IMAGE_LIBRARY.query(
            fields=fields,
            terminal_codes=terminal_codes,
            terminal_code_match=policy_match,
            archive_policy_ids=archive_policy_ids,
            archive_policy_match=archive_policy_match,
            month=month,
            businesses=businesses,
            customer_name=customer_name,
            page=page,
            page_size=page_size,
        )
        _attach_image_archive_tags(result)
        tag_map = SNOW_OUTBOUND_STORE.policy_tags_for_terminals(
            month,
            [item["field"] for item in result["items"]],
        )
        for item in result["items"]:
            item["policy_tags"] = tag_map.get(item["field"], [])
        return jsonify(
            {
                **result,
                "months": IMAGE_LIBRARY.months(),
                "businesses": IMAGE_LIBRARY.businesses(),
                "customer_names": IMAGE_LIBRARY.customer_names(),
                "policy_options": policy_options,
                "archive_policy_options": archive_policy_options,
            }
        )

    @application.post("/api/image-library/search")
    @_login_required
    def image_library_search():
        payload = request.get_json(silent=True) or {}
        month = str(payload.get("month") or "").strip()
        businesses = _parse_string_values(
            payload.get("businesses", payload.get("business"))
        )
        policy_ids = _parse_string_values(
            payload.get("policy_ids", payload.get("policy_id"))
        )
        policy_match = str(payload.get("policy_match") or "include").strip()
        if policy_match not in {"include", "exclude"}:
            return jsonify({"error": "终端政策条件必须为包含或不包含"}), 400
        archive_policy_ids = _parse_string_values(
            payload.get("archive_policy_ids", payload.get("archive_policy_id"))
        )
        archive_policy_match = str(
            payload.get("archive_policy_match") or "archived"
        ).strip()
        if archive_policy_match not in {"archived", "unarchived"}:
            return jsonify({"error": "归档条件必须为已归档或未归档"}), 400
        raw_fields = payload.get("fields", "")
        if isinstance(raw_fields, list):
            fields = [
                str(item).strip()
                for item in raw_fields
                if str(item).strip()
            ]
        else:
            fields = _parse_field_lines(str(raw_fields))
        try:
            page, page_size = _parse_pagination(
                payload.get("page"),
                payload.get("page_size"),
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        try:
            policy_options = (
                SNOW_OUTBOUND_STORE.active_policy_options(month) if month else []
            )
        except ValueError:
            policy_options = []
        archive_policy_options = _archive_policy_options(month)
        valid_archive_policy_ids = {
            item["id"] for item in archive_policy_options
        }
        if any(
            policy_id not in valid_archive_policy_ids
            for policy_id in archive_policy_ids
        ):
            return jsonify({"error": "所选归档标签在当前月份不存在归档记录"}), 400
        terminal_codes = None
        if policy_ids:
            valid_policy_ids = {item["id"] for item in policy_options}
            if any(policy_id not in valid_policy_ids for policy_id in policy_ids):
                return jsonify({"error": "所选政策标签不属于当前月份或未启用"}), 400
            terminal_codes = sorted(
                {
                    item["terminal_code"]
                    for policy_id in policy_ids
                    for item in SNOW_OUTBOUND_STORE.shipped_terminals(policy_id)
                }
            )
        result = IMAGE_LIBRARY.query(
            fields=fields,
            terminal_codes=terminal_codes,
            terminal_code_match=policy_match,
            archive_policy_ids=archive_policy_ids,
            archive_policy_match=archive_policy_match,
            month=month,
            businesses=businesses,
            customer_name=str(payload.get("customer_name") or "").strip(),
            page=page,
            page_size=page_size,
        )
        _attach_image_archive_tags(result)
        tag_map = SNOW_OUTBOUND_STORE.policy_tags_for_terminals(
            month,
            [item["field"] for item in result["items"]],
        )
        for item in result["items"]:
            item["policy_tags"] = tag_map.get(item["field"], [])
        return jsonify(
            {
                **result,
                "months": IMAGE_LIBRARY.months(),
                "businesses": IMAGE_LIBRARY.businesses(),
                "customer_names": IMAGE_LIBRARY.customer_names(),
                "policy_options": policy_options,
                "archive_policy_options": archive_policy_options,
            }
        )

    @application.get("/api/image-library/images/<image_id>/thumbnail")
    @_login_required
    def image_library_thumbnail(image_id: str):
        thumbnail = IMAGE_LIBRARY.thumbnail_for(image_id)
        if thumbnail is None:
            abort(404)
        return _serve_output_file(thumbnail, cache=True)

    @application.get("/api/photo-archive/options")
    @_login_required
    def photo_archive_options():
        month = request.args.get("month", "").strip()
        try:
            items = SNOW_OUTBOUND_STORE.active_policy_options(
                month,
                requires_photo_only=True,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        return jsonify({"items": items, "month": month})

    @application.post("/api/photo-archive")
    @_login_required
    def archive_library_images():
        _check_csrf()
        payload = request.get_json(silent=True) or {}
        image_ids = [
            str(item).strip()
            for item in payload.get("image_ids") or []
            if str(item).strip()
        ]
        policy_id = str(payload.get("policy_id") or "").strip()
        month = str(payload.get("month") or "").strip()
        policy = SNOW_OUTBOUND_STORE.get_policy(policy_id)
        if not policy:
            return jsonify({"error": "雪花政策标签不存在"}), 404
        policy_month = f"{int(policy['year']):04d}-{int(policy['month']):02d}"
        if not policy["enabled"]:
            return jsonify({"error": "该雪花政策标签未启用"}), 400
        if not policy["requires_photo"]:
            return jsonify({"error": "该雪花政策标签定义为无需拍照"}), 400
        if policy_month != month:
            return jsonify({"error": "所选政策标签与照片月份不一致"}), 400
        actor, actor_name = _customer_operator()
        try:
            result = IMAGE_LIBRARY.archive_images(
                image_ids,
                policy_id=policy_id,
                month=month,
                actor=actor,
                actor_name=actor_name,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        return jsonify(
            {
                **result,
                "policy_id": policy_id,
                "policy_name": policy["display_name"],
            }
        )

    @application.delete(
        "/api/photo-archive/images/<image_id>/policies/<policy_id>"
    )
    @_login_required
    def remove_library_image_archive_tag(image_id: str, policy_id: str):
        _check_csrf()
        policy = SNOW_OUTBOUND_STORE.get_policy(policy_id, include_deleted=True)
        if not policy:
            return jsonify({"error": "雪花政策标签不存在"}), 404
        actor, actor_name = _customer_operator()
        result = IMAGE_LIBRARY.remove_archive_tag(
            image_id,
            policy_id=policy_id,
            actor=actor,
            actor_name=actor_name,
        )
        if result is None:
            return jsonify({"error": "该照片未标注此政策标签"}), 404
        return jsonify(
            {
                **result,
                "policy_name": policy["display_name"],
            }
        )

    @application.get("/api/photo-archive/policies")
    @_login_required
    def photo_archive_policies():
        month = request.args.get("month", "").strip()
        match = re.fullmatch(r"(\d{4})-(\d{2})", month)
        if not match:
            return jsonify({"error": "请选择照片档案月份"}), 400
        try:
            page, page_size = _parse_pagination(
                request.args.get("page"),
                request.args.get("page_size"),
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        policies = SNOW_OUTBOUND_STORE.list_policies(
            year=match.group(1),
            month=str(int(match.group(2))),
            requires_photo=True,
            page=page,
            page_size=page_size,
        )
        policy_ids = [item["id"] for item in policies["items"]]
        archive_stats = IMAGE_LIBRARY.policy_archive_stats(policy_ids)
        latest_logs = IMAGE_LIBRARY.latest_photo_archive_logs(policy_ids)
        items = []
        for policy in policies["items"]:
            stats = archive_stats.get(
                policy["id"],
                {"photo_count": 0, "photographed_count": 0},
            )
            shipped_terminals = SNOW_OUTBOUND_STORE.shipped_terminals(policy["id"])
            shipped_codes = {
                terminal["terminal_code"] for terminal in shipped_terminals
            }
            photographed_codes = IMAGE_LIBRARY.archived_terminal_codes(policy["id"])
            latest = latest_logs.get(policy["id"])
            items.append(
                {
                    "policy_id": policy["id"],
                    "display_name": policy["display_name"],
                    "color": policy["color"],
                    "year": policy["year"],
                    "month": policy["month"],
                    "enabled": policy["enabled"],
                    "shipped_count": len(shipped_codes),
                    "photographed_count": stats["photographed_count"],
                    "missing_count": len(shipped_codes - photographed_codes),
                    "photo_count": stats["photo_count"],
                    "latest_operation": (
                        {
                            "operated_at": latest["operated_at"],
                            "actor_name": latest["actor_name"] or latest["actor"],
                            "action_type": latest["action_type"],
                            "action_label": (
                                "归档"
                                if latest["action_type"] == "archive"
                                else (
                                    "删除标签"
                                    if latest["action_type"] == "unarchive"
                                    else "导出"
                                )
                            ),
                            "photo_count": int(latest["photo_count"] or 0),
                            "terminal_count": int(latest["terminal_count"] or 0),
                            "skipped_count": int(latest["skipped_count"] or 0),
                        }
                        if latest
                        else None
                    ),
                }
            )
        return jsonify(
            {
                "items": items,
                "total": policies["total"],
                "page": policies["page"],
                "page_size": policies["page_size"],
                "month": month,
                "months": SNOW_OUTBOUND_STORE.policy_months(
                    requires_photo_only=True
                ),
            }
        )

    @application.get("/api/photo-archive/policies/<policy_id>/missing")
    @_login_required
    def photo_archive_missing_terminals(policy_id: str):
        policy = SNOW_OUTBOUND_STORE.get_policy(policy_id)
        if not policy:
            return jsonify({"error": "雪花政策标签不存在"}), 404
        photographed_codes = IMAGE_LIBRARY.archived_terminal_codes(policy_id)
        missing = [
            terminal
            for terminal in SNOW_OUTBOUND_STORE.shipped_terminals(policy_id)
            if terminal["terminal_code"] not in photographed_codes
        ]
        return jsonify(
            {
                "items": missing,
                "total": len(missing),
                "policy_id": policy_id,
                "policy_name": policy["display_name"],
            }
        )

    @application.post("/api/photo-archive/policies/<policy_id>/export")
    @_login_required
    def export_photo_archive(policy_id: str):
        _check_csrf()
        policy = SNOW_OUTBOUND_STORE.get_policy(policy_id)
        if not policy:
            return jsonify({"error": "雪花政策标签不存在"}), 404
        try:
            archive_path, archive_name, photo_count, terminal_count = (
                _create_photo_archive(policy)
            )
            actor, actor_name = _customer_operator()
            IMAGE_LIBRARY.record_photo_archive_export(
                policy_id,
                actor=actor,
                actor_name=actor_name,
                photo_count=photo_count,
                terminal_count=terminal_count,
            )
            response = send_file(
                archive_path,
                as_attachment=True,
                download_name=archive_name,
                mimetype="application/zip",
            )
            response.call_on_close(lambda: archive_path.unlink(missing_ok=True))
            return response
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception:
            application.logger.exception("照片档案导出失败")
            return jsonify({"error": "照片档案导出失败，请联系管理员"}), 500

    @application.get("/api/extraction-records")
    @_login_required
    def list_extraction_records():
        return jsonify({"items": EXTRACTION_RECORD_STORE.list_records()})

    @application.get("/output/<path:relative_path>")
    @_login_required
    def output_file(relative_path: str):
        first_part = Path(relative_path).parts[:1]
        if first_part in {
            ("_system",),
            ("_image_exports",),
            ("_photo_archive_exports",),
            ("_image_thumbnails",),
        }:
            abort(404)
        return _serve_output_file(
            OUTPUT_ROOT / relative_path,
            cache=first_part == ("_image_library",),
        )

    @application.post("/api/extract")
    @_login_required
    def extract():
        _check_csrf()
        _check_rate_limit()
        payload = request.get_json(silent=True) or {}
        url = str(payload.get("url", "")).strip()
        if not url:
            return jsonify({"error": "请粘贴 CRM 拜访详情链接"}), 400
        if len(url) > 4096:
            return jsonify({"error": "链接过长"}), 400

        record_id = secrets.token_hex(16)
        owner_username, owner_display_name = _record_owner()
        EXTRACTION_RECORD_STORE.start_record(
            record_id=record_id,
            owner_username=owner_username,
            owner_display_name=owner_display_name,
            method="single_link",
        )
        try:
            with EXTRACT_LOCK:
                result = extract_images(url, OUTPUT_ROOT)
                added_image_count = IMAGE_LIBRARY.add_result(result, source_url=url)
        except (ValueError, CrmApiError) as exc:
            _complete_extraction_record(
                application,
                record_id,
                status="failed",
                error_information=str(exc),
            )
            return jsonify({"error": str(exc)}), 400
        except Exception:
            application.logger.exception("提取图片失败")
            _complete_extraction_record(
                application,
                record_id,
                status="failed",
                error_information="提取失败，请联系管理员查看服务日志",
            )
            return jsonify({"error": "提取失败，请联系管理员查看服务日志"}), 500
        terminal_fields: set[str] = set()
        for image in result.images:
            try:
                terminal_fields.add(photoid_name_field(image.photoid))
            except ValueError:
                continue
        _complete_extraction_record(
            application,
            record_id,
            status="success",
            image_count=added_image_count,
            terminal_count=len(terminal_fields) if added_image_count else 0,
        )
        response = _serialize_result(result)
        response["record_id"] = record_id
        response["added_image_count"] = added_image_count
        return jsonify(response)

    @application.post("/api/batch-extract")
    @_login_required
    def batch_extract():
        _check_csrf()
        _check_rate_limit()
        upload = request.files.get("file")
        if upload is None or not upload.filename:
            return jsonify({"error": "请选择 Excel 文件"}), 400
        if Path(upload.filename).suffix.lower() != ".xlsx":
            return jsonify({"error": "仅支持 .xlsx 格式的 Excel 文件"}), 400

        job_id = secrets.token_urlsafe(18)
        owner_username, owner_display_name = _record_owner()
        EXTRACTION_RECORD_STORE.start_record(
            record_id=job_id,
            owner_username=owner_username,
            owner_display_name=owner_display_name,
            method="batch",
        )
        try:
            links, input_stats = _parse_excel_links(upload.stream)
        except ValueError as exc:
            _complete_extraction_record(
                application,
                job_id,
                status="failed",
                error_information=str(exc),
            )
            return jsonify({"error": str(exc)}), 400
        except Exception:
            application.logger.exception("批量提取图片失败")
            _complete_extraction_record(
                application,
                job_id,
                status="failed",
                error_information="批量提取失败，请联系管理员查看服务日志",
            )
            return jsonify({"error": "批量提取失败，请联系管理员查看服务日志"}), 500

        _prune_batch_jobs()
        now = time.time()
        chunk_count = max(1, (len(links) + BATCH_CHUNK_SIZE - 1) // BATCH_CHUNK_SIZE)
        _register_batch_job(
            job_id,
            {
                "owner": _current_user(),
                "record_id": job_id,
                "status": "queued",
                "processed": 0,
                "total": len(links),
                "current_row": None,
                "succeeded": 0,
                "failed": 0,
                "image_count": 0,
                "retry_count": 0,
                "chunk_index": 1,
                "chunk_count": chunk_count,
                "links": links,
                "completed_records": [],
                "errors": [],
                **input_stats,
                "created_at": now,
                "updated_at": now,
            },
        )
        BATCH_QUEUE.put(job_id)
        return jsonify(
            {
                "job_id": job_id,
                "status": "queued",
                "total": len(links),
                "retry_count": 0,
                "chunk_index": 1,
                "chunk_count": chunk_count,
                **input_stats,
            }
        ), 202

    @application.get("/api/batch-extract/<job_id>")
    @_login_required
    def batch_extract_status(job_id: str):
        with BATCH_JOBS_LOCK:
            job = BATCH_JOBS.get(job_id)
            if job is None or job.get("owner") != _current_user():
                return jsonify({"error": "批量任务不存在或已过期"}), 404
            payload = _public_batch_job(job)
        return jsonify(payload)

    @application.errorhandler(413)
    def upload_too_large(_error):
        size_mb = MAX_UPLOAD_BYTES / 1024 / 1024
        return jsonify({"error": f"上传文件不能超过 {size_mb:g} MB"}), 413

    @application.errorhandler(403)
    @application.errorhandler(429)
    def handled_error(error):
        return jsonify({"error": error.description}), error.code

    _start_batch_worker(application)
    return application


app = create_app()


if __name__ == "__main__":
    app.run(
        host=os.environ.get("INFOLENS_HOST", "0.0.0.0"),
        port=int(os.environ.get("INFOLENS_PORT", "8765")),
        debug=False,
        threaded=True,
    )
